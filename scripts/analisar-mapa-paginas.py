"""Analisa toda mapa-paginas e reporta o que alterar linha por linha."""
import openpyxl
import re
from collections import Counter

wb = openpyxl.load_workbook(r'C:\Users\danie\Downloads\planilha_geral_gravity_COMPLETA.xlsx', data_only=True)
ws = wb['6. mapa-paginas']

# Cols: A=Local, B=URL atual, C=URL DDD, D=Titulo atual, E=Titulo DDD,
# F=arquivo, G=arquivo DDD, H=componente, I=componente DDD, J=produto, K=area,
# L=tipo view, M=breadcrumb, N=auth, O=patente, P=rotas api, Q=models, R=modais,
# S=mobile, T=arquivo full, U=status DDD, V=obs

# Identifica tipo real por nome do arquivo
def tipo_real(fname, comp):
    if not fname: return 'Outro'
    base = fname.replace('.tsx','').replace('.jsx','')
    if '.stories' in fname or fname.endswith('.test.tsx'): return 'Teste/Story'
    if 'Harness' in base: return 'TestHarness'
    if base.startswith('Modal') or base.endswith('Modal'): return 'Modal'
    if base.endswith('Drawer'): return 'Drawer'
    if base.endswith('Popover'): return 'Popover'
    if base.endswith('Layout') and 'ExportLayout' not in base: return 'Layout'
    if base in ('AdminPanel','TabelaUsuarios','TabelaWorkspaces','Footer','Navbar','KPICard'): return 'Componente'
    return 'Página'

# Componentes que precisam renomear (DDD)
COMP_RENAMES = {
    'MetricasGeminiAdmin': 'MetricasLLMAdmin',
    'TenantDetail': 'OrganizacaoDetalhe',
    'TestesGeraisAdmin': 'TesteGeralAdmin',
    'EmpresasParceiros': 'EmpresasCadastros',
    'ProcessoLayout_2': '⚠️ DELETAR — duplicata',
}

paginas = []
modais_misturados = []
layouts_misturados = []
componentes_misturados = []
harness_tests = []
url_vazias = []
renames_needed = []
duplicatas_componente = {}

for r in range(2, ws.max_row+1):
    local = ws.cell(r,1).value
    url = str(ws.cell(r,2).value or '').strip()
    url_ddd = ws.cell(r,3).value
    titulo = ws.cell(r,4).value
    fname = str(ws.cell(r,6).value or '').strip()
    comp = str(ws.cell(r,8).value or '').strip()
    produto = ws.cell(r,10).value
    tipo_view = ws.cell(r,12).value
    if not fname: continue

    tipo = tipo_real(fname, comp)
    base_comp = fname.replace('.tsx','').replace('.jsx','')

    if tipo == 'Modal' or tipo == 'Drawer' or tipo == 'Popover':
        modais_misturados.append((r, fname, tipo, produto))
    elif tipo == 'Layout':
        layouts_misturados.append((r, fname, produto))
    elif tipo == 'Componente':
        componentes_misturados.append((r, fname, produto))
    elif tipo == 'TestHarness' or tipo == 'Teste/Story':
        harness_tests.append((r, fname, produto))
    else:
        paginas.append((r, fname, comp, produto, url))

    if not url and tipo == 'Página':
        url_vazias.append((r, fname, produto))

    if base_comp in COMP_RENAMES:
        renames_needed.append((r, fname, COMP_RENAMES[base_comp]))

    # Duplicatas de nome de componente (mesmo nome em produtos diferentes)
    key = base_comp
    duplicatas_componente.setdefault(key, []).append((r, produto))

dupes = {k: v for k, v in duplicatas_componente.items() if len(v) > 1}

total = len([1 for r in range(2, ws.max_row+1) if ws.cell(r,6).value])
print(f'TOTAL linhas: {total}')
print(f'  [A] Páginas reais: {len(paginas)}')
print(f'  [B] Modais misturados (mover para aba 7): {len(modais_misturados)}')
print(f'  [C] Layouts (marcar tipo): {len(layouts_misturados)}')
print(f'  [D] Componentes misturados (mover para aba 9): {len(componentes_misturados)}')
print(f'  [E] Test harness / stories (deletar): {len(harness_tests)}')
print(f'  [F] Páginas sem URL (preencher manual): {len(url_vazias)}')
print(f'  [G] Componentes c/ rename DDD: {len(renames_needed)}')
print(f'  [H] Duplicatas de nome: {len(dupes)}')

print('\n' + '='*60)
print('[B] MODAIS MISTURADOS — mover para aba "7. Modais" ou marcar Tipo=Modal')
print('='*60)
for row in modais_misturados:
    print(f'  L{row[0]:4}: {row[1]:45} [{row[2]}] ({row[3]})')

print('\n' + '='*60)
print('[C] LAYOUTS — marcar "Tipo de view" = Layout')
print('='*60)
for row in layouts_misturados:
    print(f'  L{row[0]:4}: {row[1]:40} ({row[2]})')

print('\n' + '='*60)
print('[D] COMPONENTES misturados — mover para aba "9. Componentes Locais"')
print('='*60)
for row in componentes_misturados:
    print(f'  L{row[0]:4}: {row[1]:40} ({row[2]})')

print('\n' + '='*60)
print('[E] TEST HARNESS / STORIES — DELETAR')
print('='*60)
for row in harness_tests:
    print(f'  L{row[0]:4}: {row[1]:40} ({row[2]})')

print('\n' + '='*60)
print('[G] RENAMES DDD necessários')
print('='*60)
for row in renames_needed:
    print(f'  L{row[0]:4}: {row[1]:35} → {row[2]}')

print('\n' + '='*60)
print('[F] PÁGINAS SEM URL — preencher col B manual (primeiras 30)')
print('='*60)
for row in url_vazias[:30]:
    print(f'  L{row[0]:4}: {row[1]:40} ({row[2]})')
if len(url_vazias) > 30:
    print(f'  ... (mais {len(url_vazias)-30})')

print('\n' + '='*60)
print('[H] DUPLICATAS de nome de componente (mesmo arquivo, produtos diferentes)')
print('='*60)
for k, rs in list(dupes.items())[:30]:
    prods = ' | '.join(str(p[1]) for p in rs)
    linhas = ', '.join(f'L{p[0]}' for p in rs)
    print(f'  {k:30} [{linhas}] produtos: {prods}')
