"""Analisa toda a mapa-rotas e reporta o que precisa mudar, linha por linha."""
import openpyxl
import re

wb = openpyxl.load_workbook(r'C:\Users\danie\Downloads\planilha_geral_gravity_COMPLETA.xlsx', data_only=True)
ws = wb['5. mapa-rotas']

GHOST_FILES = {
    'servicos-global/tenant/atividades/server/routes/pipelines.ts',
    'servicos-global/tenant/atividades/server/routes/kanban.ts',
    'servicos-global/tenant/atividades/server/routes/contatos.ts',
    'servicos-global/tenant/atividades/server/routes/empresas.ts',
    'servicos-global/tenant/api-cockpit/server/src/routes/observability.ts',
    'servicos-global/configurador/server/routes/serviceToken.ts',
    'servicos-global/organizacao/pedido/server/src/routes/dashboardWidgets.ts',
}

BUG_MODELS = {
    'produtoGravityConfig','catalogProduct','itemPedido','pedidoAnexo',
    'pedidoCasasDecimaisConfig','pedidoTemplatePdf','pedidoSaldoFormulaConfig',
    'configuracaoCanalTenant','contatoExterno','preferenciasUsuario',
    'nfDespesaCatalogo','nfDespesaTemplate','nfExportLayout','pedidoHistorico',
    'dashboardConfig',
}

ddd_patterns = [r'\btenants\b', r'\bworkspaces\b', r'\bcompanies\b',
                r':tenantId', r':companyId', r':userId', r':productKey',
                r'/activate', r'/deactivate', r'/subscribe', r'tenant-products',
                r'service-token', r'gemini-metrics']

total = 0
slash_only = []
ghost = []
with_ddd = []
bugs = []
duplicates = {}

for r in range(2, ws.max_row+1):
    rota = str(ws.cell(r,2).value or '').strip()
    metodo = str(ws.cell(r,1).value or '').strip()
    arquivo = str(ws.cell(r,8).value or '').strip()
    model = str(ws.cell(r,13).value or '').strip()
    if not rota: continue
    total += 1

    key = f'{metodo} {rota}'
    if rota != '/':
        duplicates.setdefault(key, []).append(r)

    if rota == '/':
        slash_only.append((r, metodo, arquivo))
        continue

    arq_norm = arquivo.replace(chr(92), '/').lstrip('./')
    if arq_norm in GHOST_FILES:
        ghost.append((r, metodo, rota, arq_norm))
        continue

    if any(re.search(p, rota) for p in ddd_patterns):
        with_ddd.append((r, metodo, rota))

    if model in BUG_MODELS:
        bugs.append((r, metodo, rota, model))

dupes = {k:v for k,v in duplicates.items() if len(v)>1}

print(f'TOTAL linhas com dados: {total}')
print(f'  [A] Rotas com path "/" (extração incompleta): {len(slash_only)}')
print(f'  [B] Ghost (deletar): {len(ghost)}')
print(f'  [C] Precisa DDD rename (col C): {len(with_ddd)}')
print(f'  [D] Bug accessor Prisma (revisar código): {len(bugs)}')
print(f'  [E] Duplicatas exatas: {len(dupes)}')

print('\n' + '='*60)
print('[B] GHOST — PINTAR DE VERMELHO / DELETAR')
print('='*60)
for row in ghost:
    fname = row[3].split("/")[-1]
    print(f'  L{row[0]:4}: {row[1]:7} {row[2]:50} [{fname}]')

print('\n' + '='*60)
print('[C] DDD RENAME — PREENCHER COL C')
print('='*60)
for row in with_ddd:
    print(f'  L{row[0]:4}: {row[1]:7} {row[2]}')

print('\n' + '='*60)
print('[D] BUG ACCESSOR PRISMA — Status DDD: "Bug no código"')
print('='*60)
for row in bugs:
    print(f'  L{row[0]:4}: {row[1]:7} {row[2]:50} [model={row[3]}]')

print('\n' + '='*60)
print('[E] DUPLICATAS EXATAS')
print('='*60)
for k, rs in dupes.items():
    print(f'  {k}: linhas {rs}')

print('\n' + '='*60)
print('[A] SLASH ONLY — resumo por arquivo')
print('='*60)
from collections import Counter
files = Counter()
for _, _, arq in slash_only:
    files[arq.split('/')[-1]] += 1
for fname, cnt in sorted(files.items(), key=lambda x:-x[1]):
    print(f'  {cnt:3}x  {fname}')
