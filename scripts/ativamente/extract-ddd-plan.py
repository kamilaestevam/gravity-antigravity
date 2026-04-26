"""
extract-ddd-plan.py — Extrai plano DDD definitivo da planilha_geral_gravity (22).xlsx

Lê aba '1.ddd_campos' e classifica cada linha pela COR de fundo das células B/D/E:
  AZUL    (~ blue)    -> EDIT/RENAME (col D atual -> col E DDD)
  VERMELHO(~ red)     -> DELETE (drop col D)
  AMARELO (~ yellow)  -> CREATE (criar col E com tipo inferido)
  Sem cor / branco    -> NOOP

Saída: scripts/ativamente/ddd-plan.json com { tabelas: { <Tabela>: { rename:[], drop:[], create:[] } } }
e cobertura por módulo (col A, "Local").
"""

from openpyxl import load_workbook
import json
import sys
import os

XLSX = r"C:\Users\danie\Downloads\planilha_geral_gravity (22).xlsx"
SHEET = "1.ddd_campos"
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ddd-plan.json")


def classify_color(hex_color: str) -> str:
    """Classifica cor RGB hex em AZUL/VERMELHO/AMARELO/NOOP.
    Threshold por dominância de canal."""
    if not hex_color:
        return "noop"
    s = hex_color.upper().lstrip("#")
    # openpyxl retorna 'FFRRGGBB' (ARGB)
    if len(s) == 8:
        s = s[2:]
    if len(s) != 6:
        return "noop"
    try:
        r = int(s[0:2], 16)
        g = int(s[2:4], 16)
        b = int(s[4:6], 16)
    except ValueError:
        return "noop"

    # Branco/transparente
    if r > 240 and g > 240 and b > 240:
        return "noop"
    # Cinza neutro
    if abs(r - g) < 15 and abs(g - b) < 15 and abs(r - b) < 15:
        return "noop"

    # Vermelho dominante (FF0000 family)
    if r > 180 and g < 100 and b < 100:
        return "delete"
    # Amarelo (FFFF00 family — alto R+G, baixo B)
    if r > 200 and g > 180 and b < 120:
        return "create"
    # Azul (0000FF / 4472C4 etc — B alto)
    if b > 150 and (b > r + 30 or b > g + 30):
        return "rename"
    return "noop"


def get_fill_hex(cell) -> str:
    """Extrai hex da cor de fundo (Pattern fill)."""
    fill = cell.fill
    if fill is None or fill.fgColor is None:
        return ""
    fg = fill.fgColor
    if fg.type == "rgb" and fg.rgb:
        return str(fg.rgb)
    if fg.type == "theme":
        # Theme colors são complicados — fallback para indexed/RGB se houver
        return ""
    return ""


def main():
    wb = load_workbook(XLSX, data_only=True)
    if SHEET not in wb.sheetnames:
        print(f"ERRO: aba {SHEET!r} não encontrada. Abas: {wb.sheetnames}", file=sys.stderr)
        sys.exit(1)
    ws = wb[SHEET]

    plan: dict = {}
    stats = {"rename": 0, "delete": 0, "create": 0, "noop": 0, "total": 0}
    modulos: dict = {}

    # Header
    headers = [c.value for c in ws[1]]
    print(f"Headers: {headers[:8]}", file=sys.stderr)

    for row_idx in range(2, ws.max_row + 1):
        local_cell = ws.cell(row=row_idx, column=1)   # A
        tabela_cell = ws.cell(row=row_idx, column=2)  # B
        atual_cell = ws.cell(row=row_idx, column=4)   # D
        ddd_cell = ws.cell(row=row_idx, column=5)     # E
        desc_cell = ws.cell(row=row_idx, column=6)    # F (descrição p/ AMARELO)

        local = (local_cell.value or "").strip() if local_cell.value else ""
        tabela = (tabela_cell.value or "").strip() if tabela_cell.value else ""
        atual = (atual_cell.value or "").strip() if atual_cell.value else ""
        ddd = (ddd_cell.value or "").strip() if ddd_cell.value else ""
        desc = (desc_cell.value or "").strip() if desc_cell.value else ""

        if not tabela:
            continue

        # Cor do B (linha) — fonte primária. Backup: cor de D ou E.
        color_b = classify_color(get_fill_hex(tabela_cell))
        color_d = classify_color(get_fill_hex(atual_cell))
        color_e = classify_color(get_fill_hex(ddd_cell))

        # Eleição: prioriza vermelho, depois amarelo, depois azul, depois noop
        votes = [color_b, color_d, color_e]
        if "delete" in votes:
            action = "delete"
        elif "create" in votes:
            action = "create"
        elif "rename" in votes:
            action = "rename"
        else:
            action = "noop"

        stats[action] += 1
        stats["total"] += 1

        if tabela not in plan:
            plan[tabela] = {
                "modulo": local,
                "rename": [],
                "delete": [],
                "create": [],
                "noop_count": 0,
            }

        modulos.setdefault(local, set()).add(tabela)

        if action == "rename":
            if atual and ddd and atual != ddd:
                plan[tabela]["rename"].append({"de": atual, "para": ddd, "row": row_idx})
        elif action == "delete":
            if atual:
                plan[tabela]["delete"].append({"col": atual, "row": row_idx})
        elif action == "create":
            if ddd:
                plan[tabela]["create"].append({"col": ddd, "desc": desc, "row": row_idx})
        else:
            plan[tabela]["noop_count"] += 1

    # Serialize modulos sets
    modulos_out = {k: sorted(list(v)) for k, v in modulos.items()}

    out = {
        "stats": stats,
        "modulos": modulos_out,
        "tabelas": plan,
    }
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    print(f"OK: {OUT}")
    print(f"Stats: {stats}")
    print(f"Módulos: {len(modulos_out)} | Tabelas: {len(plan)}")
    for m, ts in modulos_out.items():
        print(f"  [{m}] {len(ts)} tabelas")


if __name__ == "__main__":
    main()
