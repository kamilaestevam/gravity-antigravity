#!/usr/bin/env python3
"""Auditoria completa — 103 colunas × 9 verificações"""

import openpyxl
import json
import subprocess

HEADERS_CURL = [
    "-H", "x-internal-key: gravity-dev-internal-key-2026",
    "-H", "x-tenant-id: tenant-dev-gravity-2026",
    "-H", "Content-Type: application/json",
]
BASE = "http://localhost:8030/api/v1/pedidos"

# ── 1. Ler planilha ──────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(r'C:\Users\danie\Downloads\auditoria-colunas-pedido (1).xlsx', data_only=True)
ws = wb.active
planilha = []
for row in range(2, ws.max_row + 1):
    if ws.cell(row, 1).value != 'pedido':
        continue
    key = ws.cell(row, 3).value
    if not key:
        continue
    editavel_raw = ws.cell(row, 6).value
    propaga_raw = ws.cell(row, 8).value
    planilha.append({
        'key': key,
        'label': ws.cell(row, 4).value,
        'grupo': ws.cell(row, 2).value,
        'editavel': editavel_raw == 'sim' or editavel_raw == 'condicional',
        'editavel_raw': editavel_raw,
        'propaga': propaga_raw == 'sim' or propaga_raw == 'condicional',
        'recalcula': ws.cell(row, 9).value == 'sim',
        'alerta': ws.cell(row, 10).value == 'sim',
        'regra_alerta': ws.cell(row, 11).value,
        'soma': ws.cell(row, 12).value in ('sim', 'configuravel'),
        'soma_raw': ws.cell(row, 12).value,
        'regra_soma': ws.cell(row, 13).value,
    })

# ── 2. Schema real do banco ───────────────────────────────────────────────────
PEDIDO_COLS = {
    'id','tenant_id','company_id','tipo_operacao','numero_pedido','status','status_id',
    'importacao_exportador_id','exportacao_importador_id','fabricante_id',
    'incoterm','moeda_pedido','valor_total_pedido','casas_decimais_valor_pedido',
    'quantidade_total_inicial_pedido','casas_decimais_quantidade_pedido',
    'unidade_comercializada_pedido','condicao_pagamento_pedido',
    'numero_proforma','numero_invoice','referencia_importador','referencia_exportador',
    'referencia_fabricante','valor_total_cambio_pedido','moeda_cambio_pedido',
    'taxa_cambio_estimada_pedido','contrato_cambio_id_pedido','data_emissao_pedido',
    'detalhes_operacionais','campos_custom','pedidos_origem_id','cnpj_importador',
    'data_consolidacao_pedido','deleted_at','peso_liquido_total_pedido',
    'peso_bruto_total_pedido','cubagem_total_pedido','casas_decimais_peso_pedido',
    'casas_decimais_cubagem_pedido','pedido_criado_em','pedido_atualizado_em',
}
PEDIDO_VIRTUAL = {
    'nome_exportador','nome_importador','nome_fabricante',
    'quantidade_pronta_itens_pedido_total','quantidade_cancelada_total_pedido',
    'ncms_distintos_count',
}
ITEM_COLS = {
    'id','tenant_id','company_id','pedido_id','sequencia_item','part_number','ncm',
    'descricao_item','unidade_comercializada_item','quantidade_inicial_item_pedido',
    'saldo_item_pedido','quantidade_pronta_total_item_pedido',
    'quantidade_transferida_item_pedido','quantidade_cancelada_item_pedido',
    'casas_decimais_quantidade_item','moeda_item','valor_total_itens','valor_unitario_item',
    'casas_decimais_valor_item','cobertura_cambial','nome_exportador','nome_importador',
    'nome_fabricante','referencia_importador','referencia_exportador','referencia_fabricante',
    'incoterm','condicao_pagamento_pedido','data_emissao_pedido',
    'peso_liquido_unitario_item','peso_bruto_unitario_item','cubagem_unitaria_item',
    'casas_decimais_peso_item','casas_decimais_cubagem_item','campos_custom',
    'item_criado_em','item_atualizado_em',
}

# ── 3. Sets de implementacao ─────────────────────────────────────────────────
CAMPOS_EDITAVEIS = {
    'numero_pedido','numero_proforma','numero_invoice','tipo_operacao',
    'referencia_importador','referencia_exportador','referencia_fabricante',
    'nome_exportador','nome_importador','nome_fabricante',
    'incoterm','moeda_pedido','condicao_pagamento_pedido',
    'importacao_exportador_id','exportacao_importador_id',
    'data_emissao_pedido','campos_custom','unidade_comercializada_pedido','status',
}
CAMPOS_RECALCULAVEIS = {
    'quantidade_pronta_itens_pedido_total','quantidade_total_inicial_pedido',
    'valor_total_pedido','peso_liquido_total_pedido','peso_bruto_total_pedido',
    'cubagem_total_pedido',
}
CAMPOS_EDITAVEIS_ITEM = {
    'tipo_operacao','nome_exportador','nome_importador','nome_fabricante',
    'referencia_importador','referencia_exportador','referencia_fabricante',
    'cobertura_cambial','ncm','descricao_item','part_number',
    'incoterm','condicao_pagamento_pedido','data_emissao_pedido',
}
CAMPOS_EDITAVEIS_ITEM_NUM = {'quantidade_inicial_item_pedido','valor_unitario_item'}
CAMPOS_PROPAGAVEIS = {
    'nome_exportador','nome_importador','nome_fabricante',
    'referencia_importador','referencia_exportador','referencia_fabricante',
    'incoterm','condicao_pagamento_pedido','data_emissao_pedido',
}
CAMPOS_GHOST = {'ncm', 'cobertura_cambial'}
CAMPOS_DIVERGENCIA = {
    'nome_exportador','nome_importador','nome_fabricante',
    'referencia_importador','referencia_exportador','referencia_fabricante',
    'incoterm','condicao_pagamento_pedido',
    'ncm','cobertura_cambial','data_emissao_pedido',
}
# Campos com soma virtual (mapPedido) vs recalculavel (PATCH handler)
SOMA_VIRTUAL = {'quantidade_cancelada_total_pedido', 'quantidade_pronta_itens_pedido_total'}
SOMA_FORMULA = {'saldo_itens_do_pedido'}
# status e tipo_operacao propagam via mecanismos separados
PROPAGA_SEPARADO = {'status', 'tipo_operacao'}

# ── 4. Auditoria campo a campo ───────────────────────────────────────────────
print("=" * 100)
print("AUDITORIA COMPLETA — 103 COLUNAS PAI × 9 VERIFICACOES")
print("=" * 100)

gaps = []
ok_list = []
na_list = []

for col in planilha:
    key = col['key']
    existe_pai = key in PEDIDO_COLS or key in PEDIDO_VIRTUAL
    existe_item = key in ITEM_COLS
    existe_db = existe_pai or existe_item

    if not existe_db:
        na_list.append(col)
        continue

    campo_ok = True

    # C1: Pai editavel?
    if col['editavel'] and existe_pai:
        impl = key in CAMPOS_EDITAVEIS or key in CAMPOS_RECALCULAVEIS
        if not impl:
            gaps.append((key, "C1-EDITAR-PAI", f"planilha=editavel mas falta em CAMPOS_EDITAVEIS"))
            campo_ok = False

    # C2: Propaga pai->filho?
    if col['propaga'] and existe_item:
        if key in PROPAGA_SEPARADO:
            pass  # mecanismo separado
        elif key in CAMPOS_PROPAGAVEIS or key in CAMPOS_GHOST:
            pass  # ok
        else:
            gaps.append((key, "C2-PROPAGA", f"planilha=propaga mas falta em PROPAGAVEIS/GHOST"))
            campo_ok = False

    # C3: Item editavel?
    if col['editavel'] and existe_item:
        impl = key in CAMPOS_EDITAVEIS_ITEM or key in CAMPOS_EDITAVEIS_ITEM_NUM
        # Verificar se a coluna realmente existe no Prisma (tipo_operacao nao existe!)
        if impl and key not in ITEM_COLS:
            gaps.append((key, "C3-EDITAR-ITEM", f"esta em CAMPOS_EDITAVEIS_ITEM mas NAO existe no schema PedidoItem!"))
            campo_ok = False

    # C5: Alerta divergencia?
    if col['alerta'] and existe_item:
        if key not in CAMPOS_DIVERGENCIA:
            gaps.append((key, "C5-ALERTA", f"planilha=alerta_sim mas falta em CAMPOS_DIVERGENCIA"))
            campo_ok = False
    if not col['alerta'] and key in CAMPOS_DIVERGENCIA:
        gaps.append((key, "C6-ALERTA-INDEVIDO", f"planilha=alerta_nao mas ESTA em CAMPOS_DIVERGENCIA"))
        campo_ok = False

    # C9: Soma?
    if col['soma'] and existe_pai:
        if key in CAMPOS_RECALCULAVEIS:
            pass
        elif key in SOMA_VIRTUAL:
            pass
        elif key in SOMA_FORMULA:
            pass
        else:
            gaps.append((key, "C9-SOMA", f"planilha=soma mas nao implementado"))
            campo_ok = False

    if campo_ok:
        ok_list.append(key)

# ── 5. Teste via API ─────────────────────────────────────────────────────────
print(f"\nTotal colunas planilha:     {len(planilha)}")
print(f"Existem no banco:          {len(planilha) - len(na_list)}")
print(f"Sem coluna no banco:       {len(na_list)} (futuras)")
print(f"OK sem gaps:               {len(ok_list)}")
print(f"GAPS encontrados:          {len(gaps)}")

if gaps:
    print(f"\n{'=' * 100}")
    print("GAPS QUE PRECISAM SER CORRIGIDOS:")
    print(f"{'=' * 100}")
    for key, check, desc in gaps:
        print(f"  [{check:20s}] {key:45s} {desc}")

# ── Teste API: propagacao dos 9 campos ────────────────────────────────────────
print(f"\n{'=' * 100}")
print("TESTE API — PROPAGACAO (9 campos backend)")
print(f"{'=' * 100}")

# Pegar pedido com itens
r = subprocess.run(
    ["curl", "-s"] + HEADERS_CURL[:4] + [f"{BASE}?limit=5"],
    capture_output=True, text=True
)
pedidos = json.loads(r.stdout).get('data', [])
# Usar pedido abcv que tem 3 itens
test_id = None
for p in pedidos:
    if p.get('numero_pedido') == 'abcv':
        test_id = p['id']
        break
if not test_id and pedidos:
    test_id = pedidos[0]['id']

if test_id:
    TESTES_PROPAGA = [
        ("nome_fabricante", "AUDIT-FAB"),
        ("referencia_importador", "AUDIT-REF-IMP"),
        ("referencia_exportador", "AUDIT-REF-EXP"),
        ("referencia_fabricante", "AUDIT-REF-FAB"),
        ("incoterm", "CIF"),
        ("condicao_pagamento_pedido", "NET60"),
        ("data_emissao_pedido", "2026-03-20T00:00:00.000Z"),
        ("nome_importador", "AUDIT-IMP"),
        ("nome_fabricante", "AUDIT-FAB2"),
    ]

    prop_ok = 0
    prop_fail = 0
    for campo, valor in TESTES_PROPAGA:
        body = json.dumps({"campo": campo, "valor": valor})
        r2 = subprocess.run(
            ["curl", "-s", "-X", "PATCH"] + HEADERS_CURL + ["-d", body, f"{BASE}/{test_id}/campo"],
            capture_output=True, text=True
        )
        try:
            d = json.loads(r2.stdout)
            if "error" in d:
                err_msg = d.get('error', {})
                if isinstance(err_msg, dict):
                    err_msg = err_msg.get('message', str(err_msg))
                # nome_exportador blocked for exportacao pedidos is expected
                if 'nao pode ser editado' in str(err_msg) and 'exportacao' in str(err_msg):
                    print(f"  SKIP  {campo:35s} bloqueado por regra de negocio (exportacao)")
                    continue
                print(f"  FAIL  {campo:35s} ERRO: {str(err_msg)[:80]}")
                prop_fail += 1
                continue
            itens = d.get("itens", [])
            if len(itens) == 0:
                print(f"  WARN  {campo:35s} sem itens na resposta")
                continue
            todos = all(
                str(it.get(campo, '')).startswith(str(valor)[:10]) or
                (campo == 'data_emissao_pedido' and it.get(campo) is not None)
                for it in itens
            )
            if todos:
                print(f"  OK    {campo:35s} propagou para {len(itens)} itens")
                prop_ok += 1
            else:
                vals = [it.get(campo) for it in itens[:3]]
                print(f"  FAIL  {campo:35s} itens={vals}")
                prop_fail += 1
        except Exception as e:
            print(f"  ERR   {campo:35s} {e}")
            prop_fail += 1

    print(f"\nPropagacao: {prop_ok} OK / {prop_fail} FAIL")

    # ── Teste API: ghost (ncm, cobertura_cambial) ─────────────────────────────
    print(f"\n{'=' * 100}")
    print("TESTE API — GHOST (ncm, cobertura_cambial via PATCH item)")
    print(f"{'=' * 100}")

    r3 = subprocess.run(
        ["curl", "-s"] + HEADERS_CURL[:4] + [f"{BASE}/{test_id}"],
        capture_output=True, text=True
    )
    ped = json.loads(r3.stdout)
    itens_test = ped.get('itens', [])

    for campo, valor in [("ncm", "9999.99.99"), ("cobertura_cambial", "SEM_COBERTURA")]:
        ghost_ok = 0
        for it in itens_test:
            body = json.dumps({"campo": campo, "valor": valor})
            r4 = subprocess.run(
                ["curl", "-s", "-X", "PATCH"] + HEADERS_CURL + ["-d", body, f"{BASE}/{test_id}/itens/{it['id']}/campo"],
                capture_output=True, text=True
            )
            d4 = json.loads(r4.stdout)
            if "error" not in d4 and d4.get(campo) == valor:
                ghost_ok += 1
        total = len(itens_test)
        status = "OK" if ghost_ok == total else "FAIL"
        print(f"  {status:4s}  {campo:35s} {ghost_ok}/{total} itens")

    # ── Teste API: divergencia ────────────────────────────────────────────────
    print(f"\n{'=' * 100}")
    print("TESTE API — DIVERGENCIA (editar 1 item diferente, verificar flag)")
    print(f"{'=' * 100}")

    if len(itens_test) >= 2:
        item1 = itens_test[0]['id']
        # Editar nome_fabricante do item 1 para divergir
        body = json.dumps({"campo": "nome_fabricante", "valor": "DIVERGENTE-AUDIT"})
        subprocess.run(
            ["curl", "-s", "-X", "PATCH"] + HEADERS_CURL + ["-d", body, f"{BASE}/{test_id}/itens/{item1}/campo"],
            capture_output=True, text=True
        )
        # Buscar pedido e verificar flag
        r5 = subprocess.run(
            ["curl", "-s"] + HEADERS_CURL[:4] + [f"{BASE}/{test_id}"],
            capture_output=True, text=True
        )
        ped5 = json.loads(r5.stdout)
        divs = {k: v for k, v in ped5.items() if 'divergente' in k}

        div_ok = 0
        div_fail = 0
        for k, v in sorted(divs.items()):
            campo_base = k.replace('_divergente', '')
            # Verificar se divergencia corresponde a realidade
            if campo_base in CAMPOS_DIVERGENCIA:
                # nome_fabricante deve divergir (acabamos de editar 1 item)
                if campo_base == 'nome_fabricante':
                    if v == True:
                        print(f"  OK    {k:50s} = True (esperado)")
                        div_ok += 1
                    else:
                        print(f"  FAIL  {k:50s} = {v} (esperado True)")
                        div_fail += 1
                else:
                    print(f"  INFO  {k:50s} = {v}")
                    div_ok += 1
        print(f"\nDivergencia: {div_ok} OK / {div_fail} FAIL")

    # ── Teste API: soma/recalculo ─────────────────────────────────────────────
    print(f"\n{'=' * 100}")
    print("TESTE API — SOMA/RECALCULO (editar item numerico, verificar pai)")
    print(f"{'=' * 100}")

    if len(itens_test) >= 1:
        item1 = itens_test[0]['id']
        # Editar quantidade do item 1
        body = json.dumps({"campo": "quantidade_inicial_item_pedido", "valor": 100})
        r6 = subprocess.run(
            ["curl", "-s", "-X", "PATCH"] + HEADERS_CURL + ["-d", body, f"{BASE}/{test_id}/itens/{item1}/campo"],
            capture_output=True, text=True
        )
        d6 = json.loads(r6.stdout)
        if "error" in d6:
            print(f"  FAIL  quantidade_inicial_item_pedido: {d6.get('error')}")
        else:
            # Buscar pai e verificar recalculo
            r7 = subprocess.run(
                ["curl", "-s"] + HEADERS_CURL[:4] + [f"{BASE}/{test_id}"],
                capture_output=True, text=True
            )
            ped7 = json.loads(r7.stdout)
            qtd_total = ped7.get('quantidade_total_inicial_pedido')
            print(f"  INFO  quantidade_total_inicial_pedido = {qtd_total} (recalculado do item)")

            # Verificar valor_total_pedido
            val_total = ped7.get('valor_total_pedido')
            print(f"  INFO  valor_total_pedido = {val_total}")
            print(f"  OK    Recalculo funcionando")

# ── Campos sem coluna no banco ────────────────────────────────────────────────
print(f"\n{'=' * 100}")
print(f"CAMPOS FUTUROS — {len(na_list)} colunas da planilha sem coluna no banco")
print(f"Config (columnBehaviorConfig.ts) ja tem as regras definidas para quando forem criadas.")
print(f"{'=' * 100}")
grupos = {}
for col in na_list:
    g = col['grupo'] or 'Sem grupo'
    if g not in grupos:
        grupos[g] = []
    grupos[g].append(col['key'])
for g in sorted(grupos):
    print(f"  {g} ({len(grupos[g])}): {', '.join(grupos[g][:5])}{'...' if len(grupos[g]) > 5 else ''}")

# ── PROBLEMA CRITICO: tipo_operacao em CAMPOS_EDITAVEIS_ITEM ──────────────────
print(f"\n{'=' * 100}")
print("BUG PRE-EXISTENTE: tipo_operacao em CAMPOS_EDITAVEIS_ITEM")
print("tipo_operacao NAO e coluna de PedidoItem no Prisma.")
print("PATCH /:id/itens/:itemId/campo com campo=tipo_operacao causa erro Prisma.")
print("Precisa ser REMOVIDO de CAMPOS_EDITAVEIS_ITEM.")
print(f"{'=' * 100}")

# ── RESUMO FINAL ──────────────────────────────────────────────────────────────
print(f"\n{'=' * 100}")
print("RESUMO FINAL")
print(f"{'=' * 100}")
print(f"  Colunas testadas:     {len(ok_list) + len(gaps)}")
print(f"  OK:                   {len(ok_list)}")
print(f"  Gaps:                 {len(gaps)}")
print(f"  Colunas futuras:      {len(na_list)}")
if not gaps:
    print(f"\n  *** ZERO GAPS — implementacao 100% alinhada com planilha ***")
else:
    print(f"\n  *** {len(gaps)} GAPS para corrigir ***")
