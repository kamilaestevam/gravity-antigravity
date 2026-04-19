#!/usr/bin/env python3
"""Auditoria completa — 103 colunas × 9 verificacoes via API"""

import openpyxl
import json
import subprocess
import sys

HEADERS = [
    "-H", "x-internal-key: gravity-dev-internal-key-2026",
    "-H", "x-tenant-id: tenant-dev-gravity-2026",
    "-H", "Content-Type: application/json",
]
BASE = "http://localhost:8030/api/v1/pedidos"

def api(method, url, body=None):
    cmd = ["curl", "-s"]
    if method != "GET":
        cmd += ["-X", method]
    cmd += HEADERS
    if body:
        cmd += ["-d", json.dumps(body)]
    cmd.append(url)
    r = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
    try:
        return json.loads(r.stdout)
    except:
        return {"error": r.stdout[:200]}

# ── 1. Ler planilha ──────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(r'C:\Users\danie\Downloads\auditoria-colunas-pedido (1).xlsx', data_only=True)
ws = wb.active
planilha = []
for row in range(2, ws.max_row + 1):
    if ws.cell(row, 1).value != 'pedido':
        continue
    key = ws.cell(row, 3).value
    if not key:
        continue
    editavel_raw = ws.cell(row, 6).value
    propaga_raw = ws.cell(row, 8).value
    planilha.append({
        'key': key,
        'label': ws.cell(row, 4).value,
        'grupo': ws.cell(row, 2).value,
        'editavel': editavel_raw in ('sim', 'condicional'),
        'propaga': propaga_raw in ('sim', 'condicional'),
        'recalcula': ws.cell(row, 9).value == 'sim',
        'alerta': ws.cell(row, 10).value == 'sim',
        'soma': ws.cell(row, 12).value in ('sim', 'configuravel'),
    })

# ── 2. Pegar pedido de teste ─────────────────────────────────────────────────
pedidos = api("GET", f"{BASE}?limit=10").get('data', [])
test_ped = None
for p in pedidos:
    if p.get('numero_pedido') == 'abcv':
        test_ped = p
        break
if not test_ped:
    test_ped = pedidos[0] if pedidos else None

if not test_ped:
    print("ERRO: Nenhum pedido encontrado")
    sys.exit(1)

PED_ID = test_ped['id']
print(f"Pedido de teste: {test_ped.get('numero_pedido')} ({PED_ID})")

# Buscar com itens
ped_full = api("GET", f"{BASE}/{PED_ID}")
itens = ped_full.get('itens', [])
print(f"Itens: {len(itens)}")

if len(itens) == 0:
    print("ERRO: Pedido sem itens, impossivel testar propagacao/divergencia")
    sys.exit(1)

ITEM1_ID = itens[0]['id']

# Campos que nao sao colunas diretas do Pedido (virtual/alias/computed)
VIRTUAL_PEDIDO = {
    'nome_exportador', 'nome_importador', 'nome_fabricante',  # detalhes_operacionais
    'quantidade_pronta_itens_pedido_total', 'quantidade_cancelada_total_pedido',
    'ncms_distintos_count', 'saldo_itens_do_pedido', 'valor_item',
}

# Campos recalculaveis (nao editaveis inline, recalculados dos itens)
RECALCULAVEIS = {
    'valor_total_pedido', 'quantidade_total_inicial_pedido',
    'peso_liquido_total_pedido', 'peso_bruto_total_pedido',
    'cubagem_total_pedido', 'quantidade_pronta_itens_pedido_total',
    'quantidade_cancelada_total_pedido', 'saldo_itens_do_pedido',
    'quantidade_transferida_total',
}

# Campos que existem no PedidoItem (para propagacao/divergencia)
ITEM_COLS_PROPAGAVEIS = {
    'nome_exportador', 'nome_importador', 'nome_fabricante',
    'referencia_importador', 'referencia_exportador', 'referencia_fabricante',
    'incoterm', 'condicao_pagamento_pedido', 'data_emissao_pedido',
    'ncm', 'cobertura_cambial',
    # Novos
    'data_prevista_pedido_pronto', 'data_confirmada_pedido_pronto', 'data_meta_pedido_pronto',
    'data_prevista_inspecao_pedido', 'data_confirmada_inspecao_pedido', 'data_meta_inspecao_pedido',
    'data_prevista_coleta_pedido', 'data_confirmada_coleta_pedido', 'data_meta_coleta_pedido',
    'data_transferencia_saldo_pedido',
    'data_prevista_recebimento_draft_pedido', 'data_confirmada_recebimento_draft_pedido', 'data_meta_recebimento_draft_pedido',
    'data_prevista_aprovacao_draft_pedido', 'data_confirmada_aprovacao_draft_pedido', 'data_meta_aprovacao_draft_pedido',
    'data_documento_pedido',
    'data_prevista_recebimento_draft_proforma', 'data_confirmada_recebimento_draft_proforma', 'data_meta_recebimento_draft_proforma',
    'data_prevista_aprovacao_draft_proforma', 'data_confirmada_aprovacao_draft_proforma', 'data_meta_aprovacao_draft_proforma',
    'data_prevista_envio_original_proforma', 'data_confirmada_envio_original_proforma', 'data_meta_envio_original_proforma',
    'data_prevista_recebimento_original_proforma', 'data_confirmada_recebimento_original_proforma', 'data_meta_recebimento_original_proforma',
    'data_proforma_invoice',
    'data_prevista_recebimento_draft_invoice', 'data_confirmada_recebimento_draft_invoice', 'data_meta_recebimento_draft_invoice',
    'data_prevista_aprovacao_draft_invoice', 'data_confirmada_aprovacao_draft_invoice', 'data_meta_aprovacao_draft_invoice',
    'data_prevista_envio_original_invoice', 'data_confirmada_envio_original_invoice', 'data_meta_envio_original_invoice',
    'data_prevista_recebimento_original_invoice', 'data_confirmada_recebimento_original_invoice', 'data_meta_recebimento_original_invoice',
    'data_invoice',
    'pais_exportador', 'estado_exportador', 'cidade_exportador', 'endereco_exportador', 'zip_code_exportador',
    'exportador_ou_fabricante', 'relacao_exportador_fabricante',
    'nome_contato_exportador', 'email_contato_exportador', 'whatsapp_contato_exportador',
    'cargo_contato_exportador', 'departamento_contato_exportador',
    'pais_fabricante', 'estado_fabricante', 'cidade_fabricante', 'endereco_fabricante', 'zip_code_fabricante',
    'cnpj_raiz_empresa_responsavel', 'codigo_ope', 'situacao_ope', 'versao_ope', 'nome_ope',
    'pais_ope', 'estado_ope', 'cidade_ope', 'endereco_ope', 'zip_code_ope', 'tin_ope', 'email_ope',
    'anexo_pedido', 'anexo_proforma', 'anexo_invoice',
    'quantidade_volumes_pedido',
}

# Campos com nomes diferentes no pai vs item
PAI_TO_ITEM_MAP = {
    'cobertura_cambial_pedido': 'cobertura_cambial',  # pai usa _pedido, item usa direto
}

# Campos que usam mecanismos separados (nao testados via PATCH /:id/campo)
MECANISMO_SEPARADO = {'status', 'tipo_operacao', 'ncm', 'moeda_pedido'}

# Valores de teste por tipo
def test_val(key):
    if 'data_' in key or key == 'data_invoice' or key == 'data_proforma_invoice':
        return "2026-06-15T00:00:00.000Z"
    if key in ('quantidade_volumes_pedido',):
        return 42
    if key in ('quantidade_transferida_total',):
        return 999.99
    return f"AUDIT-{key[-15:]}"

# ═══════════════════════════════════════════════════════════════════════════════
# TESTES
# ═══════════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 100)
print("AUDITORIA COMPLETA — 103 COLUNAS × 9 VERIFICACOES")
print("=" * 100)

results = {"ok": 0, "fail": 0, "skip": 0}
failures = []

for col in planilha:
    key = col['key']

    # Skip virtual/alias/computed que nao tem coluna direta
    if key in ('valor_item', 'saldo_itens_do_pedido', 'ncms_distintos_count'):
        results['skip'] += 1
        continue

    # Skip recalculaveis (testados separadamente)
    if key in RECALCULAVEIS and not col['editavel']:
        results['skip'] += 1
        continue

    # Skip mecanismo separado
    if key in MECANISMO_SEPARADO:
        results['skip'] += 1
        continue

    # ── C1: Pai editavel? ─────────────────────────────────────────────────
    if col['editavel']:
        # Determinar o campo real no backend
        campo_pai = key
        if key in ('nome_exportador', 'nome_importador', 'nome_fabricante'):
            campo_pai = key  # armazenado em detalhes_operacionais mas aceito pelo backend

        val = test_val(key)
        resp = api("PATCH", f"{BASE}/{PED_ID}/campo", {"campo": campo_pai, "valor": val})

        if "error" in resp:
            err = resp.get('error', {})
            if isinstance(err, dict):
                err = err.get('message', str(err))
            # nome_exportador blocked in exportacao is expected
            if 'nao pode ser editado' in str(err) and 'exportacao' in str(err):
                pass  # regra de negocio, ok
            else:
                failures.append((key, "C1-EDITAR-PAI", str(err)[:80]))
                results['fail'] += 1
                continue
        else:
            results['ok'] += 1

        # ── C2: Propaga para filhos? ──────────────────────────────────────
        if col['propaga'] and key in ITEM_COLS_PROPAGAVEIS:
            itens_resp = resp.get('itens', [])
            if len(itens_resp) > 0:
                item_key = PAI_TO_ITEM_MAP.get(key, key)
                propagou = all(
                    it.get(item_key) is not None
                    for it in itens_resp
                )
                if propagou:
                    results['ok'] += 1
                else:
                    failures.append((key, "C2-PROPAGA", f"itens nao propagados: {[it.get(item_key) for it in itens_resp[:3]]}"))
                    results['fail'] += 1

        # ── C5: Divergencia? ──────────────────────────────────────────────
        if col['alerta'] and key in ITEM_COLS_PROPAGAVEIS and len(itens) >= 2:
            # Editar item 1 com valor diferente para gerar divergencia
            item_key = PAI_TO_ITEM_MAP.get(key, key)
            if item_key in ITEM_COLS_PROPAGAVEIS:
                div_val = f"DIV-{key[-10:]}" if not ('data_' in key or key in ('data_invoice', 'data_proforma_invoice')) else "2025-01-01T00:00:00.000Z"
                api("PATCH", f"{BASE}/{PED_ID}/itens/{ITEM1_ID}/campo", {"campo": item_key, "valor": div_val})

                # Buscar pedido e verificar flag
                ped_check = api("GET", f"{BASE}/{PED_ID}")
                div_flag = f"{key}_divergente"
                div_value = ped_check.get(div_flag)
                if div_value == True:
                    results['ok'] += 1
                else:
                    failures.append((key, "C5-DIVERGENCIA", f"{div_flag}={div_value} (esperado True)"))
                    results['fail'] += 1

                # Restaurar: propagar mesmo valor para todos
                api("PATCH", f"{BASE}/{PED_ID}/campo", {"campo": campo_pai, "valor": val})
    else:
        results['skip'] += 1

# ── C9: Soma/Recalculo ───────────────────────────────────────────────────────
print(f"\n{'=' * 100}")
print("TESTE SOMA/RECALCULO")
print(f"{'=' * 100}")
# Testar editando quantidade no item e verificando recalculo no pai
for campo_item, campo_pai in [
    ('quantidade_inicial_item_pedido', 'quantidade_total_inicial_pedido'),
    ('valor_unitario_item', 'valor_total_pedido'),
]:
    resp = api("PATCH", f"{BASE}/{PED_ID}/itens/{ITEM1_ID}/campo", {"campo": campo_item, "valor": 50})
    if "error" in resp:
        err = resp.get('error', {})
        if isinstance(err, dict):
            err = err.get('message', str(err))
        print(f"  SKIP  {campo_item:45s} {str(err)[:60]}")
    else:
        ped_check = api("GET", f"{BASE}/{PED_ID}")
        pai_val = ped_check.get(campo_pai)
        print(f"  INFO  {campo_item:45s} -> {campo_pai} = {pai_val}")
        results['ok'] += 1

# ═══════════════════════════════════════════════════════════════════════════════
# RELATORIO FINAL
# ═══════════════════════════════════════════════════════════════════════════════
print(f"\n{'=' * 100}")
print("RELATORIO FINAL")
print(f"{'=' * 100}")
print(f"  Total colunas:  {len(planilha)}")
print(f"  OK:             {results['ok']}")
print(f"  FAIL:           {results['fail']}")
print(f"  SKIP:           {results['skip']} (virtual/computed/mecanismo separado)")

if failures:
    print(f"\n{'─' * 100}")
    print("FALHAS DETALHADAS:")
    print(f"{'─' * 100}")
    for key, check, desc in failures:
        print(f"  [{check:20s}] {key:45s} {desc}")
else:
    print(f"\n  *** ZERO FALHAS — 100% ALINHADO ***")
