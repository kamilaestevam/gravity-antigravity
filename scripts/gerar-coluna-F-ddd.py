"""Gera a coluna F (Nome no banco - DDD) do ddd_campos.

Regras:
- vazio se não precisa mudar
- preenchido se precisa mudar
"""
import openpyxl, re

SRC = r'C:\Users\danie\Downloads\planilha_geral_gravity_COMPLETA.xlsx'

# Núcleo de cada tabela (default = snake_case do nome Prisma, com overrides)
NUCLEO_OVERRIDE = {
    'AssinaturaProdutoGravity': 'assinatura',
    'NegociacaoEspecial': 'negociacao_especial',
    'FornecedorOrganizacao': 'organizacao_fornecedor',
    'ProdutoGravityWorkspace': 'produtos_gravity_workspace',
    'MetricasGemini': 'metricas_llm',
}

# Renomeações fixas de campos
FK_RENAMES = {
    'tenant_id': 'id_organizacao',
    'company_id': 'id_workspace',
    'user_id': 'id_usuario',
    'product_id': 'id_produto',
    'tenant': 'id_organizacao',
    'user': 'id_usuario',
    'company': 'id_workspace',
    'organizacao': 'id_organizacao',
    'workspace': 'id_workspace',
}

# Campos que NÃO renomear (devem virar vazios)
KEEP_EMPTY = {
    # Stripe
    'stripe_subscription_id','stripe_price_id','stripe_customer_id',
    'current_period_start','current_period_end','cancelled_at',
    # Não tocar nessas siglas/ISO (valores internos de enum)
    'USD','EUR','BRL','CNY','JPY','GBP','CHF','ARS','UYU',
    'FOB','CIF','EXW','CFR','FCA','DDP','DAP','CPT','CIP','DPU','FAS',
    'II','IPI','PIS','COFINS','ICMS',
    'AWB','BL','CRT','FCL','LCL',
    # Stripe enum values
    'ACTIVE','PAST_DUE','CANCELLED','TRIALING','INCOMPLETE',
    'DRAFT','OPEN','PAID','VOID','OVERDUE','UNCOLLECTIBLE',
}

# Expansões
ABBREV = {
    'exec': 'execucao',
    'freq': 'frequencia',
    'qtd': 'quantidade',
    'desc': 'descricao',
    'prev': 'previsao',
    'conf': 'confirmacao',
}

def snake(pascal):
    """PascalCase -> snake_case"""
    s = re.sub(r'(.)([A-Z][a-z]+)', r'\1_\2', pascal)
    s = re.sub(r'([a-z0-9])([A-Z])', r'\1_\2', s)
    return s.lower()

def nucleo_de(tabela):
    if tabela in NUCLEO_OVERRIDE:
        return NUCLEO_OVERRIDE[tabela]
    return snake(tabela)

def expandir(campo):
    """Expande abreviações: ultima_exec → ultima_execucao"""
    parts = campo.split('_')
    out = [ABBREV.get(p, p) for p in parts]
    return '_'.join(out)

def gerar_ddd(tabela, atual):
    if not tabela or not atual: return ''
    atual = str(atual).strip()
    tabela = str(tabela).strip()

    # Stripe / siglas: vazio
    if atual in KEEP_EMPTY: return ''

    # FKs conhecidas
    if atual in FK_RENAMES:
        return FK_RENAMES[atual]

    # Enums (entrada de tabela = nome do enum, atual = nome do enum): vazio (salvo renomes)
    # Isso detecta pelo fato de a linha ter tabela == atual (padrão dos enums no sheet)
    # Para enums que precisam renomear:
    if atual == 'MetricasGemini': return 'MetricasLLM'
    if atual == tabela:
        # é linha de "cabeçalho" de enum. Não mexer, salvo overrides acima.
        return ''

    nucleo = nucleo_de(tabela)

    # PK id → id_<nucleo_completo>
    if atual == 'id':
        # Para PK, o núcleo usa o nome completo (não a versão abreviada das overrides)
        # Ex: id_assinatura_produto_gravity (não id_assinatura)
        nucleo_pk = snake(tabela) if tabela in NUCLEO_OVERRIDE else nucleo
        return f'id_{nucleo_pk}'

    # timestamps
    if atual == 'created_at':
        return f'data_criacao_{nucleo}'
    if atual == 'updated_at':
        return f'data_atualizacao_{nucleo}'
    if atual == 'deleted_at':
        return f'data_exclusao_{nucleo}'

    # FK com sufixo _id: <entidade>_id → id_<entidade>
    if atual.endswith('_id'):
        ent = atual[:-3]
        return f'id_{ent}'

    # Campo genérico: aplicar abreviações + sufixar com núcleo
    campo_expandido = expandir(atual)
    alvo = f'{campo_expandido}_{nucleo}'

    # Se já está assim, não mudar
    if atual == alvo: return ''

    return alvo

# Run and output column F
wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb['1. ddd_campos']

lines = []
for r in range(2, ws.max_row+1):
    tabela = ws.cell(r, 2).value
    atual = ws.cell(r, 5).value  # col E = Nome no banco - Prisma
    ddd = gerar_ddd(tabela, atual)
    lines.append(ddd if ddd else '')

# Save to file for paste
out_path = r'C:\Users\danie\Downloads\coluna_F_DDD.txt'
with open(out_path, 'w', encoding='utf-8') as f:
    for line in lines:
        f.write(line + '\n')

total_filled = sum(1 for l in lines if l)
print(f'Total linhas: {len(lines)}')
print(f'Preenchidas: {total_filled}')
print(f'Vazias: {len(lines)-total_filled}')
print(f'\nArquivo: {out_path}')
