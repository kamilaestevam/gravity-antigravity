"""Gera mapa-componentes-locais final com col D (Explicação NOVA),
corrige col E (componente) onde vier constante/helper."""
import openpyxl, csv, re

SRC = r'C:\Users\danie\Downloads\planilha_geral_gravity_COMPLETA.xlsx'
DST = r'C:\Users\danie\Downloads\mapa_componentes_locais_final.csv'

# Explicações por componente/arquivo conhecido
EXPLICACOES = {
    # Configurador
    'GabiOnboardingWidget': 'Widget de onboarding da Gabi exibido no Hub/Core — guia interativo do usuário.',
    'HubButton': 'Botão do Hub que lança um produto do catálogo no Shell.',
    'PremiumEcosystemPuzzle': 'Visualização do ecossistema Premium (puzzle de produtos integrados).',
    'AdminLayout': 'Layout das telas administrativas (sidebar + outlet).',
    'E2ENotificacoesHarness': '⚠️ Test harness E2E para notificações — não é componente de produção.',
    'WorkspaceLayout': 'Layout das telas do workspace (sidebar + outlet).',
    # Marketplace
    'OnboardingPreview': 'Preview visual do onboarding exibido na home do marketplace.',
    'Footer': 'Rodapé do marketplace público.',
    'Layout': 'Layout base compartilhado (navbar + outlet + footer).',
    'Navbar': 'Barra de navegação do marketplace público.',
    # Shell
    # (já coberto por Layout acima)
    # Tenant
    'KPICard': 'Card genérico de KPI reutilizável no dashboard do tenant.',
    # Pedido
    'CelulaAnexosColuna': 'Célula especial da tabela de pedidos que renderiza anexos (thumbnails).',
    'GerenciadorColunas': 'Gerenciador de colunas customizadas do tenant na tabela Pedido (CRUD de PedidoColuna).',
    'BarraAcoesPedido': 'Toolbar com ações em lote para pedidos selecionados (excluir, duplicar, transferir, etc.).',
    'colunasFilho': 'Definições das colunas dos itens do pedido (tabela filho) — inclui fórmulas e formatação.',
    'colunasPai': 'Definições das colunas do pedido (tabela pai) — inclui fórmulas e formatação.',
    'PainelAnexos': 'Painel lateral com lista de anexos do pedido.',
    'EtapaConfirmacao': 'Etapa final do wizard SmartImport — confirma e aplica a importação.',
    'EtapaMapeamento': 'Etapa do wizard SmartImport — mapeamento de colunas da planilha ↔ campos do pedido (IA).',
    'EtapaPreview': 'Etapa do wizard SmartImport — preview dos dados que serão importados.',
    'EtapaUpload': 'Etapa inicial do wizard SmartImport — upload do arquivo Excel/CSV.',
    # Processo
    'ProcessoLayout': 'Layout do produto Processo (abas internas).',
}

def fix_component_name(fname, current):
    """Se current é constante/helper/lowercase, usa filename."""
    if not current: return fname.replace('.tsx','').replace('.jsx','')
    c = str(current).strip()
    # Nomes inválidos
    if c.isupper() or '_' in c or (c and c[0].islower()):
        return fname.replace('.tsx','').replace('.jsx','')
    return c

def gerar_explicacao(comp_name, fname, pasta):
    if comp_name in EXPLICACOES:
        return EXPLICACOES[comp_name]
    base = fname.replace('.tsx','').replace('.jsx','')
    if base in EXPLICACOES:
        return EXPLICACOES[base]
    # Fallback
    if pasta == 'layout':
        return f'Layout local — wrapper com navegação compartilhada.'
    if pasta == 'widgets':
        return f'Widget local — componente reutilizável do dashboard.'
    # Nome humanizado
    human = re.sub(r'(.)([A-Z][a-z]+)', r'\1 \2', comp_name or base)
    human = re.sub(r'([a-z0-9])([A-Z])', r'\1 \2', human)
    return f'Componente local: {human.strip()}.'

# Processa
wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb['9. Componentes Locais']

header_atual = [ws.cell(1,c).value or '' for c in range(1, 15)]
# Insere col "Explicação" após E (componente DDD)
new_header = header_atual[:5] + ['Explicação'] + header_atual[5:]

rows_out = [new_header]
for r in range(2, ws.max_row+1):
    row = [ws.cell(r,c).value for c in range(1, 15)]
    if not row[1]: continue

    fname = str(row[1] or '').strip()
    comp_atual = str(row[3] or '').strip()
    pasta = str(row[6] or '').strip()

    # Corrige col D (Nome do componente)
    fixed = fix_component_name(fname, comp_atual)
    if fixed != comp_atual:
        row[3] = fixed

    # Explicação (inserida após E)
    explicacao = gerar_explicacao(fixed, fname, pasta)

    new_row = row[:5] + [explicacao] + row[5:]
    rows_out.append(new_row)

with open(DST, 'w', encoding='utf-8', newline='') as f:
    writer = csv.writer(f)
    for r in rows_out:
        writer.writerow(['' if v is None else v for v in r])

total = len(rows_out) - 1
filled_exp = sum(1 for r in rows_out[1:] if r[5])
print(f'Linhas: {total}')
print(f'Col F Explicação (nova): {filled_exp}')
print(f'Arquivo: {DST}')
