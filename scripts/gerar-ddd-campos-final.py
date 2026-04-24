"""Gera ddd_campos inteiro final em CSV."""
import openpyxl, csv, re

SRC = r'C:\Users\danie\Downloads\planilha_geral_gravity_COMPLETA.xlsx'
DST = r'C:\Users\danie\Downloads\ddd_campos_final.csv'

NUCLEO_OVERRIDE = {
    'AssinaturaProdutoGravity': 'assinatura',
    'NegociacaoEspecial': 'negociacao_especial',
    'FornecedorOrganizacao': 'organizacao_fornecedor',
    'ProdutoGravityWorkspace': 'produtos_gravity_workspace',
    'MetricasGemini': 'metricas_llm',
    # Short nucleos para casos longos
    'AgendamentoTeste': 'teste',  # usuário usou "teste" em ultima_exec/proxima_exec
}

FK_RENAMES = {
    'tenant_id': 'id_organizacao',
    'company_id': 'id_workspace',
    'user_id': 'id_usuario',
    'product_id': 'id_produto',
    'tenant': 'id_organizacao',
    'user': 'id_usuario',
    'company': 'id_workspace',
}

KEEP_EMPTY = {
    'stripe_subscription_id','stripe_price_id','stripe_customer_id',
    'current_period_start','current_period_end','cancelled_at',
    'USD','EUR','BRL','CNY','JPY','GBP','CHF','ARS','UYU',
    'FOB','CIF','EXW','CFR','FCA','DDP','DAP','CPT','CIP','DPU','FAS',
    'II','IPI','PIS','COFINS','ICMS',
    'AWB','BL','CRT','FCL','LCL',
    'ACTIVE','PAST_DUE','CANCELLED','TRIALING','INCOMPLETE',
    'DRAFT','OPEN','PAID','VOID','OVERDUE','UNCOLLECTIBLE',
}

ABBREV = {
    'exec': 'execucao', 'freq': 'frequencia', 'qtd': 'quantidade',
    'desc': 'descricao', 'prev': 'previsao', 'conf': 'confirmacao',
}

# Renames fixos explícitos pedidos pelo usuário
CUSTOM_RENAMES = {
    ('AssinaturaProdutoGravity','trial_ends_at'): 'teste_encerra_em_assinatura',
}

def snake(pascal):
    s = re.sub(r'(.)([A-Z][a-z]+)', r'\1_\2', pascal)
    s = re.sub(r'([a-z0-9])([A-Z])', r'\1_\2', s)
    return s.lower()

def nucleo_de(tabela):
    return NUCLEO_OVERRIDE.get(tabela, snake(tabela))

def nucleo_pk(tabela):
    """PK sempre usa nome completo, mesmo com override."""
    return snake(tabela)

def expandir(campo):
    parts = campo.split('_')
    return '_'.join(ABBREV.get(p, p) for p in parts)

def gerar_ddd_pg(tabela, atual):
    if not tabela or not atual: return ''
    atual = str(atual).strip()
    tabela = str(tabela).strip()

    if (tabela, atual) in CUSTOM_RENAMES:
        return CUSTOM_RENAMES[(tabela, atual)]
    if atual in KEEP_EMPTY: return ''
    if atual in FK_RENAMES: return FK_RENAMES[atual]
    if atual == 'MetricasGemini': return 'MetricasLLM'
    if atual == tabela: return ''

    nucleo = nucleo_de(tabela)

    if atual == 'id':
        return f'id_{nucleo_pk(tabela)}'
    if atual == 'created_at':
        return f'data_criacao_{nucleo}'
    if atual == 'updated_at':
        return f'data_atualizacao_{nucleo}'
    if atual == 'deleted_at':
        return f'data_exclusao_{nucleo}'

    if atual.endswith('_id'):
        ent = atual[:-3]
        return f'id_{ent}'

    campo_exp = expandir(atual)
    alvo = f'{campo_exp}_{nucleo}'
    if atual == alvo: return ''
    return alvo

def gerar_ddd_tela(atual, tabela):
    """Label humano (col L)."""
    if not atual: return ''
    atual = str(atual).strip()
    if atual in KEEP_EMPTY: return ''
    # PKs e timestamps geralmente ocultos
    if atual in ('id','created_at','updated_at','deleted_at'): return ''
    # FKs têm labels próprios
    labels_fixos = {
        'tenant_id': 'ID da Organização',
        'company_id': 'ID do Workspace',
        'user_id': 'ID do Usuário',
        'product_id': 'ID do Produto',
    }
    if atual in labels_fixos: return labels_fixos[atual]

    # Converte snake para Title Case pt-BR
    parts = expandir(atual).split('_')
    if parts and parts[-1] == 'id':
        parts = ['ID','de'] + parts[:-1]
    # Remove sufixo do nome da tabela se houver
    text = ' '.join(p.capitalize() for p in parts)
    # Acentos comuns
    text = text.replace('Execucao','Execução').replace('Criacao','Criação')
    text = text.replace('Atualizacao','Atualização').replace('Frequencia','Frequência')
    text = text.replace('Ultima','Última').replace('Proxima','Próxima')
    text = text.replace('Configuracao','Configuração').replace('Descricao','Descrição')
    text = text.replace('Organizacao','Organização').replace('Ultimo','Último')
    return text

# Read source
wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb['1. ddd_campos']

# Write CSV with all columns, regenerating F, H, J, L
with open(DST, 'w', encoding='utf-8', newline='') as f:
    writer = csv.writer(f)
    # Header
    header = [ws.cell(1, c).value or '' for c in range(1, 28)]
    writer.writerow(header)

    filled_f = filled_h = filled_j = filled_l = 0
    total = 0
    for r in range(2, ws.max_row+1):
        row = [ws.cell(r, c).value for c in range(1, 28)]
        # Skip fully empty rows
        if not any(row): continue
        total += 1
        tabela = row[1]  # col B
        atual_pg = row[3]     # col D
        atual_prisma = row[4] # col E (source for DDD)

        # Generate DDD columns
        ddd_pg = gerar_ddd_pg(tabela, atual_prisma)
        row[5] = ddd_pg if ddd_pg else ''  # F
        row[7] = ddd_pg if ddd_pg else ''  # H (back)
        row[9] = ddd_pg if ddd_pg else ''  # J (front)

        # L (tela) — gerar se estiver vazia; senão preservar
        if not row[11]:
            row[11] = gerar_ddd_tela(atual_prisma, tabela)

        if row[5]: filled_f += 1
        if row[7]: filled_h += 1
        if row[9]: filled_j += 1
        if row[11]: filled_l += 1

        writer.writerow(['' if v is None else v for v in row])

print(f'Linhas de dados: {total}')
print(f'Col F preenchidas: {filled_f}')
print(f'Col H preenchidas: {filled_h}')
print(f'Col J preenchidas: {filled_j}')
print(f'Col L preenchidas: {filled_l}')
print(f'\nArquivo: {DST}')
