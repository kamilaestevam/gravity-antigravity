"""Gera coluna N (Descrição) do ddd_campos com contexto real e rico."""
import openpyxl, re
from pathlib import Path

ROOT = Path(r'C:\Users\danie\gravity-antigravity')
SRC  = r'C:\Users\danie\Downloads\planilha_geral_gravity_COMPLETA.xlsx'
DST  = r'C:\Users\danie\Downloads\coluna_N_descricao.txt'

FRAGMENTS = [
    'configurador/prisma/schema.prisma',
    'servicos-global/tenant/cadastros/prisma/fragment.prisma',
    'servicos-global/tenant/agendamento/prisma/fragment.prisma',
    'servicos-global/tenant/api-cockpit/prisma/fragment.prisma',
    'servicos-global/tenant/atividades/prisma/fragment.prisma',
    'servicos-global/tenant/conector-erp/prisma/fragment.prisma',
    'servicos-global/tenant/cronometro/prisma/fragment.prisma',
    'servicos-global/tenant/dashboard/prisma/fragment.prisma',
    'servicos-global/tenant/email/prisma/fragment.prisma',
    'servicos-global/tenant/gabi/prisma/fragment.prisma',
    'servicos-global/tenant/historico-global/prisma/fragment.prisma',
    'servicos-global/tenant/ncm-sync/prisma/fragment.prisma',
    'servicos-global/tenant/notificacoes/prisma/fragment.prisma',
    'servicos-global/tenant/preferencias-usuario/prisma/fragment.prisma',
    'servicos-global/tenant/relatorios/prisma/fragment.prisma',
    'servicos-global/tenant/whatsapp/prisma/fragment.prisma',
    'servicos-global/produto/helpdesk/prisma/fragment.prisma',
    'produto/bid-cambio/server/prisma/fragment.prisma',
    'produto/bid-frete/server/prisma/fragment.prisma',
    'produto/financeiro-comex/server/prisma/fragment.prisma',
    'produto/lpco/server/prisma/fragment.prisma',
    'produto/nf-importacao/server/prisma/fragment.prisma',
    'produto/pedido/server/prisma/fragment.prisma',
    'produto/processo/server/prisma/fragment.prisma',
    'produto/simula-custo/server/prisma/fragment.prisma',
]

field_comments = {}  # (tabela, campo) -> comentário Prisma
field_types = {}     # (tabela, campo) -> tipo Prisma

for frag_path in FRAGMENTS:
    fp = ROOT / frag_path
    if not fp.exists(): continue
    lines = fp.read_text(encoding='utf-8').split('\n')
    current_model = None
    for i, line in enumerate(lines):
        stripped = line.strip()
        mm = re.match(r'^model\s+(\w+)\s*\{', stripped)
        if mm:
            current_model = mm[1]
            continue
        if stripped == '}':
            current_model = None
            continue
        if not current_model: continue
        fm = re.match(r'^(\w+)\s+(\S+)(.*)$', stripped)
        if fm and not stripped.startswith('@@') and not stripped.startswith('//'):
            campo, tipo, rest = fm[1], fm[2], fm[3] or ''
            cm = re.search(r'//\s*(.+?)$', rest)
            if cm:
                field_comments[(current_model, campo)] = cm[1].strip()
            field_types[(current_model, campo)] = tipo.strip()

# Descrições contextuais por (tabela, campo)
TABLE_CONTEXT = {
    'Organizacao': 'organização (tenant raiz)',
    'Usuario': 'usuário',
    'Empresa': 'workspace (empresa filial)',
    'Workspace': 'workspace (empresa filial)',
    'AssinaturaProdutoGravity': 'assinatura de produto Gravity',
    'ProdutoGravity': 'produto Gravity (catálogo)',
    'ProdutoGravityWorkspace': 'ativação de produto Gravity no workspace',
    'ConfiguracaoProduto': 'configuração de produto por tenant',
    'UsuarioPermissao': 'permissão granular de usuário',
    'UsuarioWorkspace': 'vínculo do usuário ao workspace',
    'PermissaoAdminGravity': 'permissão de admin interno Gravity',
    'FaixaPreco': 'faixa de preço escalonada por volume',
    'NegociacaoEspecial': 'negociação especial (preços personalizados)',
    'Deploy': 'deploy da plataforma',
    'FornecedorOrganizacao': 'vínculo cross-tenant de fornecedor',
    'Seguranca': 'evento de segurança auditado',
    'Requisicoes': 'métrica de rate limit',
    'Servicos': 'health check de microsserviço',
    'Cambio': 'cotação PTAX oficial do BCB',
    'Testes': 'execução de teste automatizado',
    'AgendamentoTeste': 'agendamento de testes (cron)',
    'PlanoTeste': 'plano de teste documentado',
    'FaturaProdutosGravity': 'fatura de serviços Gravity',
    'MetricasGemini': 'métricas de uso de LLM',
    'Moeda': 'catálogo de moedas',
    'Unidade': 'catálogo de unidades de medida',
    'NCM': 'catálogo NCM (8 dígitos)',
    'OPE': 'operador estrangeiro SISCOMEX',
    'HistoricoStatusOPE': 'histórico de mudança de status do OPE',
    'Agenda': 'agenda configurável (reuniões/consultas)',
    'Slot': 'horário disponível na agenda',
    'Reserva': 'reserva de um slot',
    'DisponibilidadeConfig': 'regras de geração dos slots',
    'TokenAPI': 'token de API do tenant',
    'WebhookConfig': 'configuração de webhook de saída',
    'WebhookLog': 'log de disparo de webhook',
    'LogConsumo': 'log de chamada via token de API',
    'ConexaoErp': 'credenciais ERP do cliente',
    'MapeamentoErp': 'de-para de campos ERP ↔ Gravity',
    'AtividadesDados': 'tarefa/atividade do Kanban',
    'AtividadesParticipantes': 'junção N:N atividade ↔ usuário',
    'AtividadesTempo': 'sessão de tempo registrada',
    'AtividadesCronometro': 'sessão de tempo concluída (histórico)',
    'AtividadesTimer': 'cronômetro em execução',
    'TempoCriacaoRelatorio': 'cache de relatório de tempo',
    'DashboardConfiguracao': 'dashboard (container de widgets)',
    'DashboardCriar': 'widget individual do dashboard',
    'DashboardMetricas': 'cache de KPIs por período',
    'DashboardAlertas': 'regra de alerta sobre dashboard',
    'DashboardCompartilhar': 'link/token de compartilhamento',
    'EmailAssuntosParticipantes': 'thread de e-mail',
    'EmailMensagem': 'mensagem individual de e-mail',
    'EmailRegistroEnvio': 'log de auditoria de envio',
    'TemplateEmail': 'template de e-mail editável',
    'EmailFilaEnvio': 'fila com retry de envio',
    'ConversaCompletaGabi': 'sessão de chat da Gabi',
    'MensagemIndividualGabiai': 'mensagem da conversa Gabi',
    'GabiaLogUso': 'auditoria de ação da Gabi',
    'GabiaTokenConsumidos': 'registro de chamada field-help',
    'GabiaTokenWorkspace': 'quota mensal de tokens por tenant+produto',
    'PersonalizacaoOrganizacaoGabiai': 'persona Gabi por tenant',
    'HistoricoLog': 'log imutável de ação relevante',
    'AlertRule': 'regra de alerta sobre histórico',
    'AlertEvent': 'alerta disparado pela regra',
    'AlertNotificationLog': 'log de notificação de alerta',
    'ExportarResultado': 'resultado de exportação em background',
    'NcmItem': 'cache da tabela NCM',
    'NcmSyncLog': 'log de sincronização NCM',
    'NcmScheduleConfig': 'cron de sync NCM (singleton)',
    'NotificacoesTituloCorpo': 'notificação do sininho',
    'ExternalContact': 'contato externo (e-mail/WhatsApp)',
    'TenantChannelConfig': 'habilitação de canais por tenant',
    'UserPreferences': 'preferências de UI do usuário',
    'RelatoriosSalvos': 'relatório customizado salvo pelo usuário',
    'RelatoriosConfiguracao': 'agendamento de envio de relatório',
    'ExportarJob': 'job de exportação de relatório',
    'WhatsappConversa': 'conversa WhatsApp',
    'WhatsappMensagem': 'mensagem WhatsApp',
    'WhatsappLog': 'log de custo WhatsApp',
    'WhatsappRegra': 'regra de automação WhatsApp',
    'HelpdeskCategoria': 'categoria de ticket do helpdesk',
    'HelpdeskSLA': 'SLA por categoria+prioridade',
    'HelpdeskTicket': 'ticket de suporte',
    'HelpdeskRespostaTemplate': 'template de resposta rápida',
    'Pedido': 'pedido de compra internacional (cabeçalho)',
    'PedidoItem': 'item do pedido',
    'Processo': 'processo COMEX (cabeçalho)',
    'ProcessoFatura': 'fatura comercial vinculada ao processo',
    'ProcessoItem': 'dado logístico por item do processo',
    'ProcessoContainer': 'container utilizado no processo',
    'PedidoStatus': 'status customizável do pedido (abas)',
    'PedidoColuna': 'coluna customizada criada pelo tenant',
    'PedidoPreferenciaUsuario': 'preferência de coluna por usuário',
    'PedidoPreferenciaPadrao': 'preferência de coluna padrão do workspace',
    'AprendizadoImportacaoDados': 'aprendizado IA de mapeamento de planilha',
    'AnexoPedido': 'anexo do pedido (PDF/XLSX/imagem)',
    'TemplatePedidoPdf': 'template de PDF do pedido',
    'TrackingItemsTransferidos': 'rastreamento de itens transferidos',
    'ColunaUsuarioPedido': 'coluna criada pelo usuário (EAV)',
    'ValorColunaUsuarioPedido': 'valor da célula de coluna customizada',
    'KanbanPreferencias': 'preferências Kanban por usuário',
    'DashboardPreferencias': 'preferências do dashboard Pedido',
    'PedidoCasasDecimais': 'configuração de casas decimais por campo',
    'PedidoSaldoFormula': 'fórmula de cálculo do saldo do pedido',
    'DashboardPainel': 'painel/tela dashboard do Pedido',
    'PedidoSnapshotEmpresa': 'snapshot imutável da empresa no pedido',
    'PedidoSnapshotOpe': 'snapshot imutável do OPE no pedido',
    'PedidoConfigAtualizacaoCadastros': 'política de sync Cadastros → Pedido',
    # Produtos externos
    'FinanceiroProcesso': 'aba financeira do processo',
    'FinanceiroLancamento': 'lançamento financeiro (movimentação)',
    'FinanceiroCategorias': 'categoria de despesa',
    'FinanceiroCondicaoPagamento': 'condição de pagamento',
    'FinanceiroNumerario': 'numerário enviado ao despachante',
    'FinanceiroNumerarioDespesa': 'despesa dentro de um numerário',
    'FinanceiroRateio': 'arquivo de rateio gerado',
    'FinanceiroHistorico': 'histórico financeiro imutável',
    # LPCO
    'Lpco': 'Licença/Permissão/Certificado/Outro (LPCO)',
    'LpcoItem': 'item (mercadoria) do LPCO',
    'LpcoExigencia': 'exigência aberta pelo órgão anuente',
    'LpcoVinculo': 'vínculo LPCO ↔ outras entidades',
    'LpcoDocumento': 'documento anexado ao LPCO',
    'LpcoHistorico': 'histórico de mudança no LPCO',
    'SiscomexCredencial': 'credenciais SISCOMEX do tenant',
    # NF Importação
    'NfImportacao': 'Nota Fiscal de Importação (cabeçalho)',
    'NfImportacaoItem': 'item da NF de Importação',
    'NfImportacaoDespesa': 'despesa associada à NF',
    'NfImportacaoRateio': 'rateio de despesa entre itens',
    'NfImportacaoDocumento': 'documento anexado à NF',
    'NfImportacaoHistorico': 'histórico de alteração da NF',
    'DespesaCatalogo': 'catálogo de tipos de despesa',
    'DespesaTemplate': 'template de despesas reutilizável',
    'DespesaTemplateItem': 'item que compõe template de despesa',
    'ExportLayout': 'layout de exportação customizado (XML/Excel)',
    'ExportLayoutCampo': 'campo que compõe layout de exportação',
    'FavoritoFiscal': 'NCM/CFOP/CST favorito para preenchimento rápido',
    # BID
    'ParcelaCambio': 'parcela de contrato de câmbio',
    'AnexoCambio': 'anexo de contrato de câmbio',
    'FormaPagamentoCambio': 'forma de pagamento de câmbio',
    'ConfigParcelaCambio': 'config default de parcelamento',
    'CotacaoCambio': 'cotação de câmbio recebida',
    'BidRequestCambio': 'BID de câmbio (pedido de cotação)',
    'BidResponseCambio': 'resposta da corretora ao BID',
    'Corretora': 'corretora de câmbio',
    'AvaliacaoCorretora': 'avaliação do tenant sobre corretora',
    'RatingCorretora': 'rating global consolidado',
    'SavingCambio': 'economia registrada em contrato',
    'PreferenciaCambio': 'preferências gerais BID Câmbio',
    'PreferenciaGridCambio': 'preferências de layout do grid',
    'Fornecedor': 'fornecedor de frete',
    'Cotacao': 'cotação de frete',
    'BidRequest': 'BID de frete',
    'BidResponse': 'resposta do fornecedor ao BID',
    'DetalheTaxa': 'detalhamento de taxas em proposta',
    'TabelaPreco': 'tabela de preço negociada',
    'Avaliacao': 'avaliação do tenant sobre fornecedor',
    'RatingFornecedor': 'rating global consolidado',
    'Saving': 'economia registrada em fechamento',
    'ConnectorConfig': 'conector automático de cotações',
    'Porto': 'porto/aeroporto',
    # SimulaCusto
    'Estimativa': 'estimativa de custo de importação',
    'TaxaEstimativa': 'taxa aplicada à estimativa',
    'TributoEstimativa': 'tributo calculado na estimativa',
    'DocumentoEstimativa': 'documento anexado à estimativa',
    'CacheAliquota': 'cache de alíquotas por NCM',
    'CacheCambio': 'cache de cotação PTAX',
    'SequenciaEstimativa': 'numeração sequencial',
}

TEMPLATES_CAMPO = {
    'id': 'Identificador único (PK).',
    'tenant_id': 'ID da organização dona do registro (isolamento multi-tenant).',
    'company_id': 'ID do workspace onde o registro existe.',
    'user_id': 'ID do usuário vinculado.',
    'product_id': 'ID do produto Gravity.',
    'created_at': 'Data e hora de criação do registro.',
    'updated_at': 'Data e hora da última alteração.',
    'deleted_at': 'Data de exclusão lógica (soft delete); nulo = ativo.',
    'ativo': 'Indica se o registro está ativo.',
    'nome': 'Nome/rótulo.',
    'titulo': 'Título.',
    'descricao': 'Descrição textual livre.',
    'slug': 'Identificador URL-friendly (kebab-case).',
    'status': 'Status atual (ver enum dedicado).',
    'tipo': 'Tipo/categoria.',
    'email': 'Endereço de e-mail.',
    'telefone': 'Número de telefone.',
    'whatsapp': 'WhatsApp formato E.164.',
    'cnpj': 'CNPJ (14 dígitos).',
    'cpf': 'CPF (11 dígitos).',
    'cidade': 'Cidade.',
    'estado': 'Estado (UF).',
    'pais': 'País (ISO-2).',
    'endereco': 'Endereço completo.',
    'zipcode': 'CEP / código postal.',
    'moeda': 'Moeda (código ISO 4217).',
    'valor': 'Valor monetário.',
    'quantidade': 'Quantidade.',
    'ordem': 'Ordem de exibição.',
}

def desc_contextual(tabela, campo):
    """Descrição template combinada com contexto da tabela."""
    ctx = TABLE_CONTEXT.get(tabela, '')
    tpl = TEMPLATES_CAMPO.get(campo)
    if tpl:
        if ctx and campo in ('id','nome','titulo','descricao','ativo','status','tipo','slug'):
            # Personaliza com contexto
            return tpl.replace('.', f' do/da {ctx}.', 1) if ctx else tpl
        return tpl
    return None

def gerar_descricao(tabela, atual, existing):
    if not tabela or not atual: return existing or ''
    t = str(tabela).strip()
    a = str(atual).strip()
    exist = str(existing).strip() if existing else ''

    # Enum (tabela == atual)
    if t == a:
        return exist or f'Definição do enum {t}.'

    # 1) Template comum + contexto
    desc = desc_contextual(t, a)

    # 2) Se não tem template, tenta comentário Prisma
    if not desc:
        pc = field_comments.get((t, a))
        if pc:
            # Se comentário é curto (lista de valores/formato), envolve com prefixo
            if re.match(r'^[A-Z_]+(\s*\|\s*[A-Z_]+)+', pc) or re.match(r'^[\d\-,\s]+$', pc):
                desc = f'Valores possíveis: {pc}.'
            elif 'ex:' in pc.lower() or 'exemplo' in pc.lower():
                desc = f'Exemplo — {pc}.'
            else:
                desc = pc.rstrip('.') + '.'
        else:
            # 3) Inferência por sufixo
            if a.endswith('_id'):
                ent = a[:-3].replace('_',' ')
                desc = f'Referência (FK) a {ent}.'
            elif a.startswith('id_'):
                ent = a[3:].replace('_',' ')
                desc = f'Referência (FK) a {ent}.'
            elif a.endswith('_at'):
                base = a[:-3].replace('_',' ')
                desc = f'Data e hora de {base}.'
            elif a.endswith('_em'):
                base = a[:-3].replace('_',' ')
                desc = f'Data e hora de {base}.'
            else:
                # Fallback: usa contexto da tabela
                ctx = TABLE_CONTEXT.get(t, '')
                if ctx:
                    desc = f'Campo "{a}" do/da {ctx}.'
                else:
                    desc = f'Campo "{a}".'

    # Enriquece com comentário Prisma se não foi fonte principal
    pc = field_comments.get((t, a))
    if pc and pc not in desc and len(pc) < 120:
        desc = f'{desc} ({pc})'

    # Preserva existente se não-trivial
    junk = {'', 'Campo padrão de auditoria / isolamento.', '—', '-'}
    if exist and exist not in junk and exist not in desc and len(exist) < 200:
        desc = f'{desc} — {exist}'

    return desc

# Processa
wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb['1. ddd_campos']

lines = []
for r in range(2, ws.max_row+1):
    tabela = ws.cell(r,2).value
    atual = ws.cell(r,5).value
    existing = ws.cell(r,14).value
    if not tabela and not atual:
        lines.append('')
        continue
    lines.append(gerar_descricao(tabela, atual, existing))

last = max((i for i,l in enumerate(lines) if l), default=0)+1
lines = lines[:last]

with open(DST, 'w', encoding='utf-8') as f:
    for l in lines:
        f.write(l + '\n')

print(f'Linhas: {len(lines)}, preenchidas: {sum(1 for l in lines if l)}')
print(f'Arquivo: {DST}')
