"""Gera mapa-enums final em CSV, aplicando regra 'vazio quando nao muda'."""
import openpyxl, csv

SRC = r'C:\Users\danie\Downloads\planilha_geral_gravity_COMPLETA.xlsx'
DST = r'C:\Users\danie\Downloads\mapa_enums_final.csv'

# Valores a MANTER em inglês mesmo no DDD (Stripe, ISO, siglas, technical)
KEEP_EN = {
    # Stripe
    'ACTIVE','PAST_DUE','CANCELLED','TRIALING','INCOMPLETE',
    'DRAFT','OPEN','PAID','VOID','OVERDUE','UNCOLLECTIBLE',
    # ISO moedas
    'USD','EUR','BRL','CNY','JPY','GBP','CHF','ARS','UYU',
    # Incoterms
    'FOB','CIF','EXW','CFR','FCA','DDP','DAP','CPT','CIP','DPU','FAS',
    # Siglas tributárias
    'II','IPI','PIS','COFINS','ICMS',
    # Cargo/docs
    'AWB','BL','CRT','FCL','LCL','FTL','LTL',
    # Patentes (mantém em inglês pois mapeam tokens Clerk)
    'SUPER_ADMIN','ADMIN','MASTER','STANDARD','SUPPLIER',
    # Technical terms
    'WEBHOOK','CRON','SERVICE',
    # D0/D1/D2 etc
    'D0','D1','D2',
    # Days
    'DAYS_30','DAYS_90','NEVER','CUSTOM',
    # Dashboard types
    'KPI_CARD','PRODUCT','GENERAL','CATALOG','GABI',
    'LINE','BAR','BAR_HORIZONTAL','DONUT','HISTOGRAM','FUNNEL',
    'GAUGE','MAP','TABLE','AREA',
}

# Traduções PT-BR amigáveis para labels em tela
TRANSLATIONS = {
    'ACTIVE': 'Ativo',
    'ATIVO': 'Ativo', 'ATIVA': 'Ativa',
    'SUSPENDED': 'Suspenso', 'SUSPENSO': 'Suspenso',
    'CANCELLED': 'Cancelado', 'CANCELADO': 'Cancelado', 'CANCELADA': 'Cancelada',
    'PENDING_SETUP': 'Configuração pendente', 'CONFIGURACAO_PENDENTE': 'Configuração pendente',
    'PAST_DUE': 'Vencida', 'VENCIDA': 'Vencida',
    'TRIALING': 'Em teste', 'EM TESTE': 'Em teste', 'EM_TESTE': 'Em teste',
    'INCOMPLETE': 'Incompleta', 'INCOMPLETA': 'Incompleta',
    'INACTIVE': 'Inativo', 'INATIVO': 'Inativo', 'INATIVA': 'Inativa',
    'SUPER_ADMIN': 'Super Admin', 'ADMIN': 'Admin', 'MASTER': 'Master',
    'STANDARD': 'Padrão', 'PADRAO': 'Padrão', 'SUPPLIER': 'Fornecedor', 'FORNECEDOR': 'Fornecedor',
    'SERVICE': 'Serviço', 'SERVICO': 'Serviço', 'WEBHOOK': 'Webhook', 'CRON': 'Cron',
    'COMING_SOON': 'Em breve', 'EM_BREVE': 'Em breve',
    'LEGACY': 'Legado', 'LEGADO': 'Legado',
    'MONTHLY': 'Mensal', 'MENSAL': 'Mensal',
    'PER_PROCESS': 'Por processo', 'POR_PROCESSO': 'Por processo',
    'PER_DOCUMENT': 'Por documento', 'POR_DOCUMENTO': 'Por documento',
    'PER_ESTIMATE': 'Por estimativa', 'POR_ESTIMATIVA': 'Por estimativa',
    'PER_DI_DUIMP': 'Por DI/DUIMP', 'POR_DI_DUIMP': 'Por DI/DUIMP',
    'PER_DUE': 'Por DUE', 'POR_DUE': 'Por DUE',
    'PER_PRODUCT': 'Por produto', 'POR_PRODUTO': 'Por produto',
    'PER_FLOW': 'Por fluxo', 'POR_FLUXO': 'Por fluxo',
    'PER_LPCO': 'Por LPCO', 'POR_LPCO': 'Por LPCO',
    'UNLIMITED': 'Ilimitado', 'ILIMITADO': 'Ilimitado',
    'LIMITED': 'Limitado', 'LIMITADO': 'Limitado',
    'DEVELOPMENT': 'Desenvolvimento', 'DESENVOLVIMENTO': 'Desenvolvimento',
    'STAGING': 'Homologação', 'HOMOLOGACAO': 'Homologação',
    'PRODUCTION': 'Produção', 'PRODUCAO': 'Produção',
    'ALL': 'Todos', 'TODOS': 'Todos',
    'SUCCESS': 'Sucesso', 'SUCESSO': 'Sucesso',
    'FAILED': 'Falhou', 'FALHOU': 'Falhou',
    'ROLLBACK': 'Revertido', 'REVERTIDO': 'Revertido',
    'IN_PROGRESS': 'Em andamento', 'EM_ANDAMENTO': 'Em andamento',
    'DRAFT': 'Rascunho', 'RASCUNHO': 'Rascunho',
    'OPEN': 'Aberto', 'ABERTO': 'Aberto', 'ABERTA': 'Aberta',
    'PAID': 'Pago', 'PAGO': 'Pago',
    'VOID': 'Anulada',
    'OVERDUE': 'Vencida',
    'UNCOLLECTIBLE': 'Incobrável',
    'READ': 'Leitura', 'LEITURA': 'Leitura',
    'WRITE': 'Escrita', 'ESCRITA': 'Escrita',
    'DELETE': 'Exclusão', 'EXCLUSAO': 'Exclusão',
    'NEVER': 'Nunca', 'NUNCA': 'Nunca',
    'DAYS_30': '30 dias', 'DAYS_90': '90 dias', 'CUSTOM': 'Customizado', 'CUSTOMIZADO': 'Customizado',
    'PRODUCT': 'Produto', 'PRODUTO': 'Produto',
    'GENERAL': 'Geral', 'GERAL': 'Geral',
    'CATALOG': 'Catálogo', 'CATALOGO': 'Catálogo',
    'KPI_CARD': 'Card KPI',
    'LINE': 'Linha', 'LINHA': 'Linha',
    'BAR': 'Barra', 'BARRA': 'Barra',
    'BAR_HORIZONTAL': 'Barra horizontal', 'BARRA_HORIZONTAL': 'Barra horizontal',
    'DONUT': 'Rosca', 'ROSCA': 'Rosca',
    'HISTOGRAM': 'Histograma', 'HISTOGRAMA': 'Histograma',
    'FUNNEL': 'Funil', 'FUNIL': 'Funil',
    'GAUGE': 'Medidor', 'MEDIDOR': 'Medidor',
    'MAP': 'Mapa', 'MAPA': 'Mapa',
    'TABLE': 'Tabela', 'TABELA': 'Tabela',
    'AREA': 'Área',
    'PENDENTE': 'Pendente', 'PROCESSANDO': 'Processando',
    'ENVIADO': 'Enviado', 'ENVIADA': 'Enviada',
    'FALHOU': 'Falhou',
    'INBOUND': 'Recebido', 'RECEBIDO': 'Recebido', 'RECEBIDA': 'Recebida',
    'OUTBOUND': 'Enviado',
    'ARQUIVADA': 'Arquivada', 'RESOLVIDA': 'Resolvida',
    'MUITO_POSITIVO': 'Muito positivo', 'POSITIVO': 'Positivo',
    'NEUTRO': 'Neutro', 'NEGATIVO': 'Negativo',
    'MUITO_NEGATIVO': 'Muito negativo',
    'BAIXA': 'Baixa', 'NORMAL': 'Normal', 'ALTA': 'Alta', 'URGENTE': 'Urgente',
    'USUARIO': 'Usuário', 'USER': 'Usuário',
    'API': 'API', 'IA': 'IA', 'AI': 'IA',
    'JOB': 'Job', 'INTEGRATION': 'Integração', 'INTEGRACAO': 'Integração',
    'FAILURE': 'Falha', 'FALHA': 'Falha', 'PARTIAL': 'Parcial', 'PARCIAL': 'Parcial',
    'PENDING': 'Pendente', 'REVIEWED': 'Revisado', 'REVISADO': 'Revisado',
    'ESCALATED': 'Escalado', 'ESCALADO': 'Escalado',
    'RUNNING': 'Executando', 'EXECUTANDO': 'Executando',
    'ERROR': 'Erro', 'ERRO': 'Erro',
    'MANUAL': 'Manual',
    'WAITING_ON_CUSTOMER': 'Aguardando cliente', 'AGUARDANDO_CLIENTE': 'Aguardando cliente',
    'RESOLVED': 'Resolvido', 'RESOLVIDO': 'Resolvido',
    'CLOSED': 'Fechado', 'FECHADO': 'Fechado',
    'LOW': 'Baixa', 'MEDIUM': 'Média', 'MEDIA': 'Média',
    'HIGH': 'Alta', 'URGENT': 'Urgente',
    'IMPORTACAO': 'Importação', 'EXPORTACAO': 'Exportação',
    'PRONTO': 'Pronto', 'FUTURO': 'Futuro',
    'AGENDADO': 'Agendado',
    'ENVIADA_CORRETORAS': 'Enviada às corretoras',
    'EM_COTACAO': 'Em cotação',
    'AGUARDANDO_APROVACAO': 'Aguardando aprovação',
    'APROVADA': 'Aprovada', 'REPROVADA': 'Reprovada',
    'EXPIRADA': 'Expirada', 'EXPIRADO': 'Expirado',
    'EMAIL': 'E-mail', 'PORTAL': 'Portal', 'WHATSAPP': 'WhatsApp',
    'VISUALIZADO': 'Visualizado', 'RESPONDIDO': 'Respondido',
    'ERRO_ENVIO': 'Erro no envio',
    'EM_ANALISE': 'Em análise',
    'MELHOR_TAXA': 'Melhor taxa', 'MELHOR_SPREAD': 'Melhor spread',
    'MELHOR_AVALIACAO': 'Melhor avaliação',
    'MELHOR_PRECO': 'Melhor preço',
    'MELHOR_TRANSIT_TIME': 'Melhor tempo de trânsito',
    'CORRETORA_CAMBIO': 'Corretora de câmbio',
    'BANCO_COMERCIAL': 'Banco comercial',
    'BANCO_CAMBIO': 'Banco de câmbio',
    'FINTECH': 'Fintech',
    'BLOQUEADA': 'Bloqueada', 'BLOQUEADO': 'Bloqueado',
    'DATA_EMBARQUE': 'Data de embarque', 'DATA_CHEGADA': 'Data de chegada',
    'DATA_REGISTRO_DI': 'Data de registro da DI',
    'DATA_DESEMBARACO': 'Data de desembaraço',
    'DATA_ENTREGA': 'Data de entrega',
    'PRONTIDAO_CARGA': 'Prontidão da carga',
    'DATA_FIXA': 'Data fixa',
    'MARITIMO': 'Marítimo', 'AEREO': 'Aéreo', 'RODOVIARIO': 'Rodoviário',
    'AEREO_GERAL': 'Aéreo geral',
    'RODOVIARIO_FTL': 'Rodoviário FTL', 'RODOVIARIO_LTL': 'Rodoviário LTL',
    'ENVIADA_FORNECEDORES': 'Enviada aos fornecedores',
    'FALTA_INFORMACAO': 'Falta informação',
    'AGENTE_CARGA': 'Agente de carga',
    'ARMADOR': 'Armador', 'CIA_AEREA': 'Cia aérea',
    'TRANSPORTADORA': 'Transportadora',
    'PENDENTE_APROVACAO': 'Pendente de aprovação',
    'DIRECIONADA': 'Direcionada',
    'API_REST': 'API REST', 'API_SOAP': 'API SOAP', 'ODATA': 'OData',
    'PEDIDO_COMPRA': 'Pedido de compra',
    'PEDIDO_VENDA': 'Pedido de venda',
    'PROFORMA': 'Proforma', 'INVOICE': 'Invoice',
    'OUTRO': 'Outro',
    # E muitos mais... continua abaixo
}

def translate_label(val):
    """Traduz valor para label amigável em tela."""
    if not val: return ''
    v = str(val).strip()
    if v in TRANSLATIONS: return TRANSLATIONS[v]
    # Default: converte UPPER_SNAKE para Title Case
    parts = v.replace('_',' ').lower().split()
    return ' '.join(p.capitalize() for p in parts)

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb['4. mapa-enums']

header = [ws.cell(1, c).value or '' for c in range(1, 28)]

with open(DST, 'w', encoding='utf-8', newline='') as f:
    writer = csv.writer(f)
    writer.writerow(header)

    for r in range(2, ws.max_row+1):
        row = [ws.cell(r, c).value for c in range(1, 28)]
        if not row[1]:  # sem nome de enum, pula
            continue

        # Col B = enum name Prisma atual
        # Col C = enum name DDD
        # Col E = valor PG atual
        # Col F = valor PG DDD
        # Col G = valor Prisma atual
        # Col H = valor DDD (novo, universal)
        # Cols I/J = back atual/DDD
        # Cols K/L = front atual/DDD
        # Cols M/N = label em tela atual/DDD

        b = str(row[1] or '').strip()
        c_ddd = str(row[2] or '').strip()
        e_pg = str(row[4] or '').strip()
        f_pg_ddd = str(row[5] or '').strip()
        g_prisma = str(row[6] or '').strip()

        # Regra 1: C vazio se C == B
        if c_ddd and c_ddd == b:
            row[2] = None

        # Regra 2: F vazio se F == E
        if f_pg_ddd and f_pg_ddd == e_pg:
            row[5] = None
            f_pg_ddd = ''

        # Regra 3: H (Valor - DDD) espelha F (mesmo valor)
        # Se F é diferente de E, preenche H também
        if f_pg_ddd:
            row[7] = f_pg_ddd  # col H
        else:
            row[7] = None

        # Regra 4: Back atual/DDD (cols I/J)
        # Formato: <EnumName>.<VALUE>
        enum_atual = b
        enum_ddd = c_ddd if c_ddd else b  # se não tem rename, usa o atual
        val_atual = g_prisma
        val_ddd = f_pg_ddd if f_pg_ddd else g_prisma

        row[8] = f'{enum_atual}.{val_atual}'  # col I (back atual)
        if (enum_ddd != enum_atual) or (val_ddd != val_atual):
            row[9] = f'{enum_ddd}.{val_ddd}'  # col J (back DDD)
        else:
            row[9] = None

        # Regra 5: Front = igual ao back (compartilha TypeScript types)
        row[10] = f'{enum_atual}.{val_atual}'  # col K (front atual)
        if (enum_ddd != enum_atual) or (val_ddd != val_atual):
            row[11] = f'{enum_ddd}.{val_ddd}'  # col L (front DDD)
        else:
            row[11] = None

        # Regra 6: Label em tela (cols M/N)
        # Atual = tradução do valor atual
        # DDD = tradução do valor DDD
        label_atual = translate_label(val_atual)
        label_ddd = translate_label(val_ddd)
        row[12] = label_atual  # col M
        if label_atual != label_ddd:
            row[13] = label_ddd  # col N
        else:
            row[13] = None

        writer.writerow(['' if v is None else v for v in row])

print(f'Arquivo: {DST}')
