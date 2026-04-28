"""Gera mapa-paginas final com col C (URL DDD), D (Explicação NOVA),
F (Titulo DDD), G (Arquivo DDD), I (Componente DDD)."""
import openpyxl, csv, re

SRC = r'C:\Users\danie\Downloads\planilha_geral_gravity_COMPLETA.xlsx'
DST = r'C:\Users\danie\Downloads\mapa_paginas_final.csv'

# Renames DDD de URL
URL_REPLACEMENTS = [
    (r'\btenants\b', 'organizacao'),
    (r'\bworkspaces\b', 'workspace'),
    (r'\bcompanies\b', 'workspace'),
    (r'\btenant-products\b', 'organizacao-produtos'),
    (r':tenantId\b', ':id_organizacao'),
    (r':companyId\b', ':id_workspace'),
    (r':userId\b', ':id_usuario'),
    (r':productKey\b', ':slug_produto'),
]

# Renames DDD de arquivo/componente
COMP_RENAMES = {
    'MetricasGeminiAdmin': 'MetricasLLMAdmin',
    'TenantDetail': 'OrganizacaoDetalhe',
    'TestesGeraisAdmin': 'TesteGeralAdmin',
    'ProcessoLayout_2': '— (duplicata, investigar)',
    'EmpresasParceiros': 'EmpresasCadastros',
}

# Explicações por componente (conhecidos)
EXPLICACOES = {
    # Configurador Admin
    'AdminLayout': 'Layout compartilhado das telas administrativas (sidebar + outlet).',
    'ApiCockpitAdmin': 'Administração global do API Cockpit — serviços, tokens, webhooks, logs.',
    'CadastrosGlobaisAdmin': 'Administração dos catálogos globais (moedas, unidades, NCM).',
    'DeployAdmin': 'Histórico de deploys da plataforma Gravity.',
    'FinanceiroAdmin': 'Faturas emitidas para tenants (serviços Gravity).',
    'HistoricoGlobalAdmin': 'Log global imutável de ações no sistema.',
    'LogTestes': 'Log de execuções de testes automatizados.',
    'MetricasGeminiAdmin': 'Métricas agregadas de uso de LLM (custo, tokens, latência).',
    'ModalAgendamentoNcmSync': 'Modal para configurar o cron de sincronização NCM.',
    'ModalAgendamentoTestes': 'Modal para configurar o agendamento de testes automatizados.',
    'ModalEditarOrganizacao': 'Modal para editar dados da organização (tenant).',
    'ModalExecutarTestes': 'Modal para disparar execução manual de testes.',
    'ModalNovaOrganizacao': 'Modal para criar nova organização (novo tenant).',
    'NcmIntegracaoAdmin': 'Administração da integração NCM com Portal Único.',
    'PlanosTesteAdmin': 'Administração dos planos de teste documentados.',
    'ProdutosGravityAdmin': 'Catálogo master dos produtos da plataforma.',
    'SegurancaAdmin': 'Eventos de segurança, rate limits e secrets.',
    'TestesGeraisAdmin': 'Painel geral de testes automatizados (E2E, unitários, etc.).',
    'UsuariosGlobaisAdmin': 'Administração dos usuários globais (Super Admin, Admins Gravity).',
    'VisaoGeralAdmin': 'Dashboard consolidado da administração Gravity.',
    # Configurador Workspace
    'AdminPanel': 'Painel administrativo do workspace.',
    'AuthPage': 'Página de autenticação (login/signup via Clerk).',
    'Contato': 'Página de contato público.',
    'CoreDashboard': 'Dashboard principal do usuário logado.',
    'Core': 'Tela principal (core) do Configurador.',
    'E2ENotificacoesHarness': 'Harness de teste E2E para notificações.',
    'Hub': 'Hub de produtos disponíveis para o workspace.',
    'Onboarding': 'Fluxo de onboarding inicial (pós-cadastro/trial).',
    'SelecionarWorkspace': 'Tela de seleção de workspace quando usuário tem múltiplos.',
    'Store': 'Loja de produtos adicionais do tenant.',
    'TenantDetail': 'Detalhes de uma organização (tenant) específica.',
    'Waitlist': 'Lista de espera para trials indisponíveis.',
    'ApiCockpit': 'API Cockpit do tenant — tokens, webhooks, logs.',
    'Assinaturas': 'Gestão de assinaturas do tenant.',
    'ConectorCargoWise': 'Configuração do conector CargoWise (ERP).',
    'Conectores': 'Lista de conectores ERP disponíveis.',
    'DocPortal': 'Portal de documentação interna do tenant.',
    'EmpresasParceiros': 'Gestão de empresas parceiras do workspace.',
    'Financeiro': 'Visão financeira do tenant (faturas, pagamentos).',
    'HistoricoOrganizacao': 'Histórico de ações na organização.',
    'ModalEditarAssinatura': 'Modal para editar uma assinatura de produto.',
    'ModalEditarEmpresa': 'Modal para editar dados de uma empresa.',
    'ModalEditarUsuario': 'Modal para editar dados de um usuário.',
    'ModalEditarWorkspace': 'Modal para editar dados do workspace.',
    'ModalExclusao': 'Modal genérico de confirmação de exclusão.',
    'ModalPermissoesUsuario': 'Modal de gestão de permissões granulares de um usuário.',
    'Organizacao': 'Página principal da organização (tenant).',
    'TabelaUsuarios': 'Componente de tabela de usuários do workspace.',
    'TabelaWorkspaces': 'Componente de tabela de workspaces da organização.',
    'TaxaCambio': 'Cotações PTAX do BCB.',
    'Usuarios': 'Lista de usuários do workspace.',
    'WorkspaceLayout': 'Layout compartilhado das telas de workspace.',
    'Workspaces': 'Lista de workspaces da organização.',
    # Marketplace
    'Checkout': 'Checkout de compra de um produto (público).',
    'Home': 'Página inicial do marketplace público.',
    'Precos': 'Página de preços dos produtos.',
    'ProdutoDetalhe': 'Página de detalhes de um produto do catálogo.',
    'SimuladorComex': 'Simulador público de custo COMEX (sem login).',
    'Produtos': 'Listagem pública de produtos do marketplace.',
    'Trial': 'Página de início de teste gratuito.',
    'DashboardGeralPage': 'Dashboard consolidado cross-produto.',
    # bid-cambio
    'Comparativo': 'Tela de comparação de propostas recebidas.',
    'Configuracoes': 'Página de configurações do produto.',
    'Corretoras': 'Lista de corretoras de câmbio cadastradas.',
    'Dashboard': 'Visão geral (KPIs) do produto.',
    'DetalheCorretora': 'Detalhes de uma corretora específica.',
    'DetalheCotacao': 'Detalhes de uma cotação de câmbio.',
    'ListaCambios': 'Lista de contratos de câmbio do workspace.',
    'ModalPagamento': 'Modal de registro de pagamento de parcela.',
    'NovaCotacao': 'Formulário de nova cotação de câmbio.',
    'ConfigCorretora': 'Configuração do portal da corretora.',
    'CotacoesPendentes': 'Cotações pendentes de resposta (portal).',
    'MeuDesempenho': 'Desempenho da corretora/fornecedor (portal).',
    'MinhasRespostas': 'Respostas enviadas pela corretora (portal).',
    'PortalDashboard': 'Dashboard do portal da corretora.',
    'ResponderCotacao': 'Formulário para responder a um BID.',
    'ResponderPublico': 'Responder a um BID via link público (sem login).',
    # bid-frete
    'Cotacoes': 'Lista de cotações de frete recebidas.',
    'DetalheFornecedor': 'Detalhes de um fornecedor de frete.',
    'Fornecedores': 'Lista de fornecedores cadastrados.',
    'ImportarBloco': 'Importação em massa (bloco) de cotações.',
    'TabelaPrecos': 'Tabela de preços negociada com fornecedor.',
    # financeiro-comex
    'CategoriasPage': 'Gestão de categorias de despesa.',
    'CondicoesPagamentoPage': 'Gestão de condições de pagamento.',
    'ModalHistorico': 'Modal com histórico de alterações.',
    'ModalImportar': 'Modal de importação de lançamentos.',
    'ModalNovoLancamento': 'Modal de criação de lançamento financeiro.',
    'MovimentacaoPage': 'Lançamentos financeiros do processo.',
    'ModalExibirAnexo': 'Modal para visualizar anexo.',
    'ModalInserirNumerario': 'Modal para inserir numerário (adiantamento).',
    'NumerarioPage': 'Gestão de numerários por processo.',
    'RateioPage': 'Rateio de despesas entre itens.',
    # lpco
    'LpcoDetalhe': 'Detalhes de um LPCO (Licença/Permissão/Certificado).',
    'LpcoLista': 'Lista de LPCOs do workspace.',
    'LpcoNovo': 'Formulário de criação de novo LPCO (multi-step).',
    'LpcoSimulador': 'Simulador de LPCO (cálculo prévio).',
    # nf-importacao
    'DespesaCatalogo': 'Catálogo de tipos de despesa reutilizáveis.',
    'DespesaTemplate': 'Templates de despesas para reuso.',
    'ExportLayout': 'Layouts customizados de exportação (XML/Excel).',
    'FavoritosFiscais': 'NCMs/CFOPs/CSTs marcados como favoritos.',
    'NfDetalhe': 'Detalhes de uma NF de Importação.',
    'NfLista': 'Lista de NFs de Importação.',
    'NfNovaDespesas': 'Etapa Despesas da criação de NF.',
    'NfNovaDuimp': 'Etapa DUIMP da criação de NF.',
    'NfNovaExportacao': 'Etapa Exportação da criação de NF.',
    'NfNovaFiscal': 'Etapa Fiscal da criação de NF.',
    'NfNovaOrigem': 'Etapa inicial (Origem) da criação de NF.',
    'NfNovaRateio': 'Etapa Rateio da criação de NF.',
    # pedido
    'MatrizSnapshotCadastros': 'Matriz de snapshot entre Cadastros e Pedido.',
    'DashboardPedido': 'Dashboard do produto Pedido.',
    'Historico': 'Histórico de alterações no pedido.',
    'ImportarArquivo': 'Importação de pedidos via arquivo (planilha).',
    'KanbanPedidos': 'Visualização Kanban de pedidos (agrupado por status).',
    'ListaPedidos': 'Lista/tabela de pedidos.',
    'NovoPedido': 'Formulário de criação/edição de pedido.',
    'SecaoKanbanColunas': 'Seção de configuração de colunas do Kanban.',
    # processo (legado)
    'DadosTecnicosPage': 'Dados técnicos do processo.',
    'EmailPage': 'E-mails do processo.',
    'PedidosPage': 'Pedidos vinculados ao processo.',
    'ProcessoLayout': 'Layout do processo (abas internas).',
    'ProcessoLayout_2': '⚠️ Duplicata suspeita — investigar e excluir.',
    'WorkflowPage': 'Workflow (etapas/marcos) do processo.',
    # simula-custo
    'DashboardSimulaCusto': 'Dashboard do SimulaCusto.',
    'Estimativas': 'Lista de estimativas de custo.',
    'EstimativasDashboard': 'Dashboard de estimativas.',
    'ModalSimulacao': 'Modal de simulação rápida de custo.',
    'ImportarMassa': 'Importação em massa de estimativas.',
    'Relatorios': 'Relatórios do SimulaCusto.',
}

# Títulos amigáveis (col F)
def gerar_titulo(comp_name):
    """Gera título de página a partir do nome do componente."""
    if not comp_name: return ''
    # Remove sufixos
    base = comp_name.replace('Page','').replace('Admin','').replace('Modal','')
    # Split camel → espaços
    s = re.sub(r'(.)([A-Z][a-z]+)', r'\1 \2', base)
    s = re.sub(r'([a-z0-9])([A-Z])', r'\1 \2', s)
    # Capitaliza e adiciona acentos comuns
    titulo = s.strip()
    replacements = {
        'Configuracoes': 'Configurações', 'Atividades': 'Atividades',
        'Historico': 'Histórico', 'Usuarios': 'Usuários',
        'Organizacao': 'Organização', 'Metricas': 'Métricas',
        'Precos': 'Preços', 'Numerario': 'Numerário',
        'Categorias': 'Categorias', 'Condicoes': 'Condições',
        'Exigencias': 'Exigências', 'Relatorios': 'Relatórios',
        'Conectores': 'Conectores',
    }
    for old, new in replacements.items():
        titulo = titulo.replace(old, new)
    return titulo

def apply_ddd_url(url):
    if not url: return ''
    ddd = url
    for pattern, repl in URL_REPLACEMENTS:
        ddd = re.sub(pattern, repl, ddd)
    return ddd if ddd != url else ''

def gerar_explicacao(comp_name, url, arquivo):
    """Retorna a explicação pré-definida ou infere pelo padrão."""
    if comp_name in EXPLICACOES:
        return EXPLICACOES[comp_name]
    # Inferência por sufixo
    if comp_name and comp_name.endswith('Layout'):
        return 'Layout compartilhado (wrapper com navegação/outlet).'
    if comp_name and comp_name.endswith('Modal'):
        return 'Modal — ver lista de modais para detalhes.'
    if comp_name and comp_name.endswith('Dashboard'):
        return f'Dashboard de {comp_name.replace("Dashboard","")}.'
    if comp_name and 'Lista' in comp_name:
        entity = comp_name.replace('Lista','')
        return f'Lista de {entity.lower()}.'
    if comp_name and 'Detalhe' in comp_name:
        entity = comp_name.replace('Detalhe','')
        return f'Detalhes de {entity.lower()}.'
    if comp_name and comp_name.startswith('Nova'):
        entity = comp_name[4:]
        return f'Formulário de nova {entity.lower()}.'
    if comp_name and comp_name.startswith('Novo'):
        entity = comp_name[4:]
        return f'Formulário de novo {entity.lower()}.'
    return f'Página {comp_name}.'

# Processa
wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb['6. mapa-paginas']

header_atual = [ws.cell(1,c).value or '' for c in range(1, 23)]
# Insere col "Explicação" após C (URL DDD)
new_header = header_atual[:3] + ['Explicação'] + header_atual[3:]

rows_out = [new_header]
for r in range(2, ws.max_row+1):
    row = [ws.cell(r,c).value for c in range(1, 23)]
    if not row[0]: continue

    url_atual = str(row[1] or '').strip()
    url_ddd = str(row[2] or '').strip()
    titulo_atual = row[3]
    titulo_ddd = row[4]
    arquivo = str(row[5] or '').strip()
    arquivo_ddd = row[6]
    componente = str(row[7] or '').strip()
    comp_ddd = row[8]

    # Col C (URL DDD): regra vazio se não muda
    calc_url_ddd = apply_ddd_url(url_atual)
    if calc_url_ddd:
        row[2] = calc_url_ddd
    else:
        row[2] = None

    # Col F (Titulo DDD) — gera se estiver vazia
    if not titulo_ddd:
        t_gerado = gerar_titulo(componente)
        if t_gerado and t_gerado != str(titulo_atual or '').strip():
            row[4] = t_gerado

    # Col G (Arquivo DDD) — só se comp tem rename
    if not arquivo_ddd and componente in COMP_RENAMES:
        new_comp = COMP_RENAMES[componente]
        if not new_comp.startswith('—'):
            row[6] = arquivo.replace(componente, new_comp) if componente in arquivo else f'{new_comp}.tsx'

    # Col I (Componente DDD)
    if not comp_ddd and componente in COMP_RENAMES:
        new_comp = COMP_RENAMES[componente]
        row[8] = new_comp

    # Explicação (col D nova)
    explicacao = gerar_explicacao(componente, url_atual, arquivo)

    # Monta linha com D inserida após C
    new_row = row[:3] + [explicacao] + row[3:]
    rows_out.append(new_row)

with open(DST, 'w', encoding='utf-8', newline='') as f:
    writer = csv.writer(f)
    for r in rows_out:
        writer.writerow(['' if v is None else v for v in r])

print(f'Linhas: {len(rows_out)-1}')
print(f'Arquivo: {DST}')

# Stats
filled_c = sum(1 for r in rows_out[1:] if r[2])
filled_d = sum(1 for r in rows_out[1:] if r[3])
filled_f = sum(1 for r in rows_out[1:] if r[5])
filled_i = sum(1 for r in rows_out[1:] if r[9])
print(f'Col C (URL DDD): {filled_c}')
print(f'Col D (Explicação): {filled_d}')
print(f'Col F (Titulo DDD): {filled_f}')
print(f'Col I (Componente DDD): {filled_i}')
