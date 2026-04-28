"""Gera mapa-rotas final com col C (Rota DDD) + col D (Explicação)."""
import openpyxl, csv, re

SRC = r'C:\Users\danie\Downloads\planilha_geral_gravity_COMPLETA.xlsx'
DST = r'C:\Users\danie\Downloads\mapa_rotas_final.csv'

# Regras DDD para transformação de URL
URL_REPLACEMENTS = [
    # Palavras inteiras
    (r'\btenants\b', 'organizacao'),
    (r'\btenant-products\b', 'organizacao-produtos'),
    (r'\bworkspaces\b', 'workspace'),
    (r'\bcompanies\b', 'workspace'),
    (r'\bproducts\b', 'produtos'),
    (r'\bservice-token\b', 'organizacao-token'),
    # Verbos
    (r'/activate\b', '/ativar'),
    (r'/deactivate\b', '/desativar'),
    (r'/subscribe\b', '/assinar'),
    (r'/sync\b', '/sincronizar'),
    # Params
    (r':tenantId\b', ':id_organizacao'),
    (r':companyId\b', ':id_workspace'),
    (r':productKey\b', ':slug_produto'),
    (r':userId\b', ':id_usuario'),
    # :key especial — precisa inspeção, mas na maioria é :id_organizacao
    # :id permanece como está
]

# Arquivo → pasta/prefixo de rota inferido
FILE_TO_PREFIX = {
    'atividades.ts': '/api/v1/atividades',
    'contatos.ts': '/api/v1/contatos',
    'empresas.ts': '/api/v1/empresas',
    'kanban.ts': '/api/v1/kanban',
    'pipelines.ts': '/api/v1/pipelines',
    'agenda.ts': '/api/v1/agenda',
    'config.ts': '/api/v1/config',
    'reserva.ts': '/api/v1/reserva',
    'slot.ts': '/api/v1/slot',
}

# Paths completos de rotas fantasma (usa full path, não só filename)
GHOST_FULL_PATHS = {
    'servicos-global/tenant/atividades/server/routes/pipelines.ts',
    'servicos-global/tenant/atividades/server/routes/kanban.ts',
    'servicos-global/tenant/atividades/server/routes/contatos.ts',
    'servicos-global/tenant/atividades/server/routes/empresas.ts',
    'servicos-global/tenant/api-cockpit/server/src/routes/observability.ts',
    'servicos-global/configurador/server/routes/serviceToken.ts',
    'servicos-global/organizacao/pedido/server/src/routes/dashboardWidgets.ts',
}

# Entidade a partir do path (ou do nome do arquivo como fallback)
def extract_entity(path, arquivo=''):
    """Extrai entidade principal do path, ou infere pelo arquivo."""
    parts = [p for p in (path or '').split('/') if p and not p.startswith(':')]
    skip = {'api','v1','v2','internal','admin','cockpit','portal'}
    for p in parts:
        if p not in skip: return p
    # Fallback: nome do arquivo
    if arquivo:
        fname = arquivo.replace('\\','/').split('/')[-1]
        fname = fname.replace('.ts','').replace('.js','')
        return fname
    return parts[-1] if parts else 'recurso'

def apply_ddd_url(url):
    """Transforma URL atual em versão DDD."""
    if not url or url == '/': return ''
    ddd = url
    for pattern, replacement in URL_REPLACEMENTS:
        ddd = re.sub(pattern, replacement, ddd)
    # Checa se mudou
    return ddd if ddd != url else ''

def gerar_rota_completa(path_atual, arquivo, metodo):
    """Se path_atual == '/', tenta inferir a partir do arquivo."""
    if path_atual and path_atual != '/':
        return path_atual
    if not arquivo: return path_atual
    fname = arquivo.rstrip('/').split('/')[-1]
    prefix = FILE_TO_PREFIX.get(fname)
    if prefix:
        return prefix
    return path_atual

def gerar_explicacao(metodo, rota, arquivo):
    """Gera explicação com base em método + padrão de URL."""
    if not rota or not metodo: return ''

    # Rotas fantasma: marca explicitamente (usa full path)
    arq_norm = (arquivo or '').replace('\\','/').lstrip('./')
    if arq_norm in GHOST_FULL_PATHS:
        return '⚠️ Rota fantasma — modelo Prisma não existe, candidata a exclusão.'

    entity = extract_entity(rota, arquivo)

    # Padrões comuns
    has_id = ':id' in rota or any(f':{k}' in rota for k in ('userId','companyId','tenantId','processoId','pedidoId','widgetId','suid','codigo','jobId','token'))
    last_segment = rota.rstrip('/').split('/')[-1]

    # Ações custom (não CRUD puro)
    action_verbs = {
        'ativar','desativar','assinar','cancelar','aprovar','reprovar','confirmar',
        'preview','sincronizar','test','testar','promote','invite','execute',
        'duplicar','consolidar','enviar','reverter','exportar','importar',
        'stream','download','send','void','apply-fix','reanalyze','generate',
        'expand','generate-spec','extract-testids','extract','analisar',
        'read','read-all','dismiss'
    }
    if last_segment in action_verbs:
        return f'Executa ação "{last_segment}" sobre {entity}.'

    if rota.endswith('/health'):
        return 'Health check do serviço.'
    if rota.endswith('/stream'):
        return f'Stream de eventos (SSE) de {entity}.'
    if rota.endswith('/stats'):
        return f'Estatísticas agregadas de {entity}.'
    if rota.endswith('/kpis'):
        return f'KPIs (indicadores-chave) de {entity}.'

    # CRUD padrão
    m = metodo.upper()
    if m == 'GET':
        if has_id:
            return f'Busca um {entity} pelo identificador.'
        return f'Lista {entity} do tenant (com filtros/paginação).'
    if m == 'POST':
        if has_id:
            return f'Executa ação sobre {entity} específico.'
        return f'Cria novo {entity}.'
    if m in ('PUT', 'PATCH'):
        if has_id:
            return f'Atualiza {entity} existente pelo identificador.'
        return f'Atualiza dados de {entity}.'
    if m == 'DELETE':
        return f'Remove {entity} pelo identificador.'

    return f'{m} em {rota}.'

# Processa
wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb['5. mapa-rotas']

header_atual = [ws.cell(1,c).value or '' for c in range(1, 25)]
# Inserir col 'Explicação' na posição 4 (após C = Rota DDD)
new_header = header_atual[:3] + ['Explicação'] + header_atual[3:]

rows_out = [new_header]
for r in range(2, ws.max_row+1):
    row = [ws.cell(r,c).value for c in range(1,25)]
    if not row[1]: continue

    metodo = str(row[0] or '').strip()
    rota_atual = str(row[1] or '').strip()
    arquivo = str(row[7] or '').strip() if len(row) > 7 else ''

    # Se "/" tenta inferir pela pasta/arquivo
    rota_completa = gerar_rota_completa(rota_atual, arquivo, metodo)
    if rota_completa != rota_atual:
        row[1] = rota_completa
        rota_atual = rota_completa

    # Col C = Rota DDD
    rota_ddd = apply_ddd_url(rota_atual)
    row[2] = rota_ddd if rota_ddd else None

    # Col D nova = Explicação
    explicacao = gerar_explicacao(metodo, rota_atual, arquivo)

    # Monta linha final com D inserida
    new_row = row[:3] + [explicacao] + row[3:]
    rows_out.append(new_row)

# Escreve CSV
with open(DST, 'w', encoding='utf-8', newline='') as f:
    writer = csv.writer(f)
    for r in rows_out:
        writer.writerow(['' if v is None else v for v in r])

print(f'Linhas: {len(rows_out)-1}')
print(f'Arquivo: {DST}')

# Stats
filled_c = sum(1 for r in rows_out[1:] if r[2])
filled_d = sum(1 for r in rows_out[1:] if r[3])
ghost = sum(1 for r in rows_out[1:] if '⚠️' in str(r[3] or ''))
print(f'Col C (Rota DDD) preenchidas: {filled_c}')
print(f'Col D (Explicação) preenchidas: {filled_d}')
print(f'Rotas fantasma marcadas: {ghost}')
