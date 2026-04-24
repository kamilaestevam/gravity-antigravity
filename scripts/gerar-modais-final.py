"""Gera mapa-modais final com col C (arquivo DDD), D (Explicação NOVA),
F (componente DDD), e corrige col E (componente)."""
import openpyxl, csv, re

SRC = r'C:\Users\danie\Downloads\planilha_geral_gravity_COMPLETA.xlsx'
DST = r'C:\Users\danie\Downloads\mapa_modais_final.csv'

# Renames DDD conhecidos
COMP_RENAMES = {
    'ModalEditarTenant': 'ModalEditarOrganizacao',
    # arquivos já em DDD, não precisam rename
}

# Explicações por arquivo/componente
EXPLICACOES = {
    # Configurador admin
    'ModalAgendamentoNcmSync': 'Modal para configurar o cron de sincronização NCM com o Portal Único.',
    'ModalAgendamentoTestes': 'Modal para configurar o agendamento (cron) dos testes automatizados.',
    'ModalEditarOrganizacao': 'Modal para editar dados da organização (tenant) — nome, CNPJ, contatos.',
    'ModalExecutarTestes': 'Modal para disparar execução manual de um plano de testes.',
    'ModalNovaOrganizacao': 'Modal para criar nova organização (tenant) no sistema.',
    # Configurador workspace
    'ModalEditarAssinatura': 'Modal para editar uma assinatura de produto Gravity do tenant.',
    'ModalEditarEmpresa': 'Modal para editar dados de uma empresa (workspace).',
    'ModalEditarUsuario': 'Modal para editar dados de um usuário e sua patente.',
    'ModalEditarWorkspace': 'Modal para editar dados do workspace (filial).',
    'ModalExclusao': 'Modal genérico de confirmação de exclusão — pede confirmação textual.',
    'ModalPermissoesUsuario': 'Modal de gestão granular de permissões por usuário × produto × workspace.',
    # Marketplace
    'ExitIntentDrawer': 'Drawer que abre ao detectar intenção de saída do site (exit intent).',
    'PaywallDrawer': 'Drawer de paywall que bloqueia conteúdo premium para usuários não assinantes.',
    # bid-cambio
    'ModalPagamento': 'Modal para registrar o pagamento de uma parcela de contrato de câmbio.',
    # financeiro-comex
    'ModalHistorico': 'Modal com histórico de alterações do lançamento/numerário.',
    'ModalImportar': 'Modal de importação de lançamentos financeiros (XML DUIMP, planilha, SmartRead).',
    'ModalNovoLancamento': 'Modal de criação de novo lançamento financeiro no processo.',
    'ModalNovoLancamento.test': '⚠️ Arquivo de teste — não é modal de produção.',
    'ModalExibirAnexo': 'Modal para visualizar um anexo (PDF/imagem) em tela cheia.',
    'ModalInserirNumerario': 'Modal para registrar numerário (adiantamento) ao despachante.',
    # pedido
    'ModalNovaColuna': 'Modal para criar coluna customizada do usuário na tabela Pedido.',
    'DrawerPedido': 'Drawer lateral com detalhes/ações do pedido selecionado.',
    'FiltroPopoverColuna': 'Popover para filtrar valores de uma coluna específica.',
    'ModalConsolidar': 'Modal para consolidar múltiplos pedidos em um único (preview + confirmar).',
    'ModalDuplicar': 'Modal para duplicar um pedido (escolher campos a copiar).',
    'ModalDuplicarItens': 'Modal para duplicar itens entre pedidos.',
    'ModalEdicaoEmMassa': 'Modal de edição em massa de múltiplos pedidos (preview + confirmar).',
    'ModalGerarPdf': 'Modal para gerar PDF do pedido com template selecionado.',
    'ModalNovoItem': 'Modal para adicionar novo item ao pedido.',
    'ModalNovoPedido': 'Modal para criar novo pedido (header + itens).',
    'ModalTransferir': 'Modal para transferir itens entre pedidos (preview + confirmar).',
    'SmartImportModal': 'Modal de importação inteligente (IA mapeia colunas de planilha).',
    # simula-custo
    'ModalSimulacao': 'Modal de simulação rápida de custo COMEX.',
    # nucleo-global
    'ModalBuscaNcm': 'Modal de busca e seleção de NCM do catálogo.',
    'WidgetEditModal': 'Modal para editar configurações de um widget do dashboard.',
    'CardKanbanModal': 'Modal que abre um card do Kanban em detalhes.',
    'ModalEnviarParaGlobal': 'Modal para enviar um item local para o contexto global.',
    'modal-formulario-abas-global': 'Modal com formulário tabbed global (múltiplas abas).',
    'modal-formulario-global': 'Modal com formulário global simples.',
    'ModalFormularioAbasGlobal': 'Modal com formulário tabbed global (múltiplas abas).',
    'ModalFormularioGlobal': 'Modal com formulário global simples.',
    'ModalGabiCaixaAviso': 'Modal de aviso da Gabi (caixa de informação).',
    'modal-overlay': 'Componente de overlay base para modais.',
}

def gerar_explicacao_fallback(fname, comp, tipo):
    """Gera explicação quando não tem no dicionário."""
    base = fname.replace('.tsx','').replace('.jsx','')
    # Remove sufixos de teste
    if '.test' in base:
        return '⚠️ Arquivo de teste — não é modal de produção.'
    # Tipo básico
    if tipo == 'Drawer': return f'Drawer lateral para {comp or base}.'
    if tipo == 'Popover': return f'Popover contextual para {comp or base}.'
    if tipo == 'Dialog': return f'Diálogo modal para {comp or base}.'
    # Tenta inferir pela ação
    name = base.replace('Modal','').strip()
    s = re.sub(r'(.)([A-Z][a-z]+)', r'\1 \2', name)
    s = re.sub(r'([a-z0-9])([A-Z])', r'\1 \2', s)
    return f'Modal para {s.lower()}.'

def fix_component_name(fname, current):
    """Se current não parece nome de componente (é constante), usa filename."""
    if not current: return fname.replace('.tsx','').replace('.jsx','')
    # Se é tudo maiúsculas ou snake_case, provavelmente é constante
    if current.isupper() or (current.islower() and '_' in current):
        return fname.replace('.tsx','').replace('.jsx','')
    # Se é camelCase (começa minúsculo), provavelmente é função helper
    if current[0].islower():
        return fname.replace('.tsx','').replace('.jsx','')
    return current

# Processa
wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb['7. Modais']

header_atual = [ws.cell(1,c).value or '' for c in range(1, 16)]
# Insere col "Explicação" após C (arquivo DDD)
new_header = header_atual[:3] + ['Explicação'] + header_atual[3:]

rows_out = [new_header]
for r in range(2, ws.max_row+1):
    row = [ws.cell(r,c).value for c in range(1, 16)]
    if not row[1]: continue

    fname = str(row[1] or '').strip()
    arquivo_ddd = row[2]
    comp_atual = str(row[3] or '').strip()
    comp_ddd = row[4]
    tipo = str(row[6] or '').strip()

    # Fix col E (componente) — substitui constantes por nome correto
    fixed_comp = fix_component_name(fname, comp_atual)
    if fixed_comp != comp_atual:
        row[3] = fixed_comp

    # Col F (comp DDD) — só se rename
    if not comp_ddd and fixed_comp in COMP_RENAMES:
        row[4] = COMP_RENAMES[fixed_comp]

    # Explicação
    explicacao = EXPLICACOES.get(fixed_comp) or EXPLICACOES.get(fname.replace('.tsx','')) \
                 or gerar_explicacao_fallback(fname, fixed_comp, tipo)

    new_row = row[:3] + [explicacao] + row[3:]
    rows_out.append(new_row)

with open(DST, 'w', encoding='utf-8', newline='') as f:
    writer = csv.writer(f)
    for r in rows_out:
        writer.writerow(['' if v is None else v for v in r])

print(f'Linhas: {len(rows_out)-1}')
filled_d = sum(1 for r in rows_out[1:] if r[3])
filled_f = sum(1 for r in rows_out[1:] if r[5])
fixed = sum(1 for r in rows_out[1:] if r[4] and not r[4].startswith('Modal') == False)
print(f'Col D (Explicação): {filled_d}')
print(f'Col F (Componente DDD): {filled_f}')
print(f'Arquivo: {DST}')
