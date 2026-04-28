"""Gera mapa-nucleo-global final com:
- col A (Nome) corrigido (meta/constantes → nome real)
- col B (DDD) preenchido só se rename
- col C (Categoria) preenchido com a pasta
- col D (Explicação) NOVA
"""
import openpyxl, csv, re

SRC = r'C:\Users\danie\Downloads\planilha_geral_gravity_COMPLETA.xlsx'
DST = r'C:\Users\danie\Downloads\mapa_nucleo_global_final.csv'

# Explicações por categoria (prefixo da pasta)
CATEGORIA_DESC = {
    'Botoes': 'Botão — componente de ação do design system.',
    'Campos': 'Campo de formulário — entrada de dados.',
    'Cards': 'Card — container visual com conteúdo agrupado.',
    'Modais': 'Modal — overlay/diálogo reutilizável.',
    'Tabelas': 'Tabela — grid de dados.',
    'Layouts': 'Layout — estrutura de página reutilizável.',
    'Filtros': 'Filtro — controle de seleção de dados exibidos.',
    'Graficos': 'Gráfico — visualização de dados.',
    'Menus': 'Menu — navegação ou lista de ações.',
    'Pills': 'Pill — badge/tag visual.',
    'Dashboard': 'Componente de dashboard (widget, painel).',
    'Kanban': 'Componente do board Kanban.',
    'Notificacao': 'Componente de notificação (toast/alerta).',
    'Gabi': 'Componente da IA Gabi.',
    'Selects': 'Select — dropdown de seleção.',
    'Dropdown': 'Dropdown — menu suspenso.',
    'Inputs': 'Input — campo de entrada básico.',
    'Utilidades': 'Utilidade — componente auxiliar.',
    'Sidebar': 'Sidebar — barra lateral.',
    'Header': 'Header — cabeçalho/topbar.',
    'Footer': 'Footer — rodapé.',
    'Loaders': 'Loader — indicador de carregamento.',
    'Tooltips': 'Tooltip — dica contextual.',
    'Tabs': 'Tabs — navegação por abas.',
    'Accordion': 'Accordion — conteúdo expansível.',
    'Stepper': 'Stepper — passos sequenciais.',
    'Pagination': 'Paginação.',
    'Tags': 'Tag — rótulo visual.',
    'Icons': 'Ícone.',
    'Avatar': 'Avatar do usuário.',
    'Dialogs': 'Diálogo modal.',
    'Datepicker': 'Seletor de data.',
}

def extract_category(filepath):
    """Extrai categoria da pasta nucleo-global/<Categoria>/..."""
    if not filepath: return ''
    m = re.search(r'nucleo-global/([^/]+)/', filepath)
    if m: return m[1]
    return ''

def fix_component_name(current, filepath):
    """Corrige nomes inválidos (meta, MAIUSCULAS, lowercase) pelo derivado do arquivo."""
    if not current or not filepath: return current
    c = str(current).strip()
    fname = filepath.split('/')[-1].replace('.tsx','').replace('.jsx','')
    # Nomes inválidos comuns
    invalid = {'meta','default'}
    if c in invalid or c.isupper() or (c[0].islower() if c else False):
        # Deriva pelo nome do arquivo (PascalCase)
        # Se o arquivo é kebab-case (botao-global), pega o PascalCase
        # Se já é PascalCase (BotaoGlobal), retorna
        if '-' in fname:
            # kebab-case → PascalCase
            parts = fname.split('-')
            return ''.join(p.capitalize() for p in parts)
        if fname[0].isupper():
            return fname
        return fname
    return c

def is_story_file(filepath):
    return '.stories' in (filepath or '')

def gerar_explicacao(comp_name, categoria, filepath, tipo):
    """Gera explicação do componente."""
    if is_story_file(filepath):
        return '⚠️ Arquivo Storybook (.stories.tsx) — definição para testes visuais, não é componente de produção.'

    base_desc = CATEGORIA_DESC.get(categoria, '')

    # Tenta inferir pelo nome
    n = comp_name or ''
    if n.endswith('Global'):
        n_clean = n[:-6]
    else:
        n_clean = n

    # Alguns nomes específicos do projeto
    especificos = {
        'BotaoGlobal': 'Botão primário/universal do design system (variantes por prop).',
        'BotaoSalvar': 'Botão padrão de salvar (variante do BotaoGlobal).',
        'BotaoCancelar': 'Botão padrão de cancelar.',
        'BotaoNovoAdminGlobal': 'Botão "+ Novo" usado em telas de admin (toolbar).',
        'GeralCampoGlobal': 'Campo de formulário genérico configurável (texto/número/data/select).',
        'LocalizarExpandidoCampoGlobal': 'Campo de busca com dropdown expandido de resultados.',
        'ModalBuscaNcm': 'Modal de busca e seleção de NCM do catálogo.',
        'SelectOpcao': 'Select genérico com opções configuráveis.',
        'InsightCard': 'Card de insight da Gabi (sugestão/alerta).',
    }
    if n in especificos:
        return especificos[n]

    # Fallback usa a categoria
    if base_desc:
        # Humaniza o nome
        human = re.sub(r'(.)([A-Z][a-z]+)', r'\1 \2', n_clean)
        human = re.sub(r'([a-z0-9])([A-Z])', r'\1 \2', human)
        if human:
            return f'{base_desc} ({human.strip()})'
        return base_desc

    return f'Componente do design system: {n_clean or comp_name}.'

# Processa
wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb['8. Nucleo Global']

header_atual = [ws.cell(1,c).value or '' for c in range(1, 12)]
# Insere col "Explicação" após C (Categoria)
new_header = header_atual[:3] + ['Explicação'] + header_atual[3:]

rows_out = [new_header]
for r in range(2, ws.max_row+1):
    row = [ws.cell(r,c).value for c in range(1, 12)]
    if not row[0] and not row[7]: continue  # sem nome e sem arquivo

    nome_atual = str(row[0] or '').strip()
    nome_ddd = row[1]
    cat_atual = row[2]
    filepath = str(row[7] or '').strip()
    tipo = str(row[4] or '').strip()

    # Corrige nome (col A) se inválido
    fixed_name = fix_component_name(nome_atual, filepath)
    if fixed_name != nome_atual:
        row[0] = fixed_name

    # Categoria (col C)
    if not cat_atual:
        cat = extract_category(filepath)
        if cat: row[2] = cat

    # Se é story file, marca no tipo
    if is_story_file(filepath):
        row[4] = 'Storybook'

    # Explicação (col D nova)
    categoria = row[2] or ''
    explicacao = gerar_explicacao(fixed_name, categoria, filepath, tipo)

    # Monta linha final
    new_row = row[:3] + [explicacao] + row[3:]
    rows_out.append(new_row)

with open(DST, 'w', encoding='utf-8', newline='') as f:
    writer = csv.writer(f)
    for r in rows_out:
        writer.writerow(['' if v is None else v for v in r])

total = len(rows_out) - 1
filled_b = sum(1 for r in rows_out[1:] if r[1])
filled_c = sum(1 for r in rows_out[1:] if r[2])
filled_d = sum(1 for r in rows_out[1:] if r[3])
stories = sum(1 for r in rows_out[1:] if '.stories' in str(r[-2] or '') or r[4] == 'Storybook')
print(f'Total linhas: {total}')
print(f'Col B (Componente DDD): {filled_b}')
print(f'Col C (Categoria): {filled_c}')
print(f'Col D (Explicação): {filled_d}')
print(f'Arquivos Storybook marcados: {stories}')
print(f'Arquivo: {DST}')
