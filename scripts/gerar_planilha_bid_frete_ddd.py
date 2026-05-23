#!/usr/bin/env python3
"""
Gera planilha DDD completa para o produto Bid Frete.
Formato idêntico à planilha mestre do Pedido.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import os

OUTPUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "planilha_bid_frete_ddd.xlsx")

# ── Styles ──────────────────────────────────────────────────────────────
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
CELL_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

# Tab colors
TAB_COLORS = {
    "1.ddd_campos": "4472C4",       # blue
    "2. ddd_api": "70AD47",          # green
    "3. tabelas-models": "ED7D31",   # orange
    "4. mapa-enums": "7030A0",       # purple
    "5. mapa-rotas": "FF0000",       # red
    "6. mapa-paginas": "FFD966",     # yellow
    "7. Modais": "A6A6A6",           # gray
    "8. nucleo-global": "00B0F0",    # teal
    "9. componentes-locais": "FF69B4", # pink
}

def style_sheet(ws, headers, col_widths):
    """Apply header styling, freeze pane, auto-filter."""
    ws.append(headers)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def add_row(ws, row_data):
    """Add a data row with styling."""
    ws.append(row_data)
    r = ws.max_row
    for col_idx in range(1, len(row_data) + 1):
        cell = ws.cell(row=r, column=col_idx)
        cell.alignment = CELL_ALIGN
        cell.border = THIN_BORDER

def add_row_center_nums(ws, row_data, center_cols=None):
    """Add row, center specific columns (for numeric data)."""
    ws.append(row_data)
    r = ws.max_row
    for col_idx in range(1, len(row_data) + 1):
        cell = ws.cell(row=r, column=col_idx)
        if center_cols and col_idx in center_cols:
            cell.alignment = Alignment(horizontal="center", vertical="top")
        else:
            cell.alignment = CELL_ALIGN
        cell.border = THIN_BORDER

# ════════════════════════════════════════════════════════════════════════
# DATA DEFINITIONS
# ════════════════════════════════════════════════════════════════════════

LOCAL = "Banco de Dados - Bid Frete"
PRODUTO = "Bid Frete"

# ── All models with fields ──────────────────────────────────────────────

MODELS = {
    "FreteIntBidFornecedores": {
        "entity": "Fornecedor",
        "fields": [
            # (field_name, tipo_dado, optional, default, ddd_name, natureza, editavel, explicacao, local_tela, nome_tela_atual, nome_tela_ddd, descricao, formato, validacao, valor_padrao, exemplo, componente, origem)
            ("id", "String", False, "@default(cuid())", "id_fornecedor_bid", "sistema", "Não", "", "", "", "", "Identificador único do fornecedor", "CUID", "", "", "clxyz123...", "", "Sistema"),
            ("id_organizacao", "String", False, None, "id_organizacao", "sistema", "Não", "", "", "", "", "ID da organização (tenant)", "CUID", "", "", "clorg456...", "", "Sistema"),
            ("product_id", "String", True, None, "id_produto", "sistema", "Não", "Prisma tem product_id, types.ts não tem", "", "", "", "ID do produto", "CUID", "", "", "clprod...", "", "Sistema"),
            ("user_id", "String", True, None, "id_usuario_criador", "sistema", "Não", "Prisma tem user_id, types.ts não tem", "", "", "", "ID do usuário que criou", "CUID", "", "", "clusr...", "", "Sistema"),
            ("nome", "String", False, None, "nome_fornecedor_bid", "usuario", "Sim", "", "Lista/Detalhe Fornecedor", "Nome", "Nome do Fornecedor", "Razão social do fornecedor", "Texto livre", "Min 2 chars", "", "Maersk Line", "InputTextoGlobal", "Usuario"),
            ("nome_fantasia", "String", True, None, "nome_fantasia_fornecedor_bid", "usuario", "Sim", "", "Detalhe Fornecedor", "Nome Fantasia", "Nome Fantasia", "Nome fantasia/comercial", "Texto livre", "", "", "Maersk", "InputTextoGlobal", "Usuario"),
            ("tipo", "BidFreteTipoFornecedor", False, None, "tipo_fornecedor_bid", "usuario", "Sim", "", "Lista/Detalhe Fornecedor", "Tipo", "Tipo do Fornecedor", "Tipo do fornecedor", "Enum", "Valor do enum", "", "ARMADOR", "SelectGlobal", "Usuario"),
            ("cnpj", "String", True, None, "cnpj_fornecedor_bid", "usuario", "Sim", "", "Detalhe Fornecedor", "CNPJ", "CNPJ", "CNPJ do fornecedor", "XX.XXX.XXX/XXXX-XX", "Validação CNPJ", "", "12.345.678/0001-90", "InputTextoGlobal", "Usuario"),
            ("email", "String", False, None, "email_fornecedor_bid", "usuario", "Sim", "", "Lista/Detalhe Fornecedor", "E-mail", "E-mail do Fornecedor", "E-mail principal de contato", "email", "Formato e-mail válido", "", "contato@maersk.com", "InputTextoGlobal", "Usuario"),
            ("telefone", "String", True, None, "telefone_fornecedor_bid", "usuario", "Sim", "", "Detalhe Fornecedor", "Telefone", "Telefone", "Telefone do fornecedor", "Texto livre", "", "", "+55 11 99999-0000", "InputTextoGlobal", "Usuario"),
            ("whatsapp", "String", True, None, "whatsapp_fornecedor_bid", "usuario", "Sim", "", "Detalhe Fornecedor", "WhatsApp", "WhatsApp", "WhatsApp do fornecedor", "Texto livre", "", "", "+55 11 99999-0000", "InputTextoGlobal", "Usuario"),
            ("website", "String", True, None, "website_fornecedor_bid", "usuario", "Sim", "", "Detalhe Fornecedor", "Website", "Website", "Website do fornecedor", "URL", "Formato URL", "", "https://maersk.com", "InputTextoGlobal", "Usuario"),
            ("pais", "String", True, None, "pais_fornecedor_bid", "usuario", "Sim", "", "Detalhe Fornecedor", "País", "País", "País sede do fornecedor", "Texto livre", "", "", "Brasil", "SelectGlobal", "Usuario"),
            ("cidade", "String", True, None, "cidade_fornecedor_bid", "usuario", "Sim", "", "Detalhe Fornecedor", "Cidade", "Cidade", "Cidade sede do fornecedor", "Texto livre", "", "", "São Paulo", "InputTextoGlobal", "Usuario"),
            ("status", "BidFreteStatusFornecedor", False, "@default(ATIVO)", "status_fornecedor_bid", "sistema", "Sim", "", "Lista Fornecedor", "Status", "Status do Fornecedor", "Status do fornecedor", "Enum", "Valor do enum", "ATIVO", "ATIVO", "BadgeGlobal", "Sistema"),
            ("clerk_user_id", "String", True, None, "id_clerk_usuario_fornecedor", "sistema", "Não", "Prisma tem clerk_user_id, types.ts não tem. Vincula ao portal do fornecedor", "", "", "", "ID Clerk do usuário fornecedor", "Texto", "", "", "user_xxx", "", "Sistema"),
            ("aceita_cotacao_aberta", "Boolean", False, "@default(true)", "aceita_cotacao_aberta_fornecedor_bid", "usuario", "Sim", "", "Detalhe Fornecedor", "Aceita Cotação Aberta", "Aceita Cotação Aberta", "Se aceita cotações abertas", "Boolean", "", "true", "true", "Checkbox", "Usuario"),
            ("cotacao_automatica", "Boolean", False, "@default(false)", "cotacao_automatica_fornecedor_bid", "usuario", "Sim", "DIVERGÊNCIA: types.ts usa 'resposta_automatica', prisma usa 'cotacao_automatica'", "Detalhe Fornecedor", "Resposta Automática", "Cotação Automática", "Se responde automaticamente via tabela", "Boolean", "", "false", "false", "Checkbox", "Usuario"),
            ("created_at", "DateTime", False, "@default(now())", "criado_em", "sistema", "Não", "", "", "", "", "Data de criação do registro", "ISO 8601", "", "", "2026-01-15T10:00:00Z", "", "Sistema"),
            ("updated_at", "DateTime", False, "@updatedAt", "atualizado_em", "sistema", "Não", "", "", "", "", "Data de última atualização", "ISO 8601", "", "", "2026-01-15T10:00:00Z", "", "Sistema"),
        ],
        "extra_ts_fields": [
            ("rating_global", "Float", True, None, "rating_global_fornecedor_bid", "sistema", "Não", "DIVERGÊNCIA: campo computado apenas no types.ts, não existe no prisma. Vem da tabela ClassificacaoFornecedor", "Lista Fornecedor", "Rating", "Rating Global", "Nota geral do fornecedor", "0.0-5.0", "", "", "4.5", "", "Calculado"),
            ("total_cotacoes", "Int", True, None, "total_cotacoes_fornecedor_bid", "sistema", "Não", "DIVERGÊNCIA: campo computado apenas no types.ts, não existe no prisma. Vem da tabela ClassificacaoFornecedor", "Lista Fornecedor", "Total Cotações", "Total de Cotações", "Qtd total de cotações recebidas", "Inteiro", "", "", "150", "", "Calculado"),
            ("taxa_resposta", "Float", True, None, "taxa_resposta_fornecedor_bid", "sistema", "Não", "DIVERGÊNCIA: campo computado apenas no types.ts, não existe no prisma. Vem da tabela ClassificacaoFornecedor", "Detalhe Fornecedor", "Taxa Resposta", "Taxa de Resposta", "Percentual de cotações respondidas", "0-100%", "", "", "85.5", "", "Calculado"),
            ("taxa_aprovacao", "Float", True, None, "taxa_aprovacao_fornecedor_bid", "sistema", "Não", "DIVERGÊNCIA: campo computado apenas no types.ts, não existe no prisma. Vem da tabela ClassificacaoFornecedor", "Detalhe Fornecedor", "Taxa Aprovação", "Taxa de Aprovação", "Percentual de cotações aprovadas", "0-100%", "", "", "72.3", "", "Calculado"),
            ("tempo_medio_resposta", "Float", True, None, "tempo_medio_resposta_fornecedor_bid", "sistema", "Não", "DIVERGÊNCIA: campo computado apenas no types.ts, não existe no prisma. Vem da tabela ClassificacaoFornecedor", "Detalhe Fornecedor", "Tempo Médio Resposta", "Tempo Médio de Resposta", "Tempo médio de resposta em horas", "Horas", "", "", "12.5", "", "Calculado"),
        ],
    },
    "FreteIntBidCotacoes": {
        "entity": "Cotacao",
        "fields": [
            ("id", "String", False, "@default(cuid())", "id_cotacao_bid", "sistema", "Não", "", "", "", "", "Identificador único da cotação", "CUID", "", "", "clxyz123...", "", "Sistema"),
            ("id_organizacao", "String", False, None, "id_organizacao", "sistema", "Não", "", "", "", "", "ID da organização (tenant)", "CUID", "", "", "clorg456...", "", "Sistema"),
            ("product_id", "String", True, None, "id_produto", "sistema", "Não", "", "", "", "", "ID do produto", "CUID", "", "", "clprod...", "", "Sistema"),
            ("user_id", "String", False, None, "id_usuario_criador", "sistema", "Não", "", "", "", "", "ID do usuário que criou", "CUID", "", "", "clusr...", "", "Sistema"),
            ("company_id", "String", True, None, "id_workspace", "sistema", "Não", "", "", "", "", "ID do workspace/filial", "CUID", "", "", "clcomp...", "", "Sistema"),
            ("numero", "String", False, None, "numero_cotacao_bid", "sistema", "Não", "", "Lista Cotações", "Número", "Número da Cotação", "Número sequencial da cotação", "Texto", "Auto-gerado", "", "BID-2026-001", "", "Sistema"),
            ("referencia_interna", "String", True, None, "referencia_interna_cotacao_bid", "usuario", "Sim", "", "Detalhe Cotação", "Ref. Interna", "Referência Interna", "Referência interna do cliente", "Texto livre", "", "", "PO-12345", "InputTextoGlobal", "Usuario"),
            ("tipo_operacao", "BidFreteTipoOperacao", False, None, "tipo_operacao_cotacao_bid", "usuario", "Sim", "", "Nova Cotação Step 1", "Tipo Operação", "Tipo de Operação", "Importação ou Exportação", "Enum", "Valor do enum", "", "IMPORTACAO", "SelectGlobal", "Usuario"),
            ("modal", "BidFreteModalidade", False, None, "modal_cotacao_bid", "usuario", "Sim", "", "Nova Cotação Step 1", "Modal", "Modal de Transporte", "Modal de transporte principal", "Enum", "Valor do enum", "", "MARITIMO", "SelectGlobal", "Usuario"),
            ("modalidade", "BidFreteCargaModalidade", False, None, "modalidade_carga_cotacao_bid", "usuario", "Sim", "", "Nova Cotação Step 1", "Modalidade Carga", "Modalidade de Carga", "Tipo de carga/container", "Enum", "Valor do enum", "", "FCL", "SelectGlobal", "Usuario"),
            ("origem_codigo", "String", False, None, "codigo_origem_cotacao_bid", "usuario", "Sim", "", "Nova Cotação Step 2", "Origem (Código)", "Código da Origem", "Código do porto/aeroporto de origem", "Texto", "", "", "BRSSZ", "SelectGlobal", "Usuario"),
            ("origem_nome", "String", False, None, "nome_origem_cotacao_bid", "usuario", "Sim", "", "Lista/Detalhe Cotação", "Origem", "Nome da Origem", "Nome do porto/aeroporto de origem", "Texto", "", "", "Santos", "", "Usuario"),
            ("origem_pais", "String", False, None, "pais_origem_cotacao_bid", "usuario", "Sim", "", "Detalhe Cotação", "País Origem", "País de Origem", "País de origem", "Texto", "", "", "Brasil", "", "Usuario"),
            ("destino_codigo", "String", False, None, "codigo_destino_cotacao_bid", "usuario", "Sim", "", "Nova Cotação Step 2", "Destino (Código)", "Código do Destino", "Código do porto/aeroporto de destino", "Texto", "", "", "CNSHA", "SelectGlobal", "Usuario"),
            ("destino_nome", "String", False, None, "nome_destino_cotacao_bid", "usuario", "Sim", "", "Lista/Detalhe Cotação", "Destino", "Nome do Destino", "Nome do porto/aeroporto de destino", "Texto", "", "", "Shanghai", "", "Usuario"),
            ("destino_pais", "String", False, None, "pais_destino_cotacao_bid", "usuario", "Sim", "", "Detalhe Cotação", "País Destino", "País de Destino", "País de destino", "Texto", "", "", "China", "", "Usuario"),
            ("descricao_mercadoria", "String", False, None, "descricao_mercadoria_cotacao_bid", "usuario", "Sim", "", "Nova Cotação Step 3", "Mercadoria", "Descrição da Mercadoria", "Descrição da mercadoria", "Texto livre", "Min 5 chars", "", "Peças automotivas", "InputTextoGlobal", "Usuario"),
            ("ncm", "String", True, None, "ncm_cotacao_bid", "usuario", "Sim", "", "Nova Cotação Step 3", "NCM", "NCM", "Código NCM da mercadoria", "XXXX.XX.XX", "8 dígitos", "", "8708.99.90", "InputTextoGlobal", "Usuario"),
            ("quantidade", "Int", False, "@default(1)", "quantidade_cotacao_bid", "usuario", "Sim", "", "Nova Cotação Step 3", "Quantidade", "Quantidade", "Quantidade de volumes/containers", "Inteiro", "Min 1", "1", "2", "InputTextoGlobal", "Usuario"),
            ("tipo_container", "String", True, None, "tipo_container_cotacao_bid", "usuario", "Sim", "", "Nova Cotação Step 3", "Tipo Container", "Tipo de Container", "Tipo do container (20DV, 40HC, etc.)", "Texto", "", "", "40HC", "SelectGlobal", "Usuario"),
            ("peso_kg", "Float", True, None, "peso_kg_cotacao_bid", "usuario", "Sim", "", "Nova Cotação Step 3", "Peso (kg)", "Peso em kg", "Peso total em quilogramas", "Decimal", "Min 0", "", "15000.50", "InputTextoGlobal", "Usuario"),
            ("cubagem_m3", "Float", True, None, "cubagem_m3_cotacao_bid", "usuario", "Sim", "", "Nova Cotação Step 3", "Cubagem (m³)", "Cubagem em m³", "Volume total em metros cúbicos", "Decimal", "Min 0", "", "65.00", "InputTextoGlobal", "Usuario"),
            ("incoterm", "String", False, None, "incoterm_cotacao_bid", "usuario", "Sim", "", "Nova Cotação Step 2", "Incoterm", "Incoterm", "Incoterm aplicável", "Texto", "Incoterm válido", "", "FOB", "SelectGlobal", "Usuario"),
            ("zip_code_origem", "String", True, None, "cep_origem_cotacao_bid", "usuario", "Sim", "types.ts não tem este campo", "Nova Cotação Step 2", "CEP Origem", "CEP de Origem", "CEP/ZIP code da origem (para rodoviário)", "Texto", "", "", "01310-100", "InputTextoGlobal", "Usuario"),
            ("zip_code_destino", "String", True, None, "cep_destino_cotacao_bid", "usuario", "Sim", "DIVERGÊNCIA: types.ts usa 'cep_destino', prisma usa 'zip_code_destino'", "Nova Cotação Step 2", "CEP Destino", "CEP de Destino", "CEP/ZIP code do destino (para rodoviário)", "Texto", "", "", "200040", "InputTextoGlobal", "Usuario"),
            ("valor_target", "Float", True, None, "valor_alvo_cotacao_bid", "usuario", "Sim", "DIVERGÊNCIA: types.ts usa 'valor_alvo', prisma usa 'valor_target'", "Nova Cotação Step 4", "Valor Alvo", "Valor Alvo", "Valor target para negociação", "Decimal", "Min 0", "", "5000.00", "InputTextoGlobal", "Usuario"),
            ("moeda_target", "String", True, "@default(\"USD\")", "moeda_alvo_cotacao_bid", "usuario", "Sim", "DIVERGÊNCIA: types.ts usa 'moeda_alvo', prisma usa 'moeda_target'", "Nova Cotação Step 4", "Moeda Alvo", "Moeda do Valor Alvo", "Moeda do valor target", "ISO 4217", "", "USD", "USD", "SelectGlobal", "Usuario"),
            ("visibilidade", "BidFreteCotacaoVisibilidade", False, "@default(DIRECIONADA)", "visibilidade_cotacao_bid", "usuario", "Sim", "", "Nova Cotação Step 5", "Visibilidade", "Visibilidade da Cotação", "Se a cotação é direcionada ou aberta", "Enum", "Valor do enum", "DIRECIONADA", "DIRECIONADA", "SelectGlobal", "Usuario"),
            ("ocultar_nome_empresa", "Boolean", False, "@default(false)", "ocultar_nome_empresa_cotacao_bid", "usuario", "Sim", "DIVERGÊNCIA: types.ts usa 'anonima', prisma usa 'ocultar_nome_empresa'", "Nova Cotação Step 5", "Anônima", "Ocultar Nome da Empresa", "Se oculta nome da empresa na cotação", "Boolean", "", "false", "false", "Checkbox", "Usuario"),
            ("status", "BidFreteCotacaoStatus", False, "@default(RASCUNHO)", "status_cotacao_bid", "sistema", "Não", "", "Lista Cotações", "Status", "Status da Cotação", "Status atual da cotação", "Enum", "Valor do enum", "RASCUNHO", "ENVIADA_FORNECEDORES", "BadgeGlobal", "Sistema"),
            ("data_limite_resposta", "DateTime", True, None, "data_limite_resposta_cotacao_bid", "usuario", "Sim", "DIVERGÊNCIA: types.ts usa 'prazo_resposta', prisma usa 'data_limite_resposta'", "Nova Cotação Step 4", "Prazo Resposta", "Data Limite de Resposta", "Data limite para fornecedores responderem", "ISO 8601", "Futuro", "", "2026-02-15T23:59:00Z", "InputTextoGlobal", "Usuario"),
            ("data_aprovacao", "DateTime", True, None, "data_aprovacao_cotacao_bid", "sistema", "Não", "Apenas no prisma, types.ts não tem", "", "", "", "Data de aprovação da cotação", "ISO 8601", "", "", "2026-02-10T14:30:00Z", "", "Sistema"),
            ("data_cancelamento", "DateTime", True, None, "data_cancelamento_cotacao_bid", "sistema", "Não", "Apenas no prisma, types.ts não tem", "", "", "", "Data de cancelamento da cotação", "ISO 8601", "", "", "2026-02-12T09:00:00Z", "", "Sistema"),
            ("motivo_reprovacao", "String", True, None, "motivo_reprovacao_cotacao_bid", "usuario", "Sim", "Apenas no prisma, types.ts não tem", "", "", "", "Motivo da reprovação", "Texto livre", "", "", "Valores acima do orçamento", "InputTextoGlobal", "Usuario"),
            ("motivo_cancelamento", "String", True, None, "motivo_cancelamento_cotacao_bid", "usuario", "Sim", "Apenas no prisma, types.ts não tem", "", "", "", "Motivo do cancelamento", "Texto livre", "", "", "Operação cancelada pelo cliente", "InputTextoGlobal", "Usuario"),
            ("fornecedor_vencedor_id", "String", True, None, "id_fornecedor_vencedor_cotacao_bid", "sistema", "Não", "Apenas no prisma, types.ts não tem. FK para FreteIntBidFornecedores", "", "", "", "ID do fornecedor vencedor", "CUID", "", "", "clforn...", "", "Sistema"),
            ("saving_valor", "Float", True, None, "saving_valor_cotacao_bid", "sistema", "Não", "", "Detalhe Cotação", "Saving (Valor)", "Saving em Valor", "Economia obtida em valor absoluto", "Decimal", "", "", "1500.00", "", "Calculado"),
            ("saving_percentual", "Float", True, None, "saving_percentual_cotacao_bid", "sistema", "Não", "", "Detalhe Cotação", "Saving (%)", "Saving Percentual", "Economia obtida em percentual", "Decimal 0-100", "", "", "23.5", "", "Calculado"),
            ("created_at", "DateTime", False, "@default(now())", "criado_em", "sistema", "Não", "", "", "", "", "Data de criação do registro", "ISO 8601", "", "", "2026-01-15T10:00:00Z", "", "Sistema"),
            ("updated_at", "DateTime", False, "@updatedAt", "atualizado_em", "sistema", "Não", "", "", "", "", "Data de última atualização", "ISO 8601", "", "", "2026-01-15T10:00:00Z", "", "Sistema"),
        ],
        "extra_ts_fields": [
            ("valor_aprovado", "Float", True, None, "valor_aprovado_cotacao_bid", "sistema", "Não", "DIVERGÊNCIA: campo existe apenas no types.ts, não no prisma. Provavelmente vem da Proposta aprovada", "Detalhe Cotação", "Valor Aprovado", "Valor Aprovado", "Valor da proposta aprovada", "Decimal", "", "", "4500.00", "", "Calculado"),
            ("moeda_aprovada", "String", True, None, "moeda_aprovada_cotacao_bid", "sistema", "Não", "DIVERGÊNCIA: campo existe apenas no types.ts, não no prisma. Provavelmente vem da Proposta aprovada", "Detalhe Cotação", "Moeda Aprovada", "Moeda Aprovada", "Moeda da proposta aprovada", "ISO 4217", "", "", "USD", "", "Calculado"),
        ],
    },
    "FreteIntBidPedidoCotacoes": {
        "entity": "PedidoCotacao",
        "fields": [
            ("id", "String", False, "@default(cuid())", "id_pedido_cotacao_bid", "sistema", "Não", "", "", "", "", "Identificador único do pedido de cotação", "CUID", "", "", "clxyz123...", "", "Sistema"),
            ("id_organizacao", "String", False, None, "id_organizacao", "sistema", "Não", "", "", "", "", "ID da organização (tenant)", "CUID", "", "", "clorg456...", "", "Sistema"),
            ("product_id", "String", True, None, "id_produto", "sistema", "Não", "Prisma tem product_id, types.ts não tem", "", "", "", "ID do produto", "CUID", "", "", "clprod...", "", "Sistema"),
            ("user_id", "String", False, None, "id_usuario_criador", "sistema", "Não", "", "", "", "", "ID do usuário que disparou", "CUID", "", "", "clusr...", "", "Sistema"),
            ("cotacao_id", "String", False, None, "id_cotacao_pedido_cotacao_bid", "sistema", "Não", "", "", "", "", "FK para a cotação", "CUID", "", "", "clcot...", "", "Sistema"),
            ("fornecedor_id", "String", False, None, "id_fornecedor_pedido_cotacao_bid", "sistema", "Não", "", "", "", "", "FK para o fornecedor", "CUID", "", "", "clforn...", "", "Sistema"),
            ("canal", "BidFreteCotacaoFreteIntCanal", False, "@default(EMAIL)", "canal_pedido_cotacao_bid", "usuario", "Sim", "", "Modal Disparo", "Canal", "Canal de Envio", "Canal de envio do pedido", "Enum", "Valor do enum", "EMAIL", "EMAIL", "SelectGlobal", "Usuario"),
            ("status", "BidFreteCotacao", False, "@default(PENDENTE)", "status_pedido_cotacao_bid", "sistema", "Não", "", "Detalhe Cotação", "Status", "Status do Pedido", "Status do envio do pedido", "Enum", "Valor do enum", "PENDENTE", "ENVIADO", "BadgeGlobal", "Sistema"),
            ("enviado_em", "DateTime", True, None, "enviado_em_pedido_cotacao_bid", "sistema", "Não", "", "", "", "", "Data/hora do envio", "ISO 8601", "", "", "2026-01-15T10:05:00Z", "", "Sistema"),
            ("visualizado_em", "DateTime", True, None, "visualizado_em_pedido_cotacao_bid", "sistema", "Não", "", "Detalhe Cotação", "Visualizado em", "Visualizado em", "Data/hora da visualização pelo fornecedor", "ISO 8601", "", "", "2026-01-15T11:00:00Z", "", "Sistema"),
            ("respondido_em", "DateTime", True, None, "respondido_em_pedido_cotacao_bid", "sistema", "Não", "", "Detalhe Cotação", "Respondido em", "Respondido em", "Data/hora da resposta", "ISO 8601", "", "", "2026-01-16T09:00:00Z", "", "Sistema"),
            ("token_resposta", "String", True, None, "token_resposta_pedido_cotacao_bid", "sistema", "Não", "DIVERGÊNCIA: types.ts usa 'token_publico', prisma usa 'token_resposta'", "", "", "", "Token único para resposta via link", "UUID/Hash", "Unique", "", "abc123-def456", "", "Sistema"),
            ("token_expira_em", "DateTime", True, None, "token_expira_em_pedido_cotacao_bid", "sistema", "Não", "DIVERGÊNCIA: types.ts usa 'expirado_em', prisma usa 'token_expira_em'", "", "", "", "Data de expiração do token", "ISO 8601", "", "", "2026-02-15T23:59:00Z", "", "Sistema"),
            ("mensagem_id", "String", True, None, "id_mensagem_pedido_cotacao_bid", "sistema", "Não", "Apenas no prisma, types.ts não tem. ID da mensagem enviada (e-mail/whatsapp)", "", "", "", "ID da mensagem enviada", "Texto", "", "", "msg_xxx", "", "Sistema"),
            ("erro_envio", "String", True, None, "erro_envio_pedido_cotacao_bid", "sistema", "Não", "Apenas no prisma, types.ts não tem. Erro ao enviar a mensagem", "", "", "", "Descrição do erro de envio", "Texto", "", "", "SMTP timeout", "", "Sistema"),
            ("created_at", "DateTime", False, "@default(now())", "criado_em", "sistema", "Não", "", "", "", "", "Data de criação do registro", "ISO 8601", "", "", "2026-01-15T10:00:00Z", "", "Sistema"),
            ("updated_at", "DateTime", False, "@updatedAt", "atualizado_em", "sistema", "Não", "", "", "", "", "Data de última atualização", "ISO 8601", "", "", "2026-01-15T10:00:00Z", "", "Sistema"),
        ],
        "extra_ts_fields": [],
    },
    "FreteIntBidPropostas": {
        "entity": "Proposta",
        "fields": [
            ("id", "String", False, "@default(cuid())", "id_proposta_bid", "sistema", "Não", "", "", "", "", "Identificador único da proposta", "CUID", "", "", "clxyz123...", "", "Sistema"),
            ("id_organizacao", "String", False, None, "id_organizacao", "sistema", "Não", "", "", "", "", "ID da organização (tenant)", "CUID", "", "", "clorg456...", "", "Sistema"),
            ("product_id", "String", True, None, "id_produto", "sistema", "Não", "Prisma tem product_id, types.ts não tem", "", "", "", "ID do produto", "CUID", "", "", "clprod...", "", "Sistema"),
            ("user_id", "String", True, None, "id_usuario_fornecedor", "sistema", "Não", "", "", "", "", "ID do usuário fornecedor que respondeu", "CUID", "", "", "clusr...", "", "Sistema"),
            ("bid_request_id", "String", False, None, "id_pedido_cotacao_proposta_bid", "sistema", "Não", "", "", "", "", "FK para o pedido de cotação (unique)", "CUID", "Unique", "", "clbid...", "", "Sistema"),
            ("cotacao_id", "String", False, None, "id_cotacao_proposta_bid", "sistema", "Não", "", "", "", "", "FK para a cotação", "CUID", "", "", "clcot...", "", "Sistema"),
            ("fornecedor_id", "String", False, None, "id_fornecedor_proposta_bid", "sistema", "Não", "", "", "", "", "FK para o fornecedor", "CUID", "", "", "clforn...", "", "Sistema"),
            ("moeda", "String", False, "@default(\"USD\")", "moeda_proposta_bid", "usuario", "Sim", "", "Comparativo/Responder", "Moeda", "Moeda", "Moeda da proposta", "ISO 4217", "", "USD", "USD", "SelectGlobal", "Usuario"),
            ("valor_frete", "Float", False, None, "valor_frete_proposta_bid", "usuario", "Sim", "", "Comparativo/Responder", "Frete", "Valor do Frete", "Valor do frete principal", "Decimal", "Min 0", "", "3500.00", "InputTextoGlobal", "Usuario"),
            ("taxas_origem", "Float", False, "@default(0)", "taxas_origem_proposta_bid", "usuario", "Sim", "", "Comparativo/Responder", "Taxas Origem", "Taxas de Origem", "Total de taxas na origem", "Decimal", "Min 0", "0", "250.00", "InputTextoGlobal", "Usuario"),
            ("taxas_destino", "Float", False, "@default(0)", "taxas_destino_proposta_bid", "usuario", "Sim", "", "Comparativo/Responder", "Taxas Destino", "Taxas de Destino", "Total de taxas no destino", "Decimal", "Min 0", "0", "180.00", "InputTextoGlobal", "Usuario"),
            ("valor_total", "Float", False, None, "valor_total_proposta_bid", "sistema", "Não", "", "Comparativo", "Total", "Valor Total", "Soma: frete + taxas_origem + taxas_destino", "Decimal", "", "", "3930.00", "", "Calculado"),
            ("transit_time_dias", "Int", False, None, "transit_time_dias_proposta_bid", "usuario", "Sim", "", "Comparativo/Responder", "Transit Time", "Transit Time (dias)", "Tempo de trânsito em dias", "Inteiro", "Min 1", "", "25", "InputTextoGlobal", "Usuario"),
            ("free_time_dias", "Int", True, None, "free_time_dias_proposta_bid", "usuario", "Sim", "", "Responder", "Free Time", "Free Time (dias)", "Dias de armazenagem gratuita", "Inteiro", "Min 0", "", "14", "InputTextoGlobal", "Usuario"),
            ("validade_cotacao", "DateTime", False, None, "validade_proposta_bid", "usuario", "Sim", "DIVERGÊNCIA: types.ts usa 'validade', prisma usa 'validade_cotacao'", "Comparativo/Responder", "Validade", "Validade da Proposta", "Data de validade da proposta", "ISO 8601", "Futuro", "", "2026-03-15T23:59:00Z", "InputTextoGlobal", "Usuario"),
            ("transbordos", "Int", False, "@default(0)", "transbordos_proposta_bid", "usuario", "Sim", "", "Responder", "Transbordos", "Número de Transbordos", "Quantidade de transbordos", "Inteiro", "Min 0", "0", "1", "InputTextoGlobal", "Usuario"),
            ("escalas", "String", True, None, "escalas_proposta_bid", "usuario", "Sim", "", "Responder", "Escalas", "Escalas", "Portos de escala/transbordo", "Texto livre", "", "", "Singapore, Colombo", "InputTextoGlobal", "Usuario"),
            ("observacoes", "String", True, None, "observacoes_proposta_bid", "usuario", "Sim", "", "Responder", "Observações", "Observações", "Observações adicionais", "Texto livre", "", "", "Sujeito a space confirmation", "InputTextoGlobal", "Usuario"),
            ("status", "BidFretePropostaStatus", False, "@default(RECEBIDA)", "status_proposta_bid", "sistema", "Não", "", "Comparativo", "Status", "Status da Proposta", "Status atual da proposta", "Enum", "Valor do enum", "RECEBIDA", "MELHOR_PRECO", "BadgeGlobal", "Sistema"),
            ("ranking_preco", "Int", True, None, "ranking_preco_proposta_bid", "sistema", "Não", "DIVERGÊNCIA: types.ts usa 'score_preco', prisma usa 'ranking_preco'", "Comparativo", "Ranking Preço", "Ranking por Preço", "Posição no ranking de preço", "Inteiro", "", "", "1", "", "Calculado"),
            ("ranking_transit", "Int", True, None, "ranking_transit_proposta_bid", "sistema", "Não", "DIVERGÊNCIA: types.ts usa 'score_transit', prisma usa 'ranking_transit'", "Comparativo", "Ranking Transit", "Ranking por Transit Time", "Posição no ranking de transit time", "Inteiro", "", "", "2", "", "Calculado"),
            ("ranking_avaliacao", "Int", True, None, "ranking_avaliacao_proposta_bid", "sistema", "Não", "DIVERGÊNCIA: types.ts usa 'score_avaliacao', prisma usa 'ranking_avaliacao'", "Comparativo", "Ranking Avaliação", "Ranking por Avaliação", "Posição no ranking de avaliação", "Inteiro", "", "", "1", "", "Calculado"),
            ("via_tabela_padrao", "Boolean", False, "@default(false)", "via_tabela_padrao_proposta_bid", "sistema", "Não", "Apenas no prisma, types.ts não tem", "", "", "", "Se veio de tabela padrão", "Boolean", "", "false", "false", "", "Sistema"),
            ("via_api", "Boolean", False, "@default(false)", "via_api_proposta_bid", "sistema", "Não", "Apenas no prisma, types.ts não tem", "", "", "", "Se veio via API", "Boolean", "", "false", "false", "", "Sistema"),
            ("via_portal", "Boolean", False, "@default(false)", "via_portal_proposta_bid", "sistema", "Não", "Apenas no prisma, types.ts não tem", "", "", "", "Se veio via portal do fornecedor", "Boolean", "", "false", "true", "", "Sistema"),
            ("via_email", "Boolean", False, "@default(false)", "via_email_proposta_bid", "sistema", "Não", "Apenas no prisma, types.ts não tem", "", "", "", "Se veio via e-mail/link público", "Boolean", "", "false", "false", "", "Sistema"),
            ("created_at", "DateTime", False, "@default(now())", "criado_em", "sistema", "Não", "", "", "", "", "Data de criação do registro", "ISO 8601", "", "", "2026-01-15T10:00:00Z", "", "Sistema"),
            ("updated_at", "DateTime", False, "@updatedAt", "atualizado_em", "sistema", "Não", "Apenas no prisma, types.ts não tem", "", "", "", "Data de última atualização", "ISO 8601", "", "", "2026-01-15T10:00:00Z", "", "Sistema"),
        ],
        "extra_ts_fields": [
            ("aprovada", "Boolean", True, None, "aprovada_proposta_bid", "sistema", "Não", "DIVERGÊNCIA: campo existe apenas no types.ts, não no prisma", "Comparativo", "Aprovada", "Aprovada", "Se a proposta foi aprovada", "Boolean", "", "", "true", "", "Calculado"),
            ("aprovada_em", "DateTime", True, None, "aprovada_em_proposta_bid", "sistema", "Não", "DIVERGÊNCIA: campo existe apenas no types.ts, não no prisma", "", "", "", "Data da aprovação", "ISO 8601", "", "", "2026-02-10T14:30:00Z", "", "Calculado"),
            ("aprovada_por", "String", True, None, "aprovada_por_proposta_bid", "sistema", "Não", "DIVERGÊNCIA: campo existe apenas no types.ts, não no prisma", "", "", "", "Usuário que aprovou", "CUID", "", "", "clusr...", "", "Calculado"),
        ],
    },
    "FreteIntBidPropostasTaxasCambio": {
        "entity": "DetalheTaxa",
        "fields": [
            ("id", "String", False, "@default(cuid())", "id_detalhe_taxa_bid", "sistema", "Não", "", "", "", "", "Identificador único da taxa", "CUID", "", "", "clxyz123...", "", "Sistema"),
            ("id_organizacao", "String", False, None, "id_organizacao", "sistema", "Não", "", "", "", "", "ID da organização (tenant)", "CUID", "", "", "clorg456...", "", "Sistema"),
            ("response_id", "String", False, None, "id_proposta_detalhe_taxa_bid", "sistema", "Não", "", "", "", "", "FK para a proposta", "CUID", "", "", "clresp...", "", "Sistema"),
            ("tipo", "String", False, None, "tipo_detalhe_taxa_bid", "usuario", "Sim", "", "Responder", "Tipo", "Tipo da Taxa", "Classificação da taxa (ORIGEM/DESTINO/FRETE)", "Texto", "", "", "ORIGEM", "SelectGlobal", "Usuario"),
            ("nome", "String", False, None, "nome_detalhe_taxa_bid", "usuario", "Sim", "", "Responder/Comparativo", "Nome", "Nome da Taxa", "Nome/descrição da taxa", "Texto livre", "", "", "THC", "InputTextoGlobal", "Usuario"),
            ("valor", "Float", False, None, "valor_detalhe_taxa_bid", "usuario", "Sim", "", "Responder/Comparativo", "Valor", "Valor da Taxa", "Valor monetário da taxa", "Decimal", "Min 0", "", "350.00", "InputTextoGlobal", "Usuario"),
            ("moeda", "String", False, "@default(\"USD\")", "moeda_detalhe_taxa_bid", "usuario", "Sim", "", "Responder", "Moeda", "Moeda da Taxa", "Moeda da taxa", "ISO 4217", "", "USD", "USD", "SelectGlobal", "Usuario"),
        ],
    },
    "FreteIntBidTabelasProntas": {
        "entity": "TabelaPreco",
        "fields": [
            ("id", "String", False, "@default(cuid())", "id_tabela_preco_bid", "sistema", "Não", "", "", "", "", "Identificador único da tabela", "CUID", "", "", "clxyz123...", "", "Sistema"),
            ("id_organizacao", "String", False, None, "id_organizacao", "sistema", "Não", "", "", "", "", "ID da organização (tenant)", "CUID", "", "", "clorg456...", "", "Sistema"),
            ("product_id", "String", True, None, "id_produto", "sistema", "Não", "", "", "", "", "ID do produto", "CUID", "", "", "clprod...", "", "Sistema"),
            ("user_id", "String", True, None, "id_usuario_criador", "sistema", "Não", "", "", "", "", "ID do usuário que criou", "CUID", "", "", "clusr...", "", "Sistema"),
            ("fornecedor_id", "String", False, None, "id_fornecedor_tabela_preco_bid", "sistema", "Não", "", "", "", "", "FK para o fornecedor", "CUID", "", "", "clforn...", "", "Sistema"),
            ("origem_codigo", "String", False, None, "codigo_origem_tabela_preco_bid", "usuario", "Sim", "", "Modal Tabela Preço", "Origem", "Código da Origem", "Código do porto de origem", "Texto", "", "", "BRSSZ", "SelectGlobal", "Usuario"),
            ("origem_nome", "String", False, None, "nome_origem_tabela_preco_bid", "usuario", "Sim", "", "Tabela Preços", "Origem", "Nome da Origem", "Nome do porto de origem", "Texto", "", "", "Santos", "", "Usuario"),
            ("destino_codigo", "String", False, None, "codigo_destino_tabela_preco_bid", "usuario", "Sim", "", "Modal Tabela Preço", "Destino", "Código do Destino", "Código do porto de destino", "Texto", "", "", "CNSHA", "SelectGlobal", "Usuario"),
            ("destino_nome", "String", False, None, "nome_destino_tabela_preco_bid", "usuario", "Sim", "", "Tabela Preços", "Destino", "Nome do Destino", "Nome do porto de destino", "Texto", "", "", "Shanghai", "", "Usuario"),
            ("modal", "BidFreteModalidade", False, None, "modal_tabela_preco_bid", "usuario", "Sim", "", "Modal Tabela Preço", "Modal", "Modal", "Modal de transporte", "Enum", "Valor do enum", "", "MARITIMO", "SelectGlobal", "Usuario"),
            ("modalidade", "BidFreteCargaModalidade", False, None, "modalidade_carga_tabela_preco_bid", "usuario", "Sim", "", "Modal Tabela Preço", "Modalidade", "Modalidade de Carga", "Tipo de carga", "Enum", "Valor do enum", "", "FCL", "SelectGlobal", "Usuario"),
            ("moeda", "String", False, "@default(\"USD\")", "moeda_tabela_preco_bid", "usuario", "Sim", "", "Modal Tabela Preço", "Moeda", "Moeda", "Moeda dos valores", "ISO 4217", "", "USD", "USD", "SelectGlobal", "Usuario"),
            ("valor_frete", "Float", False, None, "valor_frete_tabela_preco_bid", "usuario", "Sim", "", "Tabela Preços", "Frete", "Valor do Frete", "Valor do frete", "Decimal", "Min 0", "", "3200.00", "InputTextoGlobal", "Usuario"),
            ("taxas_origem", "Float", False, "@default(0)", "taxas_origem_tabela_preco_bid", "usuario", "Sim", "", "Modal Tabela Preço", "Taxas Origem", "Taxas de Origem", "Total taxas origem", "Decimal", "Min 0", "0", "200.00", "InputTextoGlobal", "Usuario"),
            ("taxas_destino", "Float", False, "@default(0)", "taxas_destino_tabela_preco_bid", "usuario", "Sim", "", "Modal Tabela Preço", "Taxas Destino", "Taxas de Destino", "Total taxas destino", "Decimal", "Min 0", "0", "150.00", "InputTextoGlobal", "Usuario"),
            ("valor_total", "Float", False, None, "valor_total_tabela_preco_bid", "sistema", "Não", "", "Tabela Preços", "Total", "Valor Total", "Soma dos valores", "Decimal", "", "", "3550.00", "", "Calculado"),
            ("transit_time_dias", "Int", False, None, "transit_time_dias_tabela_preco_bid", "usuario", "Sim", "", "Tabela Preços", "Transit Time", "Transit Time (dias)", "Tempo de trânsito", "Inteiro", "Min 1", "", "22", "InputTextoGlobal", "Usuario"),
            ("free_time_dias", "Int", True, None, "free_time_dias_tabela_preco_bid", "usuario", "Sim", "", "Modal Tabela Preço", "Free Time", "Free Time (dias)", "Dias de armazenagem gratuita", "Inteiro", "Min 0", "", "14", "InputTextoGlobal", "Usuario"),
            ("validade_inicio", "DateTime", False, None, "validade_inicio_tabela_preco_bid", "usuario", "Sim", "", "Modal Tabela Preço", "Início Validade", "Data Início Validade", "Data início da validade", "ISO 8601", "", "", "2026-01-01", "InputTextoGlobal", "Usuario"),
            ("validade_fim", "DateTime", False, None, "validade_fim_tabela_preco_bid", "usuario", "Sim", "", "Modal Tabela Preço", "Fim Validade", "Data Fim Validade", "Data fim da validade", "ISO 8601", "Após início", "", "2026-03-31", "InputTextoGlobal", "Usuario"),
            ("ativa", "Boolean", False, "@default(true)", "ativa_tabela_preco_bid", "usuario", "Sim", "", "Tabela Preços", "Ativa", "Ativa", "Se a tabela está ativa", "Boolean", "", "true", "true", "Checkbox", "Usuario"),
            ("created_at", "DateTime", False, "@default(now())", "criado_em", "sistema", "Não", "", "", "", "", "Data de criação do registro", "ISO 8601", "", "", "2026-01-15T10:00:00Z", "", "Sistema"),
            ("updated_at", "DateTime", False, "@updatedAt", "atualizado_em", "sistema", "Não", "", "", "", "", "Data de última atualização", "ISO 8601", "", "", "2026-01-15T10:00:00Z", "", "Sistema"),
        ],
        "extra_ts_fields": [],
    },
    "FreteIntBidFornecedoresAvaliacoes": {
        "entity": "Avaliacao",
        "fields": [
            ("id", "String", False, "@default(cuid())", "id_avaliacao_bid", "sistema", "Não", "", "", "", "", "Identificador único da avaliação", "CUID", "", "", "clxyz123...", "", "Sistema"),
            ("id_organizacao", "String", False, None, "id_organizacao", "sistema", "Não", "", "", "", "", "ID da organização (tenant)", "CUID", "", "", "clorg456...", "", "Sistema"),
            ("product_id", "String", True, None, "id_produto", "sistema", "Não", "Prisma tem product_id, types.ts não tem", "", "", "", "ID do produto", "CUID", "", "", "clprod...", "", "Sistema"),
            ("user_id", "String", False, None, "id_usuario_avaliador", "sistema", "Não", "Prisma tem user_id, types.ts não tem", "", "", "", "ID do usuário que avaliou", "CUID", "", "", "clusr...", "", "Sistema"),
            ("fornecedor_id", "String", False, None, "id_fornecedor_avaliacao_bid", "sistema", "Não", "", "", "", "", "FK para o fornecedor avaliado", "CUID", "", "", "clforn...", "", "Sistema"),
            ("cotacao_id", "String", True, None, "id_cotacao_avaliacao_bid", "sistema", "Não", "", "", "", "", "FK para a cotação (opcional)", "CUID", "", "", "clcot...", "", "Sistema"),
            ("nota_frete", "Int", True, None, "nota_frete_avaliacao_bid", "usuario", "Sim", "", "Modal Avaliação", "Nota Frete", "Nota do Frete", "Nota para o preço do frete (1-5)", "Inteiro 1-5", "1 a 5", "", "4", "InputTextoGlobal", "Usuario"),
            ("nota_atendimento", "Int", True, None, "nota_atendimento_avaliacao_bid", "usuario", "Sim", "", "Modal Avaliação", "Nota Atendimento", "Nota de Atendimento", "Nota para atendimento (1-5)", "Inteiro 1-5", "1 a 5", "", "5", "InputTextoGlobal", "Usuario"),
            ("nota_resposta", "Int", True, None, "nota_resposta_avaliacao_bid", "usuario", "Sim", "DIVERGÊNCIA: types.ts usa 'nota_prazo', prisma usa 'nota_resposta'", "Modal Avaliação", "Nota Prazo", "Nota de Prazo de Resposta", "Nota para prazo de resposta (1-5)", "Inteiro 1-5", "1 a 5", "", "3", "InputTextoGlobal", "Usuario"),
            ("nota_confiabilidade", "Int", True, None, "nota_confiabilidade_avaliacao_bid", "usuario", "Sim", "", "Modal Avaliação", "Nota Confiabilidade", "Nota de Confiabilidade", "Nota para confiabilidade (1-5)", "Inteiro 1-5", "1 a 5", "", "4", "InputTextoGlobal", "Usuario"),
            ("nota_geral", "Float", True, None, "nota_geral_avaliacao_bid", "sistema", "Não", "DIVERGÊNCIA: types.ts usa 'nota_global', prisma usa 'nota_geral'. Média calculada", "Detalhe Fornecedor", "Nota Global", "Nota Geral", "Média ponderada das notas", "Decimal 0-5", "", "", "4.0", "", "Calculado"),
            ("comentario", "String", True, None, "comentario_avaliacao_bid", "usuario", "Sim", "", "Modal Avaliação", "Comentário", "Comentário", "Comentário livre sobre o fornecedor", "Texto livre", "", "", "Excelente atendimento", "InputTextoGlobal", "Usuario"),
            ("created_at", "DateTime", False, "@default(now())", "criado_em", "sistema", "Não", "", "", "", "", "Data de criação do registro", "ISO 8601", "", "", "2026-01-15T10:00:00Z", "", "Sistema"),
            ("updated_at", "DateTime", False, "@updatedAt", "atualizado_em", "sistema", "Não", "Apenas no prisma, types.ts não tem", "", "", "", "Data de última atualização", "ISO 8601", "", "", "2026-01-15T10:00:00Z", "", "Sistema"),
        ],
        "extra_ts_fields": [],
    },
    "FreteIntBidClassificacaoFornecedores": {
        "entity": "ClassificacaoFornecedor",
        "fields": [
            ("id", "String", False, "@default(cuid())", "id_classificacao_fornecedor_bid", "sistema", "Não", "", "", "", "", "Identificador único da classificação", "CUID", "", "", "clxyz123...", "", "Sistema"),
            ("fornecedor_email", "String", False, None, "email_fornecedor_classificacao_bid", "sistema", "Não", "Chave única. Não tem id_organizacao — classificação é cross-org", "Ranking Fornecedores", "E-mail", "E-mail do Fornecedor", "E-mail único do fornecedor (cross-org)", "email", "Unique", "", "contato@maersk.com", "", "Sistema"),
            ("total_cotacoes_recebidas", "Int", False, "@default(0)", "total_cotacoes_recebidas_classificacao_bid", "sistema", "Não", "", "Ranking Fornecedores", "Cotações Recebidas", "Total Cotações Recebidas", "Total de cotações recebidas", "Inteiro", "Min 0", "0", "150", "", "Calculado"),
            ("total_cotacoes_respondidas", "Int", False, "@default(0)", "total_cotacoes_respondidas_classificacao_bid", "sistema", "Não", "", "Ranking Fornecedores", "Cotações Respondidas", "Total Cotações Respondidas", "Total de cotações respondidas", "Inteiro", "Min 0", "0", "128", "", "Calculado"),
            ("total_cotacoes_aprovadas", "Int", False, "@default(0)", "total_cotacoes_aprovadas_classificacao_bid", "sistema", "Não", "", "Ranking Fornecedores", "Cotações Aprovadas", "Total Cotações Aprovadas", "Total de cotações aprovadas", "Inteiro", "Min 0", "0", "95", "", "Calculado"),
            ("taxa_resposta", "Float", False, "@default(0)", "taxa_resposta_classificacao_bid", "sistema", "Não", "", "Ranking Fornecedores", "Taxa Resposta", "Taxa de Resposta (%)", "Percentual de resposta", "Decimal 0-100", "Min 0 Max 100", "0", "85.3", "", "Calculado"),
            ("taxa_aprovacao", "Float", False, "@default(0)", "taxa_aprovacao_classificacao_bid", "sistema", "Não", "", "Ranking Fornecedores", "Taxa Aprovação", "Taxa de Aprovação (%)", "Percentual de aprovação", "Decimal 0-100", "Min 0 Max 100", "0", "74.2", "", "Calculado"),
            ("tempo_medio_resposta_horas", "Float", False, "@default(0)", "tempo_medio_resposta_horas_classificacao_bid", "sistema", "Não", "", "Ranking Fornecedores", "Tempo Médio", "Tempo Médio Resposta (h)", "Tempo médio de resposta em horas", "Decimal", "Min 0", "0", "12.5", "", "Calculado"),
            ("aderencia_target", "Float", False, "@default(0)", "aderencia_target_classificacao_bid", "sistema", "Não", "", "Ranking Fornecedores", "Aderência Target", "Aderência ao Target (%)", "Percentual de aderência ao valor alvo", "Decimal 0-100", "Min 0 Max 100", "0", "92.1", "", "Calculado"),
            ("rating_global", "Float", False, "@default(0)", "rating_global_classificacao_bid", "sistema", "Não", "", "Ranking Fornecedores", "Rating", "Rating Global", "Nota global do fornecedor", "Decimal 0-5", "Min 0 Max 5", "0", "4.2", "", "Calculado"),
            ("media_frete", "Float", False, "@default(0)", "media_frete_classificacao_bid", "sistema", "Não", "", "Ranking Fornecedores", "Média Frete", "Média Nota Frete", "Média das notas de frete", "Decimal 0-5", "Min 0 Max 5", "0", "4.1", "", "Calculado"),
            ("media_atendimento", "Float", False, "@default(0)", "media_atendimento_classificacao_bid", "sistema", "Não", "", "Ranking Fornecedores", "Média Atendimento", "Média Nota Atendimento", "Média das notas de atendimento", "Decimal 0-5", "Min 0 Max 5", "0", "4.5", "", "Calculado"),
            ("media_resposta", "Float", False, "@default(0)", "media_resposta_classificacao_bid", "sistema", "Não", "", "Ranking Fornecedores", "Média Resposta", "Média Nota Resposta", "Média das notas de resposta", "Decimal 0-5", "Min 0 Max 5", "0", "3.8", "", "Calculado"),
            ("media_confiabilidade", "Float", False, "@default(0)", "media_confiabilidade_classificacao_bid", "sistema", "Não", "", "Ranking Fornecedores", "Média Confiabilidade", "Média Nota Confiabilidade", "Média das notas de confiabilidade", "Decimal 0-5", "Min 0 Max 5", "0", "4.3", "", "Calculado"),
            ("total_avaliacoes", "Int", False, "@default(0)", "total_avaliacoes_classificacao_bid", "sistema", "Não", "", "Ranking Fornecedores", "Total Avaliações", "Total de Avaliações", "Quantidade total de avaliações", "Inteiro", "Min 0", "0", "42", "", "Calculado"),
            ("updated_at", "DateTime", False, "@updatedAt", "atualizado_em", "sistema", "Não", "", "", "", "", "Data de última atualização", "ISO 8601", "", "", "2026-01-15T10:00:00Z", "", "Sistema"),
        ],
        "extra_ts_fields": [],
    },
    "FreteIntBidGanhoEstimado": {
        "entity": "GanhoEstimado",
        "fields": [
            ("id", "String", False, "@default(cuid())", "id_ganho_estimado_bid", "sistema", "Não", "", "", "", "", "Identificador único", "CUID", "", "", "clxyz123...", "", "Sistema"),
            ("id_organizacao", "String", False, None, "id_organizacao", "sistema", "Não", "", "", "", "", "ID da organização (tenant)", "CUID", "", "", "clorg456...", "", "Sistema"),
            ("product_id", "String", True, None, "id_produto", "sistema", "Não", "", "", "", "", "ID do produto", "CUID", "", "", "clprod...", "", "Sistema"),
            ("user_id", "String", False, None, "id_usuario_criador", "sistema", "Não", "", "", "", "", "ID do usuário", "CUID", "", "", "clusr...", "", "Sistema"),
            ("cotacao_id", "String", False, None, "id_cotacao_ganho_estimado_bid", "sistema", "Não", "", "", "", "", "FK para a cotação", "CUID", "", "", "clcot...", "", "Sistema"),
            ("company_id", "String", True, None, "id_workspace", "sistema", "Não", "", "", "", "", "ID do workspace", "CUID", "", "", "clcomp...", "", "Sistema"),
            ("valor_target", "Float", True, None, "valor_target_ganho_estimado_bid", "sistema", "Não", "", "Dashboard", "Valor Target", "Valor Target", "Valor alvo da cotação", "Decimal", "", "", "5000.00", "", "Sistema"),
            ("valor_aprovado", "Float", False, None, "valor_aprovado_ganho_estimado_bid", "sistema", "Não", "", "Dashboard", "Valor Aprovado", "Valor Aprovado", "Valor da proposta aprovada", "Decimal", "", "", "4500.00", "", "Sistema"),
            ("valor_medio", "Float", True, None, "valor_medio_ganho_estimado_bid", "sistema", "Não", "", "Dashboard", "Valor Médio", "Valor Médio", "Média de todas as propostas", "Decimal", "", "", "4800.00", "", "Calculado"),
            ("saving_vs_target", "Float", True, None, "saving_vs_target_ganho_estimado_bid", "sistema", "Não", "", "Dashboard", "Saving vs Target", "Saving vs Target", "Economia vs valor alvo", "Decimal", "", "", "500.00", "", "Calculado"),
            ("saving_vs_media", "Float", True, None, "saving_vs_media_ganho_estimado_bid", "sistema", "Não", "", "Dashboard", "Saving vs Média", "Saving vs Média", "Economia vs média das propostas", "Decimal", "", "", "300.00", "", "Calculado"),
            ("saving_percentual", "Float", True, None, "saving_percentual_ganho_estimado_bid", "sistema", "Não", "", "Dashboard", "Saving (%)", "Saving Percentual", "Percentual de economia", "Decimal 0-100", "", "", "10.0", "", "Calculado"),
            ("moeda", "String", False, "@default(\"USD\")", "moeda_ganho_estimado_bid", "sistema", "Não", "", "Dashboard", "Moeda", "Moeda", "Moeda dos valores", "ISO 4217", "", "USD", "USD", "", "Sistema"),
            ("created_at", "DateTime", False, "@default(now())", "criado_em", "sistema", "Não", "", "", "", "", "Data de criação do registro", "ISO 8601", "", "", "2026-01-15T10:00:00Z", "", "Sistema"),
        ],
        "extra_ts_fields": [],
    },
    "FreteIntBidIntegracoes": {
        "entity": "Integracao",
        "fields": [
            ("id", "String", False, "@default(cuid())", "id_integracao_bid", "sistema", "Não", "", "", "", "", "Identificador único da integração", "CUID", "", "", "clxyz123...", "", "Sistema"),
            ("id_organizacao", "String", False, None, "id_organizacao", "sistema", "Não", "", "", "", "", "ID da organização (tenant)", "CUID", "", "", "clorg456...", "", "Sistema"),
            ("product_id", "String", True, None, "id_produto", "sistema", "Não", "", "", "", "", "ID do produto", "CUID", "", "", "clprod...", "", "Sistema"),
            ("user_id", "String", True, None, "id_usuario_criador", "sistema", "Não", "", "", "", "", "ID do usuário que configurou", "CUID", "", "", "clusr...", "", "Sistema"),
            ("fornecedor_id", "String", True, None, "id_fornecedor_integracao_bid", "sistema", "Não", "", "Configurações", "Fornecedor", "Fornecedor", "FK para fornecedor (opcional)", "CUID", "", "", "clforn...", "SelectGlobal", "Usuario"),
            ("nome", "String", False, None, "nome_integracao_bid", "usuario", "Sim", "", "Configurações", "Nome", "Nome da Integração", "Nome/label da integração", "Texto livre", "Min 2 chars", "", "API Maersk", "InputTextoGlobal", "Usuario"),
            ("tipo", "BidFreteIntegracao", False, None, "tipo_integracao_bid", "usuario", "Sim", "", "Configurações", "Tipo", "Tipo de Integração", "Tipo da integração", "Enum", "Valor do enum", "", "API_REST", "SelectGlobal", "Usuario"),
            ("ativo", "Boolean", False, "@default(false)", "ativo_integracao_bid", "usuario", "Sim", "", "Configurações", "Ativo", "Ativo", "Se a integração está ativa", "Boolean", "", "false", "true", "Checkbox", "Usuario"),
            ("base_url", "String", True, None, "base_url_integracao_bid", "usuario", "Sim", "", "Configurações", "Base URL", "URL Base", "URL base da API", "URL", "Formato URL", "", "https://api.maersk.com/v2", "InputTextoGlobal", "Usuario"),
            ("api_key_hash", "String", True, None, "api_key_hash_integracao_bid", "sistema", "Não", "", "", "", "", "Hash da API key (nunca expor raw)", "Hash SHA-256", "", "", "a1b2c3...", "", "Sistema"),
            ("auth_type", "String", True, None, "tipo_autenticacao_integracao_bid", "usuario", "Sim", "", "Configurações", "Tipo Auth", "Tipo de Autenticação", "Tipo de autenticação (Bearer, Basic, API Key)", "Texto", "", "", "Bearer", "SelectGlobal", "Usuario"),
            ("headers_extra", "String", True, None, "headers_extra_integracao_bid", "usuario", "Sim", "", "Configurações", "Headers", "Headers Extras", "Headers customizados (JSON)", "JSON string", "JSON válido", "", '{"X-Custom":"value"}', "InputTextoGlobal", "Usuario"),
            ("config_extra", "String", True, None, "config_extra_integracao_bid", "usuario", "Sim", "", "Configurações", "Config Extra", "Configuração Extra", "Configurações adicionais (JSON)", "JSON string", "JSON válido", "", '{"timeout":30000}', "InputTextoGlobal", "Usuario"),
            ("ultimo_teste_em", "DateTime", True, None, "ultimo_teste_em_integracao_bid", "sistema", "Não", "", "Configurações", "Último Teste", "Data Último Teste", "Data do último teste de conexão", "ISO 8601", "", "", "2026-01-15T10:00:00Z", "", "Sistema"),
            ("ultimo_teste_ok", "Boolean", True, None, "ultimo_teste_ok_integracao_bid", "sistema", "Não", "", "Configurações", "Teste OK", "Último Teste OK", "Se o último teste foi bem-sucedido", "Boolean", "", "", "true", "BadgeGlobal", "Sistema"),
            ("erro_ultimo_teste", "String", True, None, "erro_ultimo_teste_integracao_bid", "sistema", "Não", "", "Configurações", "Erro", "Erro do Último Teste", "Mensagem de erro do último teste", "Texto", "", "", "Connection timeout", "", "Sistema"),
            ("created_at", "DateTime", False, "@default(now())", "criado_em", "sistema", "Não", "", "", "", "", "Data de criação do registro", "ISO 8601", "", "", "2026-01-15T10:00:00Z", "", "Sistema"),
            ("updated_at", "DateTime", False, "@updatedAt", "atualizado_em", "sistema", "Não", "", "", "", "", "Data de última atualização", "ISO 8601", "", "", "2026-01-15T10:00:00Z", "", "Sistema"),
        ],
        "extra_ts_fields": [],
    },
    "FreteIntBidPortosCadastro": {
        "entity": "Porto",
        "fields": [
            ("codigo", "String", False, "@id", "codigo_porto_bid", "sistema", "Não", "Chave primária natural (código do porto). Sem id_organizacao — tabela de referência global", "Master Data Portos", "Código", "Código do Porto", "Código UNLOCODE do porto", "Texto", "5 chars", "", "BRSSZ", "", "Sistema"),
            ("nome", "String", False, None, "nome_porto_bid", "sistema", "Não", "", "Master Data Portos", "Nome", "Nome do Porto", "Nome do porto/aeroporto", "Texto", "", "", "Santos", "", "Sistema"),
            ("pais", "String", False, None, "pais_porto_bid", "sistema", "Não", "", "Master Data Portos", "País", "País", "Nome do país", "Texto", "", "", "Brasil", "", "Sistema"),
            ("pais_codigo", "String", False, None, "codigo_pais_porto_bid", "sistema", "Não", "", "Master Data Portos", "Código País", "Código do País", "Código ISO do país", "ISO 3166-1 alpha-2", "2 chars", "", "BR", "", "Sistema"),
            ("tipo", "String", False, None, "tipo_porto_bid", "sistema", "Não", "", "Master Data Portos", "Tipo", "Tipo do Porto", "Tipo (porto, aeroporto, terminal)", "Texto", "", "", "porto", "", "Sistema"),
            ("latitude", "Float", True, None, "latitude_porto_bid", "sistema", "Não", "", "", "", "", "Latitude do porto", "Decimal", "-90 a 90", "", "-23.9544", "", "Sistema"),
            ("longitude", "Float", True, None, "longitude_porto_bid", "sistema", "Não", "", "", "", "", "Longitude do porto", "Decimal", "-180 a 180", "", "-46.3033", "", "Sistema"),
            ("ativo", "Boolean", False, "@default(true)", "ativo_porto_bid", "sistema", "Não", "", "Master Data Portos", "Ativo", "Ativo", "Se o porto está ativo", "Boolean", "", "true", "true", "", "Sistema"),
        ],
        "extra_ts_fields": [],
    },
}

# ── Enums ──────────────────────────────────────────────────────────────

ENUMS = {
    "BidFreteTipoOperacao": {
        "values": [
            ("IMPORTACAO", "Importação", "Import", "Operação de importação"),
            ("EXPORTACAO", "Exportação", "Export", "Operação de exportação"),
        ],
        "used_in_model": "FreteIntBidCotacoes",
        "used_in_field": "tipo_operacao",
        "categoria": "Operação",
    },
    "BidFreteModalidade": {
        "values": [
            ("MARITIMO", "Marítimo", "Maritime", "Transporte por via marítima"),
            ("AEREO", "Aéreo", "Air", "Transporte por via aérea"),
            ("RODOVIARIO", "Rodoviário", "Road", "Transporte por via rodoviária"),
        ],
        "used_in_model": "FreteIntBidCotacoes, FreteIntBidTabelasProntas",
        "used_in_field": "modal",
        "categoria": "Modal",
    },
    "BidFreteCargaModalidade": {
        "values": [
            ("FCL", "FCL (Container Cheio)", "FCL (Full Container Load)", "Container cheio"),
            ("LCL", "LCL (Carga Consolidada)", "LCL (Less Container Load)", "Carga consolidada em container"),
            ("AEREO_GERAL", "Aéreo Geral", "General Air Cargo", "Carga aérea geral"),
            ("RODOVIARIO_FTL", "Rodoviário FTL (Carga Completa)", "FTL (Full Truck Load)", "Caminhão completo"),
            ("RODOVIARIO_LTL", "Rodoviário LTL (Carga Fracionada)", "LTL (Less Truck Load)", "Carga fracionada"),
        ],
        "used_in_model": "FreteIntBidCotacoes, FreteIntBidTabelasProntas",
        "used_in_field": "modalidade",
        "categoria": "Carga",
    },
    "BidFreteCotacaoStatus": {
        "values": [
            ("RASCUNHO", "Rascunho", "Draft", "Cotação em rascunho, não enviada"),
            ("ENVIADA_FORNECEDORES", "Enviada p/ Fornecedores", "Sent to Suppliers", "Cotação enviada aos fornecedores selecionados"),
            ("EM_COTACAO", "Em Cotação", "Quoting", "Fornecedores estão respondendo"),
            ("AGUARDANDO_APROVACAO", "Aguardando Aprovação", "Pending Approval", "Propostas recebidas, aguardando decisão"),
            ("APROVADA", "Aprovada", "Approved", "Proposta vencedora aprovada"),
            ("REPROVADA", "Reprovada", "Rejected", "Todas as propostas reprovadas"),
            ("CANCELADA", "Cancelada", "Cancelled", "Cotação cancelada"),
            ("FALTA_INFORMACAO", "Falta Informação", "Missing Info", "Informações insuficientes para prosseguir"),
            ("EXPIRADA", "Expirada", "Expired", "Prazo de resposta expirou"),
        ],
        "used_in_model": "FreteIntBidCotacoes",
        "used_in_field": "status",
        "categoria": "Status Cotação",
    },
    "BidFreteCotacaoFreteIntCanal": {
        "values": [
            ("EMAIL", "E-mail", "Email", "Envio por e-mail"),
            ("WHATSAPP", "WhatsApp", "WhatsApp", "Envio por WhatsApp"),
            ("API", "API", "API", "Envio via integração API"),
            ("PORTAL", "Portal", "Portal", "Resposta via portal do fornecedor"),
        ],
        "used_in_model": "FreteIntBidPedidoCotacoes",
        "used_in_field": "canal",
        "categoria": "Canal",
    },
    "BidFreteCotacao": {
        "values": [
            ("PENDENTE", "Pendente", "Pending", "Pedido aguardando envio"),
            ("ENVIADO", "Enviado", "Sent", "Pedido enviado ao fornecedor"),
            ("VISUALIZADO", "Visualizado", "Viewed", "Fornecedor visualizou o pedido"),
            ("RESPONDIDO", "Respondido", "Responded", "Fornecedor respondeu"),
            ("EXPIRADO", "Expirado", "Expired", "Token expirou sem resposta"),
            ("ERRO_ENVIO", "Erro no Envio", "Send Error", "Falha ao enviar o pedido"),
        ],
        "used_in_model": "FreteIntBidPedidoCotacoes",
        "used_in_field": "status",
        "categoria": "Status Pedido Cotação",
    },
    "BidFretePropostaStatus": {
        "values": [
            ("RECEBIDA", "Recebida", "Received", "Proposta recebida, pendente análise"),
            ("EM_ANALISE", "Em Análise", "Under Review", "Proposta em análise comparativa"),
            ("MELHOR_PRECO", "Melhor Preço", "Best Price", "Melhor preço entre as propostas"),
            ("MELHOR_TRANSIT", "Melhor Transit Time", "Best Transit", "Menor tempo de trânsito"),
            ("MELHOR_AVALIACAO", "Melhor Avaliação", "Best Rating", "Melhor avaliação do fornecedor"),
            ("APROVADA", "Aprovada", "Approved", "Proposta aprovada como vencedora"),
            ("REPROVADA", "Reprovada", "Rejected", "Proposta reprovada"),
        ],
        "used_in_model": "FreteIntBidPropostas",
        "used_in_field": "status",
        "categoria": "Status Proposta",
    },
    "BidFreteTipoFornecedor": {
        "values": [
            ("AGENTE_CARGA", "Agente de Carga", "Freight Forwarder", "Agente/broker de carga"),
            ("ARMADOR", "Armador", "Shipping Line", "Companhia marítima / armador"),
            ("CIA_AEREA", "Companhia Aérea", "Airline", "Companhia aérea de carga"),
            ("TRANSPORTADORA", "Transportadora", "Trucking Company", "Transportadora rodoviária"),
        ],
        "used_in_model": "FreteIntBidFornecedores",
        "used_in_field": "tipo",
        "categoria": "Tipo Fornecedor",
    },
    "BidFreteStatusFornecedor": {
        "values": [
            ("ATIVO", "Ativo", "Active", "Fornecedor ativo e disponível"),
            ("INATIVO", "Inativo", "Inactive", "Fornecedor inativado"),
            ("PENDENTE_APROVACAO", "Pendente Aprovação", "Pending Approval", "Cadastro pendente de aprovação"),
            ("BLOQUEADO", "Bloqueado", "Blocked", "Fornecedor bloqueado"),
        ],
        "used_in_model": "FreteIntBidFornecedores",
        "used_in_field": "status",
        "categoria": "Status Fornecedor",
    },
    "BidFreteCotacaoVisibilidade": {
        "values": [
            ("DIRECIONADA", "Direcionada", "Directed", "Cotação enviada a fornecedores selecionados"),
            ("ABERTA", "Aberta", "Open", "Cotação aberta a todos os fornecedores"),
        ],
        "used_in_model": "FreteIntBidCotacoes",
        "used_in_field": "visibilidade",
        "categoria": "Visibilidade",
    },
    "BidFreteIntegracao": {
        "values": [
            ("API_REST", "API REST", "REST API", "Integração via API RESTful"),
            ("API_SOAP", "API SOAP", "SOAP API", "Integração via API SOAP/XML"),
            ("ODATA", "OData", "OData", "Integração via protocolo OData"),
            ("MANUAL", "Manual", "Manual", "Sem integração, entrada manual"),
        ],
        "used_in_model": "FreteIntBidIntegracoes",
        "used_in_field": "tipo",
        "categoria": "Tipo Integração",
    },
}

# ── Routes ──────────────────────────────────────────────────────────────

ROUTES = [
    # (method, prefix, full_route, ddd_route, router_var, file, middleware, auth, zod_schema, entity, action, query_params, body_params, response_type, status_code, pagination, filters, status_ddd, obs)
    # cotacoes.ts
    ("POST", "/bid-frete", "/bid-frete/cotacoes", "/bid-frete/cotacoes", "cotacoesRouter", "cotacoes.ts", "authMiddleware, tenantMiddleware", "JWT + Clerk", "criarCotacaoSchema", "Cotacao", "Criar", "", "tipo_operacao, modal, modalidade, origem_*, destino_*, descricao_mercadoria, incoterm, ...", "Cotacao", "201", "Não", "", "Atual", ""),
    ("GET", "/bid-frete", "/bid-frete/cotacoes", "/bid-frete/cotacoes", "cotacoesRouter", "cotacoes.ts", "authMiddleware, tenantMiddleware", "JWT + Clerk", "", "Cotacao[]", "Listar", "status, page, limit, busca", "", "Cotacao[] + paginação", "200", "Sim", "status, busca (texto livre)", "Atual", ""),
    ("GET", "/bid-frete", "/bid-frete/cotacoes/:id", "/bid-frete/cotacoes/:id_cotacao_bid", "cotacoesRouter", "cotacoes.ts", "authMiddleware, tenantMiddleware", "JWT + Clerk", "", "Cotacao", "Detalhe", "", "", "Cotacao + relações", "200", "Não", "", "Atual", "Inclui fornecedores e propostas"),
    ("PATCH", "/bid-frete", "/bid-frete/cotacoes/:id", "/bid-frete/cotacoes/:id_cotacao_bid", "cotacoesRouter", "cotacoes.ts", "authMiddleware, tenantMiddleware", "JWT + Clerk", "atualizarCotacaoSchema", "Cotacao", "Atualizar", "", "campos parciais da cotação", "Cotacao", "200", "Não", "", "Atual", ""),
    ("DELETE", "/bid-frete", "/bid-frete/cotacoes/:id", "/bid-frete/cotacoes/:id_cotacao_bid", "cotacoesRouter", "cotacoes.ts", "authMiddleware, tenantMiddleware", "JWT + Clerk", "", "Cotacao", "Excluir", "", "", "", "204", "Não", "", "Atual", "Soft delete ou hard delete"),
    ("PATCH", "/bid-frete", "/bid-frete/cotacoes/:id/status", "/bid-frete/cotacoes/:id_cotacao_bid/status", "cotacoesRouter", "cotacoes.ts", "authMiddleware, tenantMiddleware", "JWT + Clerk", "mudarStatusSchema", "Cotacao", "Mudar Status", "", "status, motivo_reprovacao?, motivo_cancelamento?", "Cotacao", "200", "Não", "", "Atual", ""),
    ("POST", "/bid-frete", "/bid-frete/cotacoes/bloco", "/bid-frete/cotacoes/bloco", "cotacoesRouter", "cotacoes.ts", "authMiddleware, tenantMiddleware", "JWT + Clerk", "", "Cotacao[]", "Importar Bloco", "", "arquivo CSV (multipart)", "{ importadas: number, erros: [] }", "200", "Não", "", "Atual", "Upload CSV"),
    # fornecedores.ts
    ("POST", "/bid-frete", "/bid-frete/fornecedores", "/bid-frete/fornecedores", "fornecedoresRouter", "fornecedores.ts", "authMiddleware, tenantMiddleware", "JWT + Clerk", "criarFornecedorSchema", "Fornecedor", "Criar", "", "nome, tipo, email, ...", "Fornecedor", "201", "Não", "", "Atual", ""),
    ("GET", "/bid-frete", "/bid-frete/fornecedores", "/bid-frete/fornecedores", "fornecedoresRouter", "fornecedores.ts", "authMiddleware, tenantMiddleware", "JWT + Clerk", "", "Fornecedor[]", "Listar", "page, limit, busca, status, tipo", "", "Fornecedor[] + paginação", "200", "Sim", "status, tipo, busca", "Atual", ""),
    ("GET", "/bid-frete", "/bid-frete/fornecedores/:id", "/bid-frete/fornecedores/:id_fornecedor_bid", "fornecedoresRouter", "fornecedores.ts", "authMiddleware, tenantMiddleware", "JWT + Clerk", "", "Fornecedor", "Detalhe", "", "", "Fornecedor + classificação + avaliações", "200", "Não", "", "Atual", "Inclui dados agregados"),
    ("PUT", "/bid-frete", "/bid-frete/fornecedores/:id", "/bid-frete/fornecedores/:id_fornecedor_bid", "fornecedoresRouter", "fornecedores.ts", "authMiddleware, tenantMiddleware", "JWT + Clerk", "atualizarFornecedorSchema", "Fornecedor", "Atualizar", "", "campos do fornecedor", "Fornecedor", "200", "Não", "", "Atual", "PUT full replace"),
    ("DELETE", "/bid-frete", "/bid-frete/fornecedores/:id", "/bid-frete/fornecedores/:id_fornecedor_bid", "fornecedoresRouter", "fornecedores.ts", "authMiddleware, tenantMiddleware", "JWT + Clerk", "", "Fornecedor", "Excluir", "", "", "", "204", "Não", "", "Atual", ""),
    ("PATCH", "/bid-frete", "/bid-frete/fornecedores/:id/status", "/bid-frete/fornecedores/:id_fornecedor_bid/status", "fornecedoresRouter", "fornecedores.ts", "authMiddleware, tenantMiddleware", "JWT + Clerk", "mudarStatusFornecedorSchema", "Fornecedor", "Mudar Status", "", "status", "Fornecedor", "200", "Não", "", "Atual", ""),
    ("POST", "/bid-frete", "/bid-frete/fornecedores/:id/tabela", "/bid-frete/fornecedores/:id_fornecedor_bid/tabela-preco", "fornecedoresRouter", "fornecedores.ts", "authMiddleware, tenantMiddleware", "JWT + Clerk", "criarTabelaSchema", "TabelaPreco", "Criar", "", "origem_*, destino_*, modal, valores, validade_*", "TabelaPreco", "201", "Não", "", "Atual", ""),
    ("GET", "/bid-frete", "/bid-frete/fornecedores/:id/tabela", "/bid-frete/fornecedores/:id_fornecedor_bid/tabela-preco", "fornecedoresRouter", "fornecedores.ts", "authMiddleware, tenantMiddleware", "JWT + Clerk", "", "TabelaPreco[]", "Listar Tabelas", "", "", "TabelaPreco[]", "200", "Não", "", "Atual", ""),
    ("PUT", "/bid-frete", "/bid-frete/fornecedores/:id/tabela/:tpId", "/bid-frete/fornecedores/:id_fornecedor_bid/tabela-preco/:id_tabela_preco_bid", "fornecedoresRouter", "fornecedores.ts", "authMiddleware, tenantMiddleware", "JWT + Clerk", "atualizarTabelaSchema", "TabelaPreco", "Atualizar Tabela", "", "campos da tabela", "TabelaPreco", "200", "Não", "", "Atual", ""),
    ("DELETE", "/bid-frete", "/bid-frete/fornecedores/:id/tabela/:tpId", "/bid-frete/fornecedores/:id_fornecedor_bid/tabela-preco/:id_tabela_preco_bid", "fornecedoresRouter", "fornecedores.ts", "authMiddleware, tenantMiddleware", "JWT + Clerk", "", "TabelaPreco", "Excluir Tabela", "", "", "", "204", "Não", "", "Atual", ""),
    ("POST", "/bid-frete", "/bid-frete/fornecedores/:id/avaliar", "/bid-frete/fornecedores/:id_fornecedor_bid/avaliar", "fornecedoresRouter", "fornecedores.ts", "authMiddleware, tenantMiddleware", "JWT + Clerk", "criarAvaliacaoSchema", "Avaliacao", "Criar Avaliação", "", "notas (frete, atendimento, resposta, confiabilidade) + comentario", "Avaliacao", "201", "Não", "", "Atual", ""),
    # bids.ts
    ("POST", "/bid-frete", "/bid-frete/bids/disparar", "/bid-frete/bids/disparar", "bidsRouter", "bids.ts", "authMiddleware, tenantMiddleware", "JWT + Clerk", "dispararBidSchema", "PedidoCotacao", "Disparar Bid", "", "cotacao_id, fornecedores: [{id, canal}]", "PedidoCotacao[]", "201", "Não", "", "Atual", "Cria múltiplos PedidoCotacao"),
    ("GET", "/bid-frete", "/bid-frete/bids/cotacao/:cotacaoId", "/bid-frete/bids/cotacao/:id_cotacao_bid", "bidsRouter", "bids.ts", "authMiddleware, tenantMiddleware", "JWT + Clerk", "", "PedidoCotacao[]", "Listar Bids", "", "", "PedidoCotacao[] + fornecedor", "200", "Não", "", "Atual", ""),
    ("POST", "/bid-frete", "/bid-frete/bids/cotacao-aberta", "/bid-frete/bids/cotacao-aberta", "bidsRouter", "bids.ts", "authMiddleware, tenantMiddleware", "JWT + Clerk", "cotacaoAbertaSchema", "Cotacao", "Cotação Aberta", "", "cotacao_id", "Cotacao", "201", "Não", "", "Atual", "Publica cotação para todos"),
    # comparativo.ts
    ("GET", "/bid-frete", "/bid-frete/comparativo/:cotacaoId/ranking", "/bid-frete/comparativo/:id_cotacao_bid/ranking", "comparativoRouter", "comparativo.ts", "authMiddleware, tenantMiddleware", "JWT + Clerk", "", "Proposta[]", "Ranking", "", "", "Proposta[] ordenadas por ranking", "200", "Não", "", "Atual", "Inclui rankings calculados"),
    ("GET", "/bid-frete", "/bid-frete/comparativo/:cotacaoId/analise-ia", "/bid-frete/comparativo/:id_cotacao_bid/analise-ia", "comparativoRouter", "comparativo.ts", "authMiddleware, tenantMiddleware", "JWT + Clerk", "", "AnaliseIA", "Análise IA", "", "", "{ recomendacao, justificativa, riscos }", "200", "Não", "", "Atual", "Stub — não implementado"),
    ("POST", "/bid-frete", "/bid-frete/comparativo/:cotacaoId/aprovar", "/bid-frete/comparativo/:id_cotacao_bid/aprovar", "comparativoRouter", "comparativo.ts", "authMiddleware, tenantMiddleware", "JWT + Clerk", "aprovarSchema", "Proposta", "Aprovar", "", "proposta_id", "Proposta + GanhoEstimado", "200", "Não", "", "Atual", "Aprova proposta e calcula saving"),
    ("POST", "/bid-frete", "/bid-frete/comparativo/:cotacaoId/reprovar", "/bid-frete/comparativo/:id_cotacao_bid/reprovar", "comparativoRouter", "comparativo.ts", "authMiddleware, tenantMiddleware", "JWT + Clerk", "reprovarSchema", "Cotacao", "Reprovar Todas", "", "motivo_reprovacao", "Cotacao", "200", "Não", "", "Atual", "Reprova todas as propostas"),
    # portal.ts
    ("GET", "/bid-frete", "/bid-frete/portal/pendentes", "/bid-frete/portal/pendentes", "portalRouter", "portal.ts", "portalAuthMiddleware", "JWT Fornecedor", "", "PedidoCotacao[]", "Listar Pendentes", "", "", "PedidoCotacao[] + cotação", "200", "Não", "", "Atual", "Portal do fornecedor"),
    ("POST", "/bid-frete", "/bid-frete/portal/responder/:bidRequestId", "/bid-frete/portal/responder/:id_pedido_cotacao_bid", "portalRouter", "portal.ts", "portalAuthMiddleware", "JWT Fornecedor", "responderBidSchema", "Proposta", "Responder Bid", "", "valor_frete, taxas_*, transit_time, validade, ...", "Proposta", "201", "Não", "", "Atual", "Portal do fornecedor"),
    ("GET", "/bid-frete", "/bid-frete/portal/respostas", "/bid-frete/portal/respostas", "portalRouter", "portal.ts", "portalAuthMiddleware", "JWT Fornecedor", "", "Proposta[]", "Histórico Respostas", "page, limit", "", "Proposta[] + paginação", "200", "Sim", "", "Atual", "Portal do fornecedor"),
    ("GET", "/bid-frete", "/bid-frete/portal/dashboard", "/bid-frete/portal/dashboard", "portalRouter", "portal.ts", "portalAuthMiddleware", "JWT Fornecedor", "", "DashboardPortal", "Dashboard", "", "", "{ kpis, cotacoes_recentes, grafico_respostas }", "200", "Não", "", "Atual", "Portal do fornecedor"),
    ("GET", "/bid-frete", "/bid-frete/portal/desempenho", "/bid-frete/portal/desempenho", "portalRouter", "portal.ts", "portalAuthMiddleware", "JWT Fornecedor", "", "ClassificacaoFornecedor", "Desempenho", "", "", "ClassificacaoFornecedor + histórico", "200", "Não", "", "Atual", "Portal do fornecedor"),
    ("GET", "/bid-frete", "/bid-frete/portal/meu-billing", "/bid-frete/portal/meu-billing", "portalRouter", "portal.ts", "portalAuthMiddleware", "JWT Fornecedor", "", "BillingFornecedor", "Billing", "", "", "{ plano, uso, faturas }", "200", "Não", "", "Atual", "Portal do fornecedor"),
    # cotacoes-publicas.ts
    ("GET", "/bid-frete", "/bid-frete/portal/public/:token", "/bid-frete/portal/public/:token", "publicRouter", "cotacoes-publicas.ts", "Nenhum", "Token (sem auth)", "", "CotacaoPublica", "Ver Cotação Pública", "", "", "Cotacao (dados limitados)", "200", "Não", "", "Atual", "Sem autenticação, via token"),
    ("POST", "/bid-frete", "/bid-frete/portal/public/:token/responder", "/bid-frete/portal/public/:token/responder", "publicRouter", "cotacoes-publicas.ts", "Nenhum", "Token (sem auth)", "responderPublicoSchema", "Proposta", "Responder Público", "", "valor_frete, taxas_*, transit_time, validade, taxas_detalhes[]", "Proposta", "201", "Não", "", "Atual", "Sem autenticação, via token"),
    # avaliacoes.ts
    ("POST", "/bid-frete", "/bid-frete/avaliacoes", "/bid-frete/avaliacoes", "avaliacoesRouter", "avaliacoes.ts", "authMiddleware, tenantMiddleware", "JWT + Clerk", "criarAvaliacaoSchema", "Avaliacao", "Criar", "", "fornecedor_id, cotacao_id?, notas, comentario?", "Avaliacao", "201", "Não", "", "Atual", ""),
    ("GET", "/bid-frete", "/bid-frete/avaliacoes/fornecedor/:id", "/bid-frete/avaliacoes/fornecedor/:id_fornecedor_bid", "avaliacoesRouter", "avaliacoes.ts", "authMiddleware, tenantMiddleware", "JWT + Clerk", "", "Avaliacao[]", "Listar por Fornecedor", "", "", "Avaliacao[]", "200", "Não", "", "Atual", ""),
    ("GET", "/bid-frete", "/bid-frete/avaliacoes/ranking", "/bid-frete/avaliacoes/ranking", "avaliacoesRouter", "avaliacoes.ts", "authMiddleware, tenantMiddleware", "JWT + Clerk", "", "ClassificacaoFornecedor[]", "Ranking Geral", "page, limit", "", "ClassificacaoFornecedor[] ordenados", "200", "Sim", "", "Atual", "Ranking geral de fornecedores"),
    # dashboard.ts
    ("GET", "/bid-frete", "/bid-frete/dashboard/kpis", "/bid-frete/dashboard/kpis", "dashboardRouter", "dashboard.ts", "authMiddleware, tenantMiddleware", "JWT + Clerk", "", "DashboardKPIs", "KPIs", "", "", "{ total_cotacoes, total_saving, taxa_aprovacao, ... }", "200", "Não", "", "Atual", ""),
    ("GET", "/bid-frete", "/bid-frete/dashboard/calendario", "/bid-frete/dashboard/calendario", "dashboardRouter", "dashboard.ts", "authMiddleware, tenantMiddleware", "JWT + Clerk", "", "AlertaCalendario[]", "Calendário", "", "", "AlertaCalendario[] (datas limite)", "200", "Não", "", "Atual", "Alertas de vencimento"),
    ("POST", "/bid-frete", "/bid-frete/dashboard/widgets", "/bid-frete/dashboard/widgets", "dashboardRouter", "dashboard.ts", "authMiddleware, tenantMiddleware", "JWT + Clerk", "widgetsSchema", "Widget[]", "Widgets", "", "layout de widgets", "Widget[]", "200", "Não", "", "Atual", "Widgets customizados"),
    # master-data
    ("GET", "/bid-frete", "/bid-frete/master-data/portos", "/bid-frete/master-data/portos", "masterDataRouter", "master-data.ts", "authMiddleware", "JWT + Clerk", "", "Porto[]", "Listar Portos", "busca, pais, tipo", "", "Porto[]", "200", "Não", "busca, pais, tipo", "Atual", ""),
    ("GET", "/bid-frete", "/bid-frete/master-data/incoterms", "/bid-frete/master-data/incoterms", "masterDataRouter", "master-data.ts", "authMiddleware", "JWT + Clerk", "", "Incoterm[]", "Listar Incoterms", "", "", "Incoterm[] (dados estáticos)", "200", "Não", "", "Atual", "Lista fixa"),
    ("GET", "/bid-frete", "/bid-frete/master-data/modais", "/bid-frete/master-data/modais", "masterDataRouter", "master-data.ts", "authMiddleware", "JWT + Clerk", "", "Modal[]", "Listar Modais", "", "", "Modal[] (do enum)", "200", "Não", "", "Atual", "Derivado do enum"),
    ("GET", "/bid-frete", "/bid-frete/master-data/moedas", "/bid-frete/master-data/moedas", "masterDataRouter", "master-data.ts", "authMiddleware", "JWT + Clerk", "", "Moeda[]", "Listar Moedas", "", "", "Moeda[] (dados estáticos)", "200", "Não", "", "Atual", "Lista fixa"),
    ("GET", "/bid-frete", "/bid-frete/master-data/containers", "/bid-frete/master-data/containers", "masterDataRouter", "master-data.ts", "authMiddleware", "JWT + Clerk", "", "Container[]", "Listar Containers", "", "", "Container[] (dados estáticos)", "200", "Não", "", "Atual", "Lista fixa"),
    ("GET", "/bid-frete", "/bid-frete/master-data/paises", "/bid-frete/master-data/paises", "masterDataRouter", "master-data.ts", "authMiddleware", "JWT + Clerk", "", "Pais[]", "Listar Países", "", "", "Pais[] (dados estáticos)", "200", "Não", "", "Atual", "Lista fixa"),
]

# ── Pages ──────────────────────────────────────────────────────────────

PAGES = [
    # (url, titulo_atual, titulo_ddd, arquivo_atual, arquivo_ddd, tipo, componentes, hooks, api_calls, entidades, status_ddd, obs)
    ("/", "Redirect", "Redirect", "App.tsx", "App.tsx", "Redirect", "", "", "", "", "Atual", "Redireciona para /visao-geral"),
    ("/visao-geral", "Dashboard", "Visão Geral", "Dashboard.tsx", "VisaoGeral.tsx", "Dashboard", "CabecalhoGlobal, PaginaGlobal, BadgeGlobal, BotaoGlobal", "useState, useEffect, useFetch", "GET /dashboard/kpis, GET /dashboard/calendario", "DashboardKPIs, AlertaCalendario", "Atual", "KPIs, calendário, gráficos"),
    ("/cotacoes", "Cotações", "Lista de Cotações", "Cotacoes.tsx", "CotacoesLista.tsx", "Lista + Kanban", "CabecalhoGlobal, PaginaGlobal, TabelaGlobal, BadgeGlobal, TabsGlobal", "useState, useEffect, useFetch", "GET /cotacoes", "Cotacao", "Atual", "Visão lista e kanban"),
    ("/cotacoes/nova", "Nova Cotação", "Nova Cotação", "NovaCotacao.tsx", "NovaCotacao.tsx", "Formulário Wizard", "CabecalhoGlobal, PaginaGlobal, StepperPassoPassoGlobal, InputTextoGlobal, SelectGlobal, BotaoGlobal", "useState, useForm", "POST /cotacoes, GET /master-data/*", "Cotacao", "Atual", "Wizard 7 passos"),
    ("/cotacoes/importar", "Importar Cotações", "Importar Cotações em Bloco", "CotacoesImportar.tsx", "CotacoesImportar.tsx", "Formulário Upload", "CabecalhoGlobal, PaginaGlobal, BotaoGlobal, ModalGlobal", "useState", "POST /cotacoes/bloco", "Cotacao", "Atual", "Upload CSV"),
    ("/cotacoes/:id", "Detalhe Cotação", "Detalhe da Cotação", "DetalheCotacao.tsx", "CotacaoDetalhe.tsx", "Detalhe", "CabecalhoGlobal, PaginaGlobal, BadgeGlobal, BotaoGlobal, TabsGlobal, TabelaGlobal, ModalGlobal", "useState, useEffect, useParams, useFetch", "GET /cotacoes/:id, POST /bids/disparar, PATCH /cotacoes/:id/status", "Cotacao, PedidoCotacao, Fornecedor", "Atual", "Tabs: dados, fornecedores, propostas"),
    ("/cotacoes/:id/comparativo", "Comparativo", "Comparativo de Propostas", "Comparativo.tsx", "Comparativo.tsx", "Análise", "CabecalhoGlobal, PaginaGlobal, TabelaGlobal, BadgeGlobal, BotaoGlobal, ModalGlobal", "useState, useEffect, useParams, useFetch", "GET /comparativo/:id/ranking, POST /comparativo/:id/aprovar, POST /comparativo/:id/reprovar", "Proposta, Cotacao", "Atual", "Ranking, aprovação, rejeição"),
    ("/fornecedores", "Fornecedores", "Lista de Fornecedores", "Fornecedores.tsx", "FornecedoresLista.tsx", "Lista", "CabecalhoGlobal, PaginaGlobal, TabelaGlobal, BadgeGlobal, BotaoGlobal, ModalGlobal", "useState, useEffect, useFetch", "GET /fornecedores, POST /fornecedores", "Fornecedor", "Atual", "Lista com modal de criação"),
    ("/fornecedores/:id", "Detalhe Fornecedor", "Detalhe do Fornecedor", "DetalheFornecedor.tsx", "FornecedorDetalhe.tsx", "Detalhe", "CabecalhoGlobal, PaginaGlobal, BadgeGlobal, BotaoGlobal, TabsGlobal, TabelaGlobal, ModalGlobal", "useState, useEffect, useParams, useFetch", "GET /fornecedores/:id, POST /fornecedores/:id/avaliar, GET /fornecedores/:id/tabela", "Fornecedor, Avaliacao, TabelaPreco, ClassificacaoFornecedor", "Atual", "Tabs: dados, tabelas, avaliações"),
    ("/configuracoes", "Configurações", "Configurações do Bid Frete", "Configuracoes.tsx", "Configuracoes.tsx", "Configuração", "CabecalhoGlobal, PaginaGlobal, TabsGlobal, InputTextoGlobal, SelectGlobal, BotaoGlobal", "useState, useEffect, useFetch", "GET/POST integrações", "Integracao", "Atual", "Configurações e integrações"),
    ("/portal/dashboard", "Portal Dashboard", "Dashboard do Fornecedor", "PortalDashboard.tsx", "PortalDashboard.tsx", "Dashboard", "CabecalhoGlobal, PaginaGlobal, BadgeGlobal", "useState, useEffect, useFetch", "GET /portal/dashboard", "DashboardPortal", "Atual", "Portal do fornecedor"),
    ("/portal/pendentes", "Cotações Pendentes", "Cotações Pendentes do Fornecedor", "CotacoesPendentes.tsx", "PortalPendentes.tsx", "Lista", "CabecalhoGlobal, PaginaGlobal, TabelaGlobal, BadgeGlobal, BotaoGlobal", "useState, useEffect, useFetch", "GET /portal/pendentes", "PedidoCotacao, Cotacao", "Atual", "Portal do fornecedor"),
    ("/portal/responder/:id", "Responder Cotação", "Responder Cotação (Portal)", "ResponderCotacao.tsx", "PortalResponder.tsx", "Formulário", "CabecalhoGlobal, PaginaGlobal, InputTextoGlobal, SelectGlobal, BotaoGlobal", "useState, useForm, useParams", "POST /portal/responder/:bidRequestId", "Proposta, DetalheTaxa", "Atual", "Formulário de resposta do fornecedor"),
    ("/portal/public/:token", "Responder (Público)", "Responder Cotação Pública", "ResponderPublico.tsx", "ResponderPublico.tsx", "Formulário Público", "PaginaGlobal, InputTextoGlobal, SelectGlobal, BotaoGlobal", "useState, useForm, useParams", "GET /portal/public/:token, POST /portal/public/:token/responder", "CotacaoPublica, Proposta", "Atual", "Sem autenticação, via token"),
    ("/portal/respostas", "Respostas", "Histórico de Respostas", "Respostas.tsx", "PortalRespostas.tsx", "Lista", "CabecalhoGlobal, PaginaGlobal, TabelaGlobal, BadgeGlobal", "useState, useEffect, useFetch", "GET /portal/respostas", "Proposta", "Atual", "Portal do fornecedor"),
    ("/portal/desempenho", "Desempenho", "Desempenho do Fornecedor", "Desempenho.tsx", "PortalDesempenho.tsx", "Dashboard", "CabecalhoGlobal, PaginaGlobal, BadgeGlobal", "useState, useEffect, useFetch", "GET /portal/desempenho", "ClassificacaoFornecedor", "Atual", "Portal do fornecedor"),
    ("/portal/tabela-precos", "Tabela de Preços", "Tabelas de Preço do Fornecedor", "TabelaPrecos.tsx", "PortalTabelaPrecos.tsx", "Lista", "CabecalhoGlobal, PaginaGlobal, TabelaGlobal, BotaoGlobal, ModalGlobal", "useState, useEffect, useFetch", "GET/POST/PUT/DELETE tabelas", "TabelaPreco", "Atual", "Portal do fornecedor"),
]

# ── Modals ──────────────────────────────────────────────────────────────

MODALS = [
    # (nome_atual, nome_ddd, arq_atual, arq_ddd, tipo, pagina_pai, trigger, campos, entidade, acao, api_call, tamanho, status_ddd, obs)
    ("Modal Aprovar", "ModalAprovarProposta", "Comparativo.tsx (inline)", "Comparativo.tsx (inline)", "Confirmação", "Comparativo", "Botão 'Aprovar' na proposta", "proposta_id (hidden)", "Proposta", "Aprovar proposta vencedora", "POST /comparativo/:id/aprovar", "sm", "Atual", "Confirmação simples"),
    ("Modal Reprovar", "ModalReprovarCotacao", "Comparativo.tsx (inline)", "Comparativo.tsx (inline)", "Formulário", "Comparativo", "Botão 'Reprovar Todas'", "motivo_reprovacao (textarea)", "Cotacao", "Reprovar todas as propostas", "POST /comparativo/:id/reprovar", "md", "Atual", "Exige motivo"),
    ("Modal Novo Fornecedor", "ModalCriarFornecedor", "Fornecedores.tsx (inline)", "Fornecedores.tsx (inline)", "Formulário", "Fornecedores", "Botão '+ Novo Fornecedor'", "nome, tipo, email, telefone, pais, cidade", "Fornecedor", "Criar novo fornecedor", "POST /fornecedores", "lg", "Atual", "Formulário completo"),
    ("Modal Tabela Preço", "ModalTabelaPreco", "DetalheFornecedor.tsx / TabelaPrecos.tsx (inline)", "FornecedorDetalhe.tsx / PortalTabelaPrecos.tsx (inline)", "Formulário", "DetalheFornecedor / TabelaPrecos", "Botão '+ Tabela'", "origem_*, destino_*, modal, modalidade, moeda, valores, transit_time, validade_*", "TabelaPreco", "Criar/editar tabela de preço", "POST/PUT /fornecedores/:id/tabela", "lg", "Atual", "Usado em 2 páginas"),
    ("Modal Avaliação", "ModalAvaliarFornecedor", "DetalheFornecedor.tsx (inline)", "FornecedorDetalhe.tsx (inline)", "Formulário", "DetalheFornecedor", "Botão 'Avaliar'", "nota_frete, nota_atendimento, nota_resposta, nota_confiabilidade, comentario", "Avaliacao", "Criar avaliação", "POST /fornecedores/:id/avaliar", "md", "Atual", "Rating 1-5 por critério"),
    ("Modal Disparo", "ModalDispararBid", "DetalheCotacao.tsx (inline)", "CotacaoDetalhe.tsx (inline)", "Formulário", "DetalheCotacao", "Botão 'Disparar Cotação'", "fornecedores[] (checkboxes), canal por fornecedor", "PedidoCotacao", "Disparar bid para fornecedores", "POST /bids/disparar", "lg", "Atual", "Seleção múltipla de fornecedores"),
    ("Modal Importar Bloco", "ModalImportarCotacoes", "CotacoesImportar.tsx (inline)", "CotacoesImportar.tsx (inline)", "Upload", "CotacoesImportar", "Botão 'Importar CSV'", "arquivo CSV (drag & drop)", "Cotacao", "Importar cotações em massa", "POST /cotacoes/bloco", "lg", "Atual", "Upload CSV com preview"),
]

# ── Nucleo Global ──────────────────────────────────────────────────────

NUCLEO_COMPONENTS = [
    # (componente, categoria, caminho, usado_em, qtd_usos, props, tipo, substituivel, status_ddd, obs)
    ("LogoBidFrete", "Logo", "nucleo-global/Logo/", "App.tsx, CabecalhoGlobal", "2", "tamanho?", "Visual", "Não", "Atual", "Logo do produto Bid Frete"),
    ("CabecalhoGlobal", "Layout", "nucleo-global/Layout/cabecalho-global/", "Todas as páginas", "17", "titulo, acoes?, migalhas?", "Layout", "Não", "Atual", "Header de todas as páginas"),
    ("PaginaGlobal", "Layout", "nucleo-global/Layout/pagina-global/", "Todas as páginas", "17", "children, className?", "Layout", "Não", "Atual", "Container principal de página"),
    ("StepperPassoPassoGlobal", "Campos", "nucleo-global/Campos/stepper-passo-passo-global/", "NovaCotacao", "1", "passos, passoAtual, onPassoMudar", "Navegação", "Não", "Atual", "Wizard steps na criação de cotação"),
    ("TabelaGlobal", "Tabelas", "nucleo-global/Tabelas/tabela-global/", "Cotacoes, Fornecedores, Respostas, Comparativo, TabelaPrecos, CotacoesPendentes", "6", "colunas, dados, onOrdenar?, onSelecionar?", "Dados", "Não", "Atual", "Componente de tabela reutilizável"),
    ("BadgeGlobal", "Campos", "nucleo-global/Campos/badge-global/", "Cotacoes, Fornecedores, Comparativo, Portal*", "8", "texto, variante, cor?", "Visual", "Não", "Atual", "Badges de status"),
    ("BotaoGlobal", "Campos", "nucleo-global/Campos/botao-global/", "Todas as páginas com ações", "15+", "texto, onClick, variante?, tamanho?, icone?", "Ação", "Não", "Atual", "Botões de ação"),
    ("InputTextoGlobal", "Campos", "nucleo-global/Campos/input-texto-global/", "NovaCotacao, Fornecedores, Configuracoes, Portal*", "10+", "label, valor, onChange, erro?, placeholder?", "Formulário", "Não", "Atual", "Input de texto"),
    ("SelectGlobal", "Campos", "nucleo-global/Campos/select-global/", "NovaCotacao, Fornecedores, Configuracoes", "8+", "label, opcoes, valor, onChange, erro?", "Formulário", "Não", "Atual", "Dropdown/select"),
    ("ModalGlobal", "Layout", "nucleo-global/Layout/modal-global/", "Comparativo, Fornecedores, DetalheFornecedor, DetalheCotacao, CotacoesImportar", "5+", "aberto, onFechar, titulo, children, tamanho?", "Layout", "Não", "Atual", "Modal wrapper"),
    ("DrawerGlobal", "Layout", "nucleo-global/Layout/drawer-global/", "Configuracoes", "1", "aberto, onFechar, titulo, children, lado?", "Layout", "Não", "Atual", "Drawer lateral"),
    ("TabsGlobal", "Layout", "nucleo-global/Layout/tabs-global/", "Cotacoes, DetalheCotacao, DetalheFornecedor, Configuracoes", "4", "abas, abaAtiva, onMudarAba", "Navegação", "Não", "Atual", "Navegação por abas"),
]


# ════════════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ════════════════════════════════════════════════════════════════════════

wb = openpyxl.Workbook()

# ── Sheet 1: ddd_campos ────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "1.ddd_campos"
ws1.sheet_properties.tabColor = TAB_COLORS["1.ddd_campos"]

headers1 = [
    "Local", "Tabela", "Entidade",
    "Nome no banco/back/front - ATUAL", "Nome no banco/back/front - DDD",
    "EXPLICACAO", "Produto Gravity", "Local Tela",
    "Nome em tela - Atual", "Nome em tela - DDD",
    "Descricao", "Tipo Dado", "Natureza", "Tipo de Dado", "Formato",
    "Validacao", "Obrigatorio", "Editavel", "Valor Padrao", "Exemplo",
    "Componente", "Origem"
]
widths1 = [28, 35, 22, 35, 38, 55, 14, 28, 22, 28, 35, 18, 12, 28, 18, 20, 12, 10, 14, 22, 18, 12]
style_sheet(ws1, headers1, widths1)

row_count_1 = 0
for model_name, model_data in MODELS.items():
    entity = model_data["entity"]
    for f in model_data["fields"]:
        field_name, tipo_dado, optional, default, ddd_name, natureza, editavel, explicacao, local_tela, nome_tela_atual, nome_tela_ddd, descricao, formato, validacao, valor_padrao, exemplo, componente, origem = f
        obrigatorio = "Não" if optional else "Sim"
        add_row(ws1, [
            LOCAL, model_name, entity,
            field_name, ddd_name,
            explicacao, PRODUTO, local_tela,
            nome_tela_atual, nome_tela_ddd,
            descricao, tipo_dado, natureza, tipo_dado, formato,
            validacao, obrigatorio, editavel, valor_padrao, exemplo,
            componente, origem
        ])
        row_count_1 += 1
    # Extra ts fields
    for f in model_data.get("extra_ts_fields", []):
        field_name, tipo_dado, optional, default, ddd_name, natureza, editavel, explicacao, local_tela, nome_tela_atual, nome_tela_ddd, descricao, formato, validacao, valor_padrao, exemplo, componente, origem = f
        obrigatorio = "Não" if optional else "Sim"
        add_row(ws1, [
            LOCAL + " (types.ts only)", model_name, entity,
            field_name + " (types.ts)", ddd_name,
            explicacao, PRODUTO, local_tela,
            nome_tela_atual, nome_tela_ddd,
            descricao, tipo_dado, natureza, tipo_dado, formato,
            validacao, obrigatorio, editavel, valor_padrao, exemplo,
            componente, origem
        ])
        row_count_1 += 1

# ── Sheet 2: ddd_api ──────────────────────────────────────────────────
ws2 = wb.create_sheet("2. ddd_api")
ws2.sheet_properties.tabColor = TAB_COLORS["2. ddd_api"]

headers2 = [
    "Metodo", "Rota Atual", "ROTA DDD", "EXPLICACAO ROTA",
    "ENTIDADE RETORNADA", "ENTIDADE RETORNADA - DDD",
    "EXPLICACAO ENTIDADE RETORNADA", "PREFIXO MOUNT",
    "ROTA NO ARQUIVO", "RESPONSE SCHEMA", "CONSUMIDOR"
]
widths2 = [8, 45, 50, 35, 25, 30, 35, 18, 30, 30, 20]
style_sheet(ws2, headers2, widths2)

row_count_2 = 0
for r in ROUTES:
    method, prefix, full_route, ddd_route, router_var, file_name, mw, auth, zod, entity, action, qp, bp, resp, sc, pag, filt, status, obs = r
    explicacao_rota = action
    explicacao_ent = f"Retorna {entity}"
    consumidor = "Frontend Bid Frete"
    if "portal" in full_route.lower() and "public" not in full_route.lower():
        consumidor = "Portal do Fornecedor"
    elif "public" in full_route.lower():
        consumidor = "Link público (sem auth)"
    add_row(ws2, [
        method, full_route, ddd_route, explicacao_rota,
        entity, entity, explicacao_ent, prefix,
        f"{file_name} ({router_var})", resp or entity, consumidor
    ])
    row_count_2 += 1

# ── Sheet 3: tabelas-models ──────────────────────────────────────────
ws3 = wb.create_sheet("3. tabelas-models")
ws3.sheet_properties.tabColor = TAB_COLORS["3. tabelas-models"]

headers3 = [
    "Local", "Nome no Prisma", "Nome DDD - Prisma", "Explicacao", "Tipo",
    "Produto Gravity", "Natureza", "Descricao",
    "O que faz / Proposito", "Tem tenant_id?", "Chave primaria",
    "Qtd FKs", "Relacoes (lista)", "Qtd indices", "Indices (lista)",
    "Soft delete?", "Auditoria?", "Qtd de campos", "Volume estimado",
    "Origem do dado", "Consumidores principais", "Escritores principais",
    "Arquivo fragment", "Status DDD", "Observacoes"
]
widths3 = [28, 38, 38, 40, 10, 14, 12, 40, 50, 14, 18, 8, 40, 8, 40, 12, 12, 10, 18, 18, 30, 30, 35, 12, 30]
style_sheet(ws3, headers3, widths3)

# Model metadata
MODEL_META = {
    "FreteIntBidFornecedores": {
        "ddd": "FreteIntBidFornecedores", "tipo": "Model", "natureza": "Operacional",
        "desc": "Cadastro de fornecedores de frete", "proposito": "Armazena fornecedores (armadores, agentes, cias aéreas, transportadoras) que participam dos processos de cotação",
        "tenant": "Sim", "pk": "id (CUID)", "qtd_fk": 0, "relacoes": "PedidoCotacao (1:N), Proposta (1:N), Avaliacao (1:N), TabelaPreco (1:N), Integracao (1:N)",
        "qtd_idx": 3, "indices": "@@index([id_organizacao]), @@index([id_organizacao, product_id]), @@index([id_organizacao, email])",
        "soft_del": "Não (status INATIVO)", "audit": "created_at, updated_at", "qtd_campos": 20,
        "volume": "100-10k por org", "origem": "Usuario (cadastro manual ou importação)", "consumidores": "Frontend (lista, detalhe), Portal, Comparativo",
        "escritores": "Frontend CRUD, Portal auto-cadastro", "fragment": "servicos-global/produto/bid-frete/prisma/fragment.prisma",
    },
    "FreteIntBidCotacoes": {
        "ddd": "FreteIntBidCotacoes", "tipo": "Model", "natureza": "Operacional",
        "desc": "Processos de cotação de frete", "proposito": "Representa uma solicitação de cotação (RFQ) enviada a fornecedores para obter propostas de frete",
        "tenant": "Sim", "pk": "id (CUID)", "qtd_fk": 1, "relacoes": "PedidoCotacao (1:N), Proposta (1:N), GanhoEstimado (1:1), Fornecedor vencedor (N:1)",
        "qtd_idx": 3, "indices": "@@index([id_organizacao]), @@index([id_organizacao, product_id]), @@index([id_organizacao, status])",
        "soft_del": "Não (status CANCELADA)", "audit": "created_at, updated_at", "qtd_campos": 40,
        "volume": "1k-100k por org/ano", "origem": "Usuario (formulário ou import CSV)", "consumidores": "Frontend (lista, detalhe, comparativo), Dashboard",
        "escritores": "Frontend CRUD, Import CSV", "fragment": "servicos-global/produto/bid-frete/prisma/fragment.prisma",
    },
    "FreteIntBidPedidoCotacoes": {
        "ddd": "FreteIntBidPedidoCotacoes", "tipo": "Model", "natureza": "Operacional",
        "desc": "Pedidos de cotação enviados a cada fornecedor", "proposito": "Registra o envio de uma cotação para um fornecedor específico, com status de tracking (enviado, visualizado, respondido)",
        "tenant": "Sim", "pk": "id (CUID)", "qtd_fk": 2, "relacoes": "Cotacao (N:1), Fornecedor (N:1), Proposta (1:1)",
        "qtd_idx": 3, "indices": "@@index([id_organizacao]), @@index([id_organizacao, cotacao_id]), @@index([token_resposta])",
        "soft_del": "Não", "audit": "created_at, updated_at", "qtd_campos": 17,
        "volume": "5x-20x volume de cotações", "origem": "Sistema (disparo automático)", "consumidores": "Frontend (detalhe cotação), Portal fornecedor",
        "escritores": "Backend (disparo bid), Portal (atualiza status)", "fragment": "servicos-global/produto/bid-frete/prisma/fragment.prisma",
    },
    "FreteIntBidPropostas": {
        "ddd": "FreteIntBidPropostas", "tipo": "Model", "natureza": "Operacional",
        "desc": "Propostas/respostas dos fornecedores", "proposito": "Armazena a proposta de preço enviada por um fornecedor em resposta a um pedido de cotação",
        "tenant": "Sim", "pk": "id (CUID)", "qtd_fk": 3, "relacoes": "PedidoCotacao (1:1), Cotacao (N:1), Fornecedor (N:1), DetalheTaxa (1:N)",
        "qtd_idx": 3, "indices": "@@index([id_organizacao]), @@index([id_organizacao, cotacao_id]), @@index([bid_request_id])",
        "soft_del": "Não", "audit": "created_at, updated_at", "qtd_campos": 28,
        "volume": "1x-5x volume de pedidos cotação", "origem": "Fornecedor (portal, email, API)", "consumidores": "Frontend (comparativo, detalhe), Dashboard",
        "escritores": "Portal fornecedor, Link público, API integração", "fragment": "servicos-global/produto/bid-frete/prisma/fragment.prisma",
    },
    "FreteIntBidPropostasTaxasCambio": {
        "ddd": "FreteIntBidPropostasTaxasCambio", "tipo": "Model", "natureza": "Operacional",
        "desc": "Detalhamento de taxas das propostas", "proposito": "Detalha cada taxa individual (THC, BL fee, customs, etc.) de uma proposta",
        "tenant": "Sim", "pk": "id (CUID)", "qtd_fk": 1, "relacoes": "Proposta (N:1)",
        "qtd_idx": 2, "indices": "@@index([id_organizacao]), @@index([response_id])",
        "soft_del": "Não", "audit": "Não", "qtd_campos": 7,
        "volume": "5x-15x volume de propostas", "origem": "Fornecedor (junto com proposta)", "consumidores": "Frontend (comparativo detalhado)",
        "escritores": "Portal fornecedor, Link público", "fragment": "servicos-global/produto/bid-frete/prisma/fragment.prisma",
    },
    "FreteIntBidTabelasProntas": {
        "ddd": "FreteIntBidTabelasProntas", "tipo": "Model", "natureza": "Referência",
        "desc": "Tabelas de preço pré-definidas dos fornecedores", "proposito": "Armazena tabelas de preço padrão dos fornecedores para cotação automática ou referência",
        "tenant": "Sim", "pk": "id (CUID)", "qtd_fk": 1, "relacoes": "Fornecedor (N:1)",
        "qtd_idx": 3, "indices": "@@index([id_organizacao]), @@index([id_organizacao, fornecedor_id]), @@index([id_organizacao, ativa])",
        "soft_del": "Não (flag ativa)", "audit": "created_at, updated_at", "qtd_campos": 23,
        "volume": "10x-50x volume de fornecedores", "origem": "Fornecedor (portal) ou Usuario (cadastro)", "consumidores": "Frontend (detalhe fornecedor), Portal, Cotação automática",
        "escritores": "Frontend CRUD, Portal fornecedor", "fragment": "servicos-global/produto/bid-frete/prisma/fragment.prisma",
    },
    "FreteIntBidFornecedoresAvaliacoes": {
        "ddd": "FreteIntBidFornecedoresAvaliacoes", "tipo": "Model", "natureza": "Operacional",
        "desc": "Avaliações dos fornecedores", "proposito": "Registra avaliações (notas 1-5 por critério + comentário) feitas pelos usuários sobre fornecedores",
        "tenant": "Sim", "pk": "id (CUID)", "qtd_fk": 2, "relacoes": "Fornecedor (N:1), Cotacao (N:1 opcional)",
        "qtd_idx": 3, "indices": "@@index([id_organizacao]), @@index([id_organizacao, fornecedor_id]), @@index([id_organizacao, user_id])",
        "soft_del": "Não", "audit": "created_at, updated_at", "qtd_campos": 14,
        "volume": "1x volume de cotações aprovadas", "origem": "Usuario (avaliação pós-cotação)", "consumidores": "Frontend (detalhe fornecedor), Ranking, ClassificacaoFornecedor",
        "escritores": "Frontend (modal avaliação)", "fragment": "servicos-global/produto/bid-frete/prisma/fragment.prisma",
    },
    "FreteIntBidClassificacaoFornecedores": {
        "ddd": "FreteIntBidClassificacaoFornecedores", "tipo": "Model", "natureza": "Agregação",
        "desc": "Classificação agregada dos fornecedores (cross-org)", "proposito": "Tabela materializada com métricas agregadas de desempenho dos fornecedores, indexada por email (cross-organização)",
        "tenant": "Não (cross-org por email)", "pk": "id (CUID)", "qtd_fk": 0, "relacoes": "Nenhuma (lookup por email)",
        "qtd_idx": 1, "indices": "@@unique([fornecedor_email])",
        "soft_del": "Não", "audit": "updated_at", "qtd_campos": 16,
        "volume": "1x volume de fornecedores únicos", "origem": "Sistema (recalculado automaticamente)", "consumidores": "Frontend (ranking, detalhe fornecedor), Portal (desempenho)",
        "escritores": "Backend (job de recálculo após avaliação/cotação)", "fragment": "servicos-global/produto/bid-frete/prisma/fragment.prisma",
    },
    "FreteIntBidGanhoEstimado": {
        "ddd": "FreteIntBidGanhoEstimado", "tipo": "Model", "natureza": "Analítico",
        "desc": "Ganho/saving estimado por cotação", "proposito": "Registra o saving obtido ao aprovar uma proposta vs valor target e média das propostas",
        "tenant": "Sim", "pk": "id (CUID)", "qtd_fk": 1, "relacoes": "Cotacao (1:1)",
        "qtd_idx": 3, "indices": "@@index([id_organizacao]), @@index([id_organizacao, cotacao_id]), @@index([id_organizacao, company_id])",
        "soft_del": "Não", "audit": "created_at", "qtd_campos": 14,
        "volume": "1x volume de cotações aprovadas", "origem": "Sistema (calculado na aprovação)", "consumidores": "Frontend (dashboard, detalhe cotação)",
        "escritores": "Backend (ao aprovar proposta)", "fragment": "servicos-global/produto/bid-frete/prisma/fragment.prisma",
    },
    "FreteIntBidIntegracoes": {
        "ddd": "FreteIntBidIntegracoes", "tipo": "Model", "natureza": "Configuração",
        "desc": "Integrações com APIs de fornecedores", "proposito": "Configura integrações com APIs externas de fornecedores (REST, SOAP, OData) para cotação automática",
        "tenant": "Sim", "pk": "id (CUID)", "qtd_fk": 1, "relacoes": "Fornecedor (N:1 opcional)",
        "qtd_idx": 3, "indices": "@@index([id_organizacao]), @@index([id_organizacao, fornecedor_id]), @@index([id_organizacao, ativo])",
        "soft_del": "Não (flag ativo)", "audit": "created_at, updated_at", "qtd_campos": 18,
        "volume": "1-50 por org", "origem": "Usuario (configuração manual)", "consumidores": "Backend (engine de cotação automática)",
        "escritores": "Frontend (configurações)", "fragment": "servicos-global/produto/bid-frete/prisma/fragment.prisma",
    },
    "FreteIntBidPortosCadastro": {
        "ddd": "FreteIntBidPortosCadastro", "tipo": "Model", "natureza": "Referência (Master Data)",
        "desc": "Cadastro de portos e aeroportos", "proposito": "Tabela de referência global com portos, aeroportos e terminais para seleção em cotações",
        "tenant": "Não (global)", "pk": "codigo (String natural)", "qtd_fk": 0, "relacoes": "Nenhuma (referência por código)",
        "qtd_idx": 2, "indices": "@@index([pais_codigo]), @@index([tipo, ativo])",
        "soft_del": "Não (flag ativo)", "audit": "Não", "qtd_campos": 8,
        "volume": "5k-10k registros fixos", "origem": "Sistema (seed/importação UNLOCODE)", "consumidores": "Frontend (selects de origem/destino), Master Data API",
        "escritores": "Admin (seed/importação)", "fragment": "servicos-global/produto/bid-frete/prisma/fragment.prisma",
    },
}

row_count_3 = 0
for model_name, meta in MODEL_META.items():
    add_row(ws3, [
        LOCAL, model_name, meta["ddd"], "", meta["tipo"],
        PRODUTO, meta["natureza"], meta["desc"],
        meta["proposito"], meta["tenant"], meta["pk"],
        meta["qtd_fk"], meta["relacoes"], meta["qtd_idx"], meta["indices"],
        meta["soft_del"], meta["audit"], meta["qtd_campos"], meta["volume"],
        meta["origem"], meta["consumidores"], meta["escritores"],
        meta["fragment"], "Atual", ""
    ])
    row_count_3 += 1

# Add enums as rows too
for enum_name, enum_data in ENUMS.items():
    add_row(ws3, [
        LOCAL, enum_name, enum_name, "", "Enum",
        PRODUTO, "Definição", f"Enum com {len(enum_data['values'])} valores",
        f"Define valores possíveis para {enum_data['used_in_field']} em {enum_data['used_in_model']}",
        "N/A", "N/A",
        0, "N/A", 0, "N/A",
        "N/A", "N/A", len(enum_data["values"]), "N/A",
        "Definição fixa", enum_data["used_in_model"], "Schema Prisma",
        "servicos-global/produto/bid-frete/prisma/fragment.prisma", "Atual", f"Categoria: {enum_data['categoria']}"
    ])
    row_count_3 += 1

# ── Sheet 4: mapa-enums ──────────────────────────────────────────────
ws4 = wb.create_sheet("4. mapa-enums")
ws4.sheet_properties.tabColor = TAB_COLORS["4. mapa-enums"]

headers4 = [
    "Enum Prisma", "Enum DDD", "Valor", "Valor DDD", "Label PT-BR",
    "Label EN", "Descricao", "Produto Gravity", "Usado em Model",
    "Usado em Campo", "Tipo", "Categoria", "Ordem", "Ativo",
    "Cor Badge", "Icone", "Agrupamento", "Status DDD", "Observacoes"
]
widths4 = [30, 30, 25, 25, 30, 25, 40, 14, 35, 20, 12, 22, 8, 8, 12, 12, 18, 12, 25]
style_sheet(ws4, headers4, widths4)

# Badge colors
BADGE_COLORS = {
    "RASCUNHO": "gray", "ENVIADA_FORNECEDORES": "blue", "EM_COTACAO": "cyan", "AGUARDANDO_APROVACAO": "yellow",
    "APROVADA": "green", "REPROVADA": "red", "CANCELADA": "gray", "FALTA_INFORMACAO": "orange", "EXPIRADA": "gray",
    "PENDENTE": "yellow", "ENVIADO": "blue", "VISUALIZADO": "cyan", "RESPONDIDO": "green", "EXPIRADO": "gray", "ERRO_ENVIO": "red",
    "RECEBIDA": "blue", "EM_ANALISE": "cyan", "MELHOR_PRECO": "green", "MELHOR_TRANSIT": "green", "MELHOR_AVALIACAO": "green",
    "ATIVO": "green", "INATIVO": "gray", "PENDENTE_APROVACAO": "yellow", "BLOQUEADO": "red",
    "IMPORTACAO": "blue", "EXPORTACAO": "orange",
    "MARITIMO": "blue", "AEREO": "cyan", "RODOVIARIO": "orange",
    "FCL": "blue", "LCL": "cyan", "AEREO_GERAL": "cyan", "RODOVIARIO_FTL": "orange", "RODOVIARIO_LTL": "orange",
    "EMAIL": "blue", "WHATSAPP": "green", "API": "purple", "PORTAL": "cyan",
    "DIRECIONADA": "blue", "ABERTA": "green",
    "AGENTE_CARGA": "blue", "ARMADOR": "cyan", "CIA_AEREA": "purple", "TRANSPORTADORA": "orange",
    "API_REST": "blue", "API_SOAP": "orange", "ODATA": "purple", "MANUAL": "gray",
}

BADGE_ICONS = {
    "IMPORTACAO": "ArrowDown", "EXPORTACAO": "ArrowUp",
    "MARITIMO": "Anchor", "AEREO": "AirplaneTilt", "RODOVIARIO": "Truck",
    "EMAIL": "Envelope", "WHATSAPP": "WhatsappLogo", "API": "Code", "PORTAL": "Globe",
    "ATIVO": "CheckCircle", "INATIVO": "XCircle", "PENDENTE_APROVACAO": "Clock", "BLOQUEADO": "Lock",
    "APROVADA": "CheckCircle", "REPROVADA": "XCircle", "CANCELADA": "Prohibit",
}

row_count_4 = 0
for enum_name, enum_data in ENUMS.items():
    for idx, (valor, label_pt, label_en, desc) in enumerate(enum_data["values"], 1):
        add_row(ws4, [
            enum_name, enum_name, valor, valor, label_pt,
            label_en, desc, PRODUTO, enum_data["used_in_model"],
            enum_data["used_in_field"], "Enum", enum_data["categoria"], idx, "Sim",
            BADGE_COLORS.get(valor, ""), BADGE_ICONS.get(valor, ""), enum_data["categoria"], "Atual", ""
        ])
        row_count_4 += 1

# ── Sheet 5: mapa-rotas ──────────────────────────────────────────────
ws5 = wb.create_sheet("5. mapa-rotas")
ws5.sheet_properties.tabColor = TAB_COLORS["5. mapa-rotas"]

headers5 = [
    "Metodo", "Prefixo", "Rota Completa", "Rota DDD",
    "Router Variable", "Arquivo", "Middleware", "Auth",
    "Zod Schema", "Entidade Principal", "Acao",
    "Query Params", "Body Params", "Response Type",
    "Status Code", "Paginacao", "Filtros", "Status DDD", "Observacoes"
]
widths5 = [8, 14, 50, 55, 22, 22, 35, 18, 25, 22, 22, 30, 50, 35, 10, 10, 30, 12, 35]
style_sheet(ws5, headers5, widths5)

row_count_5 = 0
for r in ROUTES:
    method, prefix, full_route, ddd_route, router_var, file_name, mw, auth, zod, entity, action, qp, bp, resp, sc, pag, filt, status, obs = r
    add_row(ws5, [
        method, prefix, full_route, ddd_route,
        router_var, file_name, mw, auth,
        zod, entity, action,
        qp, bp, resp,
        sc, pag, filt, status, obs
    ])
    row_count_5 += 1

# ── Sheet 6: mapa-paginas ────────────────────────────────────────────
ws6 = wb.create_sheet("6. mapa-paginas")
ws6.sheet_properties.tabColor = TAB_COLORS["6. mapa-paginas"]

headers6 = [
    "URL", "Titulo Atual", "Titulo DDD", "Arquivo Atual", "Arquivo DDD",
    "Tipo", "Produto Gravity", "Componentes Usados", "Hooks Usados",
    "API Calls", "Entidades", "Status DDD", "Observacoes"
]
widths6 = [30, 25, 30, 25, 25, 22, 14, 60, 35, 55, 40, 12, 40]
style_sheet(ws6, headers6, widths6)

row_count_6 = 0
for p in PAGES:
    url, titulo_atual, titulo_ddd, arq_atual, arq_ddd, tipo, comps, hooks, api, ents, status, obs = p
    add_row(ws6, [
        url, titulo_atual, titulo_ddd, arq_atual, arq_ddd,
        tipo, PRODUTO, comps, hooks,
        api, ents, status, obs
    ])
    row_count_6 += 1

# ── Sheet 7: Modais ──────────────────────────────────────────────────
ws7 = wb.create_sheet("7. Modais")
ws7.sheet_properties.tabColor = TAB_COLORS["7. Modais"]

headers7 = [
    "Nome Atual", "Nome DDD", "Arquivo Atual", "Arquivo DDD",
    "Tipo", "Produto Gravity", "Pagina Pai", "Trigger",
    "Campos", "Entidade", "Acao Principal", "API Call",
    "Tamanho", "Status DDD", "Observacoes"
]
widths7 = [22, 25, 40, 40, 14, 14, 25, 30, 55, 18, 30, 40, 10, 12, 30]
style_sheet(ws7, headers7, widths7)

row_count_7 = 0
for m in MODALS:
    nome, nome_ddd, arq, arq_ddd, tipo, pag, trigger, campos, ent, acao, api, tam, status, obs = m
    add_row(ws7, [
        nome, nome_ddd, arq, arq_ddd,
        tipo, PRODUTO, pag, trigger,
        campos, ent, acao, api,
        tam, status, obs
    ])
    row_count_7 += 1

# ── Sheet 8: nucleo-global ───────────────────────────────────────────
ws8 = wb.create_sheet("8. nucleo-global")
ws8.sheet_properties.tabColor = TAB_COLORS["8. nucleo-global"]

headers8 = [
    "Componente", "Categoria", "Caminho", "Usado em", "Qtd Usos",
    "Props Principais", "Produto Gravity", "Tipo", "Substituivel",
    "Status DDD", "Observacoes"
]
widths8 = [22, 12, 45, 55, 10, 45, 14, 14, 12, 12, 30]
style_sheet(ws8, headers8, widths8)

row_count_8 = 0
for c in NUCLEO_COMPONENTS:
    comp, cat, caminho, usado, qtd, props, tipo, subst, status, obs = c
    add_row(ws8, [
        comp, cat, caminho, usado, qtd,
        props, PRODUTO, tipo, subst,
        status, obs
    ])
    row_count_8 += 1

# ── Sheet 9: componentes-locais ──────────────────────────────────────
ws9 = wb.create_sheet("9. componentes-locais")
ws9.sheet_properties.tabColor = TAB_COLORS["9. componentes-locais"]

headers9 = [
    "Nome Atual", "Nome DDD", "Arquivo Atual", "Arquivo DDD",
    "Tipo", "Produto Gravity", "Pagina Pai", "Props",
    "Entidade", "Reutilizavel", "Dependencias",
    "Nucleo Global Usado", "Status DDD", "Observacoes"
]
widths9 = [22, 22, 25, 25, 14, 14, 20, 25, 14, 12, 22, 30, 12, 45]
style_sheet(ws9, headers9, widths9)

add_row(ws9, [
    "N/A", "N/A", "N/A", "N/A",
    "N/A", PRODUTO, "N/A", "N/A",
    "N/A", "N/A", "N/A",
    "N/A", "N/A", "Todos os componentes estao inline nas paginas. Nao existem arquivos de componentes locais separados."
])
row_count_9 = 1

# ── Save ─────────────────────────────────────────────────────────────
wb.save(OUTPUT)

print(f"\nPlanilha salva em: {OUTPUT}")
print(f"\n{'='*50}")
print(f"CONTAGEM DE LINHAS POR SHEET:")
print(f"{'='*50}")
print(f"  1. ddd_campos:          {row_count_1} linhas")
print(f"  2. ddd_api:             {row_count_2} linhas")
print(f"  3. tabelas-models:      {row_count_3} linhas ({row_count_3 - len(ENUMS)} models + {len(ENUMS)} enums)")
print(f"  4. mapa-enums:          {row_count_4} linhas (valores de {len(ENUMS)} enums)")
print(f"  5. mapa-rotas:          {row_count_5} linhas")
print(f"  6. mapa-paginas:        {row_count_6} linhas")
print(f"  7. Modais:              {row_count_7} linhas")
print(f"  8. nucleo-global:       {row_count_8} linhas")
print(f"  9. componentes-locais:  {row_count_9} linhas")
print(f"{'='*50}")
print(f"  TOTAL:                  {row_count_1 + row_count_2 + row_count_3 + row_count_4 + row_count_5 + row_count_6 + row_count_7 + row_count_8 + row_count_9} linhas")
