"""
Limpa as colunas DDD em '1. ddd_campos': esvazia onde o nome DDD é igual ao atual.

Regra: coluna DDD só deve ter valor quando o nome mudar.
- Col D (PG atual) vs F (PG DDD)    → esvazia F se iguais
- Col G (back atual) vs H (back DDD) → esvazia H se iguais
- Col I (front atual) vs J (front DDD) → esvazia J se iguais
- Col K (tela atual) vs L (tela DDD)   → esvazia L se iguais
"""
import openpyxl
from openpyxl.styles import PatternFill

SRC = r'C:\Users\danie\Downloads\planilha_geral_gravity (15).xlsx'
DST = r'C:\Users\danie\Downloads\planilha_geral_gravity_CORRIGIDA.xlsx'

wb = openpyxl.load_workbook(SRC)
ws = wb['1. ddd_campos']

pairs = [(4,6),(7,8),(9,10),(11,12)]  # (atual, ddd)
cleared = {k:0 for k in ['F','H','J','L']}

for r in range(2, ws.max_row+1):
    for atual_col, ddd_col in pairs:
        a = ws.cell(r, atual_col).value
        d = ws.cell(r, ddd_col).value
        if a is None or d is None: continue
        if str(a).strip() == str(d).strip():
            ws.cell(r, ddd_col).value = None
            letter = {6:'F',8:'H',10:'J',12:'L'}[ddd_col]
            cleared[letter] += 1

wb.save(DST)
total = sum(cleared.values())
print(f'Células esvaziadas: {total}')
for k,v in cleared.items():
    print(f'  Col {k}: {v}')
print(f'\nArquivo salvo: {DST}')
