#!/usr/bin/env python3
"""
gerar-atlas-ddd.py

Gera o Atlas DDD canonico em `documentos-tecnicos/ddd-atlas/` a partir da
planilha mestre `planilha_geral_gravity.xlsx`.

Saida: 9 arquivos `.md` (um por aba), cada um com:
  - Header: aba, versao da planilha, data de geracao, este script
  - Como ler: legenda das colunas
  - Tabela: linhas DDD-finais (ordenadas por Local/Produto, depois Nome DDD)
  - Apendice: linhas marcadas com `--` na planilha (nao acionaveis)

Uso:
    python scripts/sob-demanda/gerar-atlas-ddd.py <caminho-da-planilha.xlsx>

Exemplo:
    python scripts/sob-demanda/gerar-atlas-ddd.py .claude/planilha-tmp.xlsx
"""
from __future__ import annotations

import os
import re
import sys
from datetime import date
from pathlib import Path
from typing import Callable, Iterable

import openpyxl

# ---------------------------------------------------------------------------
# Configuracao geral
# ---------------------------------------------------------------------------

PLANILHA_VERSAO = "52"  # ajustar quando a planilha for atualizada (vide README)

EM_DASH = "—"
HYPHENS_EXEMPT = {EM_DASH, "-", "--", ""}


def is_exempt(value) -> bool:
    """Linha exempt: valor None/vazio/em-dash."""
    if value is None:
        return True
    s = str(value).strip()
    return s in HYPHENS_EXEMPT or s == EM_DASH


def cell(ws, row: int, col: int):
    v = ws.cell(row=row, column=col).value
    if v is None:
        return ""
    return v


def md_escape(value) -> str:
    """Escapa pipe e quebras de linha para celula de tabela Markdown."""
    if value is None:
        return ""
    s = str(value).replace("\r\n", " ").replace("\n", " ").replace("\r", " ")
    s = s.replace("|", "\\|")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def render_table(headers: list[str], rows: list[list]) -> str:
    """Renderiza tabela Markdown com padding por coluna (alinhamento visual no raw)."""
    if not rows:
        return "_(sem linhas acionaveis nesta aba)_\n"

    # Escape e normalize todas as celulas
    header_cells = [md_escape(h) for h in headers]
    body_cells = [[md_escape(c) for c in r] for r in rows]

    # Calcula largura maxima por coluna (cabecalho + corpo)
    n_cols = len(header_cells)
    widths = [len(header_cells[i]) for i in range(n_cols)]
    for r in body_cells:
        for i in range(n_cols):
            if i < len(r):
                widths[i] = max(widths[i], len(r[i]))
    # Largura minima 3 (para o "---" do separador)
    widths = [max(w, 3) for w in widths]

    def pad_row(cells: list[str]) -> str:
        padded = [
            (cells[i] if i < len(cells) else "").ljust(widths[i])
            for i in range(n_cols)
        ]
        return "| " + " | ".join(padded) + " |"

    out = [pad_row(header_cells)]
    out.append("| " + " | ".join("-" * w for w in widths) + " |")
    for r in body_cells:
        out.append(pad_row(r))
    return "\n".join(out) + "\n"


def write_atlas(
    output_dir: Path,
    filename: str,
    aba_index: int,
    aba_titulo: str,
    como_ler: str,
    table_headers: list[str],
    rows_acionaveis: list[list],
    rows_apendice: list[list],
) -> Path:
    today = date.today().isoformat()
    parts: list[str] = []
    parts.append(f"# Atlas DDD - Aba {aba_index}: {aba_titulo}")
    parts.append("")
    parts.append(
        f"> Gerado de planilha v{PLANILHA_VERSAO} em {today} por "
        "`scripts/sob-demanda/gerar-atlas-ddd.py`."
    )
    parts.append(
        "> NAO edite manualmente. Re-execute o script apos mudanca na planilha mestre."
    )
    parts.append("")
    parts.append("## Como ler")
    parts.append("")
    parts.append(como_ler.rstrip())
    parts.append("")
    parts.append("Convencoes:")
    parts.append("- Apenas valores DDD-finais. Nao mostra estado pre-rename.")
    parts.append(
        "- Onde aplicavel, coluna \"Alias historico\" mostra nome legado "
        "(util para grep e git log --follow)."
    )
    parts.append(
        f"- Linhas marcadas `{EM_DASH}` na planilha estao no apendice "
        "(nao sao acionaveis)."
    )
    parts.append("")
    parts.append(f"## Tabela ({len(rows_acionaveis)} linhas)")
    parts.append("")
    parts.append(render_table(table_headers, rows_acionaveis))
    parts.append("")
    parts.append(f"## Apendice - Linhas SKIP / exempt ({len(rows_apendice)} linhas)")
    parts.append("")
    if rows_apendice:
        parts.append(
            "Linhas onde o nome DDD principal foi marcado como `"
            f"{EM_DASH}` (nao acionavel: arquivo de teste, definicao Storybook, "
            "ruido de parsing, etc.)."
        )
        parts.append("")
        parts.append(render_table(table_headers, rows_apendice))
    else:
        parts.append("_(nenhuma linha exempt nesta aba)_")

    parts.append("")
    out_path = output_dir / filename
    out_path.write_text("\n".join(parts), encoding="utf-8")
    return out_path


# ---------------------------------------------------------------------------
# Geradores por aba
# ---------------------------------------------------------------------------


def gerar_aba_01_campos(wb, output_dir: Path) -> Path:
    """Aba 1: 1.ddd_campos."""
    ws = wb["1.ddd_campos"]
    headers = [
        "Local",
        "Tabela",
        "Nome DDD (db/back/front)",
        "Nome em tela DDD",
        "Tipo Dado",
        "Validacao",
        "Obrigatorio",
        "Editavel",
        "Descricao",
        "Alias historico",
    ]
    como_ler = (
        "- **Local**: produto/servico onde o campo vive (configurador, pedido, dashboard, ...).\n"
        "- **Tabela**: model Prisma (ou agrupamento logico) ao qual o campo pertence.\n"
        "- **Nome DDD (db/back/front)**: nome unico, snake_case, valido em banco/back/front "
        "(REGRA 5 da skill `ddd-nomenclatura`).\n"
        "- **Nome em tela DDD**: label canonical PT-BR exibido na UI.\n"
        "- **Tipo Dado / Validacao**: tipo logico e regra Zod/Prisma esperada.\n"
        "- **Obrigatorio / Editavel**: contrato de UI e API.\n"
        "- **Descricao**: explicacao curta do significado do campo.\n"
        "- **Alias historico**: nome anterior (antes do rename DDD), util para grep."
    )
    acionaveis: list[list] = []
    apendice: list[list] = []
    for r in range(2, ws.max_row + 1):
        local = cell(ws, r, 1)
        tabela = cell(ws, r, 2)
        nome_atual = cell(ws, r, 4)
        nome_ddd = cell(ws, r, 5)
        nome_tela_ddd = cell(ws, r, 10)
        descricao = cell(ws, r, 11)
        tipo_dado = cell(ws, r, 12)
        validacao = cell(ws, r, 16)
        obrigatorio = cell(ws, r, 17)
        editavel = cell(ws, r, 18)
        if not local and not tabela and not nome_ddd:
            continue
        row = [
            local,
            tabela,
            nome_ddd,
            nome_tela_ddd,
            tipo_dado,
            validacao,
            obrigatorio,
            editavel,
            descricao,
            nome_atual,
        ]
        if is_exempt(nome_ddd):
            apendice.append(row)
        else:
            acionaveis.append(row)
    acionaveis.sort(key=lambda x: (str(x[0]).lower(), str(x[1]).lower(), str(x[2]).lower()))
    apendice.sort(key=lambda x: (str(x[0]).lower(), str(x[1]).lower()))
    return write_atlas(
        output_dir,
        "01-campos.md",
        1,
        "Campos (db/back/front)",
        como_ler,
        headers,
        acionaveis,
        apendice,
    )


def gerar_aba_02_rotas_api(wb, output_dir: Path) -> Path:
    """Aba 2: 2. ddd_api - rotas backend."""
    ws = wb["2. ddd_api"]
    headers = [
        "Metodo",
        "Rota DDD",
        "Arquivo",
        "Mount",
        "Path interno",
        "Response Schema",
        "Consumidor",
        "Descricao",
        "Alias historico",
    ]
    como_ler = (
        "- **Metodo**: verbo HTTP (GET, POST, PUT, PATCH, DELETE).\n"
        "- **Rota DDD**: caminho final canonico (`/api/v1/...`) - este e o contrato.\n"
        "- **Arquivo**: caminho do arquivo de rota Express.\n"
        "- **Mount**: prefixo onde o router e montado (`app.use(mount, router)`).\n"
        "- **Path interno**: path declarado dentro do router (`router.get(path, ...)`).\n"
        "- **Response Schema**: schema Zod que valida o payload de resposta.\n"
        "- **Consumidor**: produto/servico que invoca esta rota.\n"
        "- **Descricao**: o que a rota faz, em uma frase.\n"
        "- **Alias historico**: rota antes do rename DDD."
    )
    acionaveis: list[list] = []
    apendice: list[list] = []
    for r in range(2, ws.max_row + 1):
        metodo = cell(ws, r, 1)
        rota_atual = cell(ws, r, 2)
        rota_ddd = cell(ws, r, 3)
        descricao = cell(ws, r, 4)
        arquivo = cell(ws, r, 5)
        # F: arquivo DDD (mesmo arquivo na maioria); H: mount; I: path; J: consumidor
        mount = cell(ws, r, 8)
        path_interno = cell(ws, r, 9)
        response_schema = cell(ws, r, 10)
        consumidor = cell(ws, r, 11) or cell(ws, r, 10)
        if not metodo and not rota_ddd:
            continue
        row = [
            metodo,
            rota_ddd,
            arquivo,
            mount,
            path_interno,
            response_schema,
            consumidor,
            descricao,
            rota_atual,
        ]
        if is_exempt(rota_ddd):
            apendice.append(row)
        else:
            acionaveis.append(row)
    acionaveis.sort(key=lambda x: (str(x[1]).lower(), str(x[0]).lower()))
    return write_atlas(
        output_dir,
        "02-rotas-api.md",
        2,
        "Rotas API (backend)",
        como_ler,
        headers,
        acionaveis,
        apendice,
    )


def gerar_aba_03_models(wb, output_dir: Path) -> Path:
    """Aba 3: 3. tabelas-models - models Prisma."""
    ws = wb["3. tabelas-models"]
    headers = [
        "Local",
        "Nome DDD",
        "Tipo",
        "Tem tenant_id?",
        "PK",
        "Relacoes",
        "Indices",
        "Soft delete",
        "Auditoria",
        "Arquivo fragment",
        "Descricao",
        "Alias historico",
    ]
    como_ler = (
        "- **Local**: produto/servico que possui o model.\n"
        "- **Nome DDD**: nome PascalCase do model Prisma (com `@@map(\"snake_case\")`).\n"
        "- **Tipo**: agregado, entidade, value-object, etc.\n"
        "- **Tem tenant_id?**: se possui `id_organizacao` (REGRA universal de isolamento).\n"
        "- **PK**: chave primaria.\n"
        "- **Relacoes / Indices**: relacionamentos e indices declarados.\n"
        "- **Soft delete / Auditoria**: marcadores de governanca.\n"
        "- **Arquivo fragment**: caminho do `fragment.prisma`.\n"
        "- **Descricao**: o que o model representa.\n"
        "- **Alias historico**: nome antes do rename DDD."
    )
    acionaveis: list[list] = []
    apendice: list[list] = []
    for r in range(2, ws.max_row + 1):
        local = cell(ws, r, 1)
        nome_atual = cell(ws, r, 2)
        nome_ddd = cell(ws, r, 3)
        descricao = cell(ws, r, 8)
        tipo = cell(ws, r, 5)
        tem_tenant = cell(ws, r, 10)
        pk = cell(ws, r, 11)
        relacoes = cell(ws, r, 13)
        indices = cell(ws, r, 15)
        soft_delete = cell(ws, r, 16)
        auditoria = cell(ws, r, 17)
        arquivo_fragment = cell(ws, r, 23)
        if not local and not nome_ddd:
            continue
        row = [
            local,
            nome_ddd,
            tipo,
            tem_tenant,
            pk,
            relacoes,
            indices,
            soft_delete,
            auditoria,
            arquivo_fragment,
            descricao,
            nome_atual,
        ]
        if is_exempt(nome_ddd):
            apendice.append(row)
        else:
            acionaveis.append(row)
    acionaveis.sort(key=lambda x: (str(x[0]).lower(), str(x[1]).lower()))
    return write_atlas(
        output_dir,
        "03-models.md",
        3,
        "Models Prisma",
        como_ler,
        headers,
        acionaveis,
        apendice,
    )


def gerar_aba_04_enums(wb, output_dir: Path) -> Path:
    """Aba 4: 4. mapa-enums."""
    ws = wb["4. mapa-enums"]
    headers = [
        "Local",
        "Enum DDD",
        "Valor Prisma",
        "Usado em models",
        "Cor",
        "Icone",
        "E default?",
        "Descricao",
        "Alias historico",
    ]
    como_ler = (
        "- **Local**: produto/servico onde o enum esta declarado.\n"
        "- **Enum DDD**: nome canonical do enum (PascalCase, `<Entidade><Atributo>`).\n"
        "- **Valor Prisma**: valor literal (UPPER_SNAKE_CASE - constante tecnica do banco).\n"
        "- **Usado em models**: models Prisma que referenciam o enum.\n"
        "- **Cor / Icone**: marcadores de UI (badge).\n"
        "- **E default?**: se este e o valor padrao da coluna.\n"
        "- **Descricao**: significado do valor.\n"
        "- **Alias historico**: valor anterior (antes do rename DDD)."
    )
    acionaveis: list[list] = []
    apendice: list[list] = []
    for r in range(2, ws.max_row + 1):
        local = cell(ws, r, 1)
        enum_atual = cell(ws, r, 2)
        enum_ddd = cell(ws, r, 3)
        descricao = cell(ws, r, 4)
        valor_pg = cell(ws, r, 5)
        valor_pg_ddd = cell(ws, r, 6)
        valor_prisma = cell(ws, r, 7) or valor_pg_ddd or valor_pg
        usado_models = cell(ws, r, 10)
        cor = cell(ws, r, 13)
        icone = cell(ws, r, 14)
        is_default = cell(ws, r, 15)
        if not local and not enum_ddd and not valor_prisma:
            continue
        row = [
            local,
            enum_ddd,
            valor_prisma,
            usado_models,
            cor,
            icone,
            is_default,
            descricao,
            enum_atual,
        ]
        # Exempt se nao tem nem enum DDD nem valor
        if is_exempt(enum_ddd) and is_exempt(valor_prisma):
            apendice.append(row)
        else:
            acionaveis.append(row)
    acionaveis.sort(
        key=lambda x: (str(x[0]).lower(), str(x[1]).lower(), str(x[2]).lower())
    )
    return write_atlas(
        output_dir,
        "04-enums.md",
        4,
        "Enums",
        como_ler,
        headers,
        acionaveis,
        apendice,
    )


def gerar_aba_05_rotas_fe(wb, output_dir: Path) -> Path:
    """Aba 5: 5. mapa-rotas - rotas frontend (consumo do back)."""
    ws = wb["5. mapa-rotas"]
    headers = [
        "Produto",
        "Metodo",
        "Rota DDD",
        "Arquivo",
        "Patente minima",
        "Request Schema",
        "Response Schema",
        "Model Prisma",
        "Descricao",
        "Alias historico",
    ]
    como_ler = (
        "- **Produto**: produto/servico que consome a rota (chamada do front ou inter-servico).\n"
        "- **Metodo**: verbo HTTP.\n"
        "- **Rota DDD**: caminho final canonical.\n"
        "- **Arquivo**: arquivo da rota (no servico que a expoe).\n"
        "- **Patente minima**: `tipo_usuario` minimo permitido.\n"
        "- **Request / Response Schema**: schemas Zod do contrato bilateral.\n"
        "- **Model Prisma**: model principal envolvido.\n"
        "- **Descricao**: o que a rota faz.\n"
        "- **Alias historico**: rota antes do rename DDD."
    )
    acionaveis: list[list] = []
    apendice: list[list] = []
    for r in range(2, ws.max_row + 1):
        metodo = cell(ws, r, 1)
        rota_atual = cell(ws, r, 2)
        rota_ddd = cell(ws, r, 3)
        descricao = cell(ws, r, 4) or cell(ws, r, 17)
        arquivo = cell(ws, r, 5)
        produto = cell(ws, r, 9) or cell(ws, r, 8)
        req_schema = cell(ws, r, 12)
        resp_schema = cell(ws, r, 13)
        model_prisma = cell(ws, r, 14)
        patente = cell(ws, r, 19)
        if not metodo and not rota_ddd:
            continue
        row = [
            produto,
            metodo,
            rota_ddd,
            arquivo,
            patente,
            req_schema,
            resp_schema,
            model_prisma,
            descricao,
            rota_atual,
        ]
        if is_exempt(rota_ddd):
            apendice.append(row)
        else:
            acionaveis.append(row)
    acionaveis.sort(
        key=lambda x: (str(x[0]).lower(), str(x[2]).lower(), str(x[1]).lower())
    )
    return write_atlas(
        output_dir,
        "05-rotas-fe.md",
        5,
        "Rotas (consumo frontend / inter-servico)",
        como_ler,
        headers,
        acionaveis,
        apendice,
    )


def gerar_aba_06_paginas(wb, output_dir: Path) -> Path:
    """Aba 6: 6. mapa-paginas."""
    ws = wb["6. mapa-paginas"]
    headers = [
        "Produto",
        "URL DDD",
        "Arquivo DDD",
        "Titulo DDD",
        "Area",
        "Tipo view",
        "Path completo",
        "Descricao",
        "Alias historico (URL)",
    ]
    como_ler = (
        "- **Produto**: produto/servico onde a pagina vive.\n"
        "- **URL DDD**: URL final canonical da rota frontend (React Router).\n"
        "- **Arquivo DDD**: nome do arquivo `.tsx` apos rename DDD.\n"
        "- **Titulo DDD**: titulo exibido na pagina (label canonical).\n"
        "- **Area**: agrupamento navegacional (sidebar, area do produto).\n"
        "- **Tipo view**: lista, dashboard, kanban, formulario, etc.\n"
        "- **Path completo**: caminho do arquivo no monorepo.\n"
        "- **Descricao**: o que a pagina faz.\n"
        "- **Alias historico (URL)**: URL antes do rename DDD."
    )
    acionaveis: list[list] = []
    apendice: list[list] = []
    for r in range(2, ws.max_row + 1):
        local = cell(ws, r, 1)
        url_atual = cell(ws, r, 2)
        url_ddd = cell(ws, r, 3)
        descricao = cell(ws, r, 4)
        titulo_ddd = cell(ws, r, 6)
        arquivo_ddd = cell(ws, r, 8)
        area = cell(ws, r, 9)
        tipo_view = cell(ws, r, 10)
        path_completo = cell(ws, r, 11)
        if not local and not url_ddd and not arquivo_ddd:
            continue
        row = [
            local,
            url_ddd,
            arquivo_ddd,
            titulo_ddd,
            area,
            tipo_view,
            path_completo,
            descricao,
            url_atual,
        ]
        if is_exempt(url_ddd) and is_exempt(arquivo_ddd):
            apendice.append(row)
        else:
            acionaveis.append(row)
    acionaveis.sort(key=lambda x: (str(x[0]).lower(), str(x[1]).lower()))
    return write_atlas(
        output_dir,
        "06-paginas.md",
        6,
        "Paginas",
        como_ler,
        headers,
        acionaveis,
        apendice,
    )


def gerar_aba_07_modais(wb, output_dir: Path) -> Path:
    """Aba 7: 7. Modais."""
    ws = wb["7. Modais"]
    headers = [
        "Produto",
        "Arquivo DDD",
        "Componente DDD",
        "Paginas que abrem",
        "Acoes",
        "Rotas API",
        "Patente minima",
        "Path completo",
        "Descricao",
        "Alias historico",
    ]
    como_ler = (
        "- **Produto**: produto/servico onde o modal vive.\n"
        "- **Arquivo DDD**: nome do arquivo apos rename DDD.\n"
        "- **Componente DDD**: nome do componente React (PascalCase).\n"
        "- **Paginas que abrem**: paginas que disparam o modal.\n"
        "- **Acoes**: botoes/operacoes principais do modal.\n"
        "- **Rotas API**: rotas backend consumidas pelo modal.\n"
        "- **Patente minima**: `tipo_usuario` minimo permitido.\n"
        "- **Path completo**: caminho do arquivo no monorepo.\n"
        "- **Descricao**: o que o modal faz.\n"
        "- **Alias historico**: nome antes do rename DDD."
    )
    acionaveis: list[list] = []
    apendice: list[list] = []
    for r in range(2, ws.max_row + 1):
        local = cell(ws, r, 1)
        arquivo_atual = cell(ws, r, 2)
        arquivo_ddd = cell(ws, r, 3)
        descricao = cell(ws, r, 5)
        componente_atual = cell(ws, r, 6)
        componente_ddd = cell(ws, r, 7)
        paginas = cell(ws, r, 10)
        acoes = cell(ws, r, 11)
        rotas_api = cell(ws, r, 12)
        patente = cell(ws, r, 14)
        path_completo = cell(ws, r, 15)
        if not local and not arquivo_ddd and not componente_ddd:
            continue
        alias_hist = arquivo_atual
        if componente_atual and componente_atual != componente_ddd:
            alias_hist = (
                f"{arquivo_atual} | {componente_atual}" if arquivo_atual else componente_atual
            )
        row = [
            local,
            arquivo_ddd,
            componente_ddd,
            paginas,
            acoes,
            rotas_api,
            patente,
            path_completo,
            descricao,
            alias_hist,
        ]
        if is_exempt(arquivo_ddd) and is_exempt(componente_ddd):
            apendice.append(row)
        else:
            acionaveis.append(row)
    acionaveis.sort(key=lambda x: (str(x[0]).lower(), str(x[2]).lower()))
    return write_atlas(
        output_dir,
        "07-modais.md",
        7,
        "Modais",
        como_ler,
        headers,
        acionaveis,
        apendice,
    )


def gerar_aba_08_nucleo_global(wb, output_dir: Path) -> Path:
    """Aba 8: 8. Nucleo Global."""
    ws = wb["8. Nucleo Global"]
    headers = [
        "Categoria",
        "Componente DDD",
        "Tipo",
        "Usado por (produtos)",
        "Path",
        "Descricao",
        "Alias historico",
    ]
    como_ler = (
        "- **Categoria**: pasta sob `nucleo-global/` (Botoes, Tabelas, Formularios, ...).\n"
        "- **Componente DDD**: nome do componente apos rename DDD (PascalCase).\n"
        "- **Tipo**: ComponenteDSL, Modal, Form, Display, Hook, etc.\n"
        "- **Usado por (produtos)**: produtos/servicos que importam o componente.\n"
        "- **Path**: caminho do arquivo no monorepo.\n"
        "- **Descricao**: o que o componente faz.\n"
        "- **Alias historico**: nome antes do rename DDD."
    )
    acionaveis: list[list] = []
    apendice: list[list] = []
    for r in range(2, ws.max_row + 1):
        nome_atual = cell(ws, r, 1)
        nome_ddd = cell(ws, r, 2)
        descricao = cell(ws, r, 3)
        # Categoria nao esta preenchida na col D - extraimos do path
        path_componente = cell(ws, r, 9)
        categoria = ""
        if path_componente:
            parts = str(path_componente).replace("\\", "/").split("/")
            if len(parts) >= 2 and parts[0] == "nucleo-global":
                categoria = parts[1]
            elif len(parts) >= 1:
                categoria = parts[0]
        tipo = cell(ws, r, 6)
        usado_por = cell(ws, r, 7)
        if not nome_atual and not nome_ddd and not path_componente:
            continue
        row = [
            categoria,
            nome_ddd,
            tipo,
            usado_por,
            path_componente,
            descricao,
            nome_atual,
        ]
        if is_exempt(nome_ddd):
            apendice.append(row)
        else:
            acionaveis.append(row)
    acionaveis.sort(key=lambda x: (str(x[0]).lower(), str(x[1]).lower()))
    apendice.sort(key=lambda x: (str(x[4]).lower(),))
    return write_atlas(
        output_dir,
        "08-nucleo-global.md",
        8,
        "Nucleo Global",
        como_ler,
        headers,
        acionaveis,
        apendice,
    )


def gerar_aba_09_componentes_locais(wb, output_dir: Path) -> Path:
    """Aba 9: 9. Componentes Locais."""
    ws = wb["9. Componentes Locais"]
    headers = [
        "Produto",
        "Pasta",
        "Arquivo DDD",
        "Componente DDD",
        "Paginas que usam",
        "Path completo",
        "Descricao",
        "Alias historico",
    ]
    como_ler = (
        "- **Produto**: produto/servico onde o componente vive.\n"
        "- **Pasta**: pasta local dentro do produto (`components/`, `secoes/`, ...).\n"
        "- **Arquivo DDD**: nome do arquivo apos rename DDD.\n"
        "- **Componente DDD**: nome do componente React (PascalCase).\n"
        "- **Paginas que usam**: paginas locais que importam o componente.\n"
        "- **Path completo**: caminho do arquivo no monorepo.\n"
        "- **Descricao**: o que o componente faz.\n"
        "- **Alias historico**: nome antes do rename DDD."
    )
    acionaveis: list[list] = []
    apendice: list[list] = []
    for r in range(2, ws.max_row + 1):
        local = cell(ws, r, 1)
        arquivo_atual = cell(ws, r, 2)
        arquivo_ddd = cell(ws, r, 3)
        descricao = cell(ws, r, 4)
        componente_atual = cell(ws, r, 5)
        componente_ddd = cell(ws, r, 6)
        pasta = cell(ws, r, 8)
        usado_por = cell(ws, r, 12)
        path_completo = cell(ws, r, 13)
        if not local and not arquivo_ddd and not componente_ddd:
            continue
        alias_hist = arquivo_atual
        if componente_atual and componente_atual != componente_ddd:
            alias_hist = (
                f"{arquivo_atual} | {componente_atual}" if arquivo_atual else componente_atual
            )
        row = [
            local,
            pasta,
            arquivo_ddd,
            componente_ddd,
            usado_por,
            path_completo,
            descricao,
            alias_hist,
        ]
        if is_exempt(arquivo_ddd) and is_exempt(componente_ddd):
            apendice.append(row)
        else:
            acionaveis.append(row)
    acionaveis.sort(key=lambda x: (str(x[0]).lower(), str(x[3]).lower()))
    return write_atlas(
        output_dir,
        "09-componentes-locais.md",
        9,
        "Componentes Locais",
        como_ler,
        headers,
        acionaveis,
        apendice,
    )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> int:
    if len(sys.argv) < 2:
        print(__doc__)
        return 1
    planilha = Path(sys.argv[1]).resolve()
    if not planilha.exists():
        print(f"ERRO: planilha nao encontrada: {planilha}")
        return 2

    repo_root = Path(__file__).resolve().parents[2]
    output_dir = repo_root / "documentos-tecnicos" / "ddd-atlas"
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"Lendo planilha: {planilha}")
    wb = openpyxl.load_workbook(planilha, data_only=True, read_only=False)

    geradores: list[Callable] = [
        gerar_aba_01_campos,
        gerar_aba_02_rotas_api,
        gerar_aba_03_models,
        gerar_aba_04_enums,
        gerar_aba_05_rotas_fe,
        gerar_aba_06_paginas,
        gerar_aba_07_modais,
        gerar_aba_08_nucleo_global,
        gerar_aba_09_componentes_locais,
    ]

    out_files: list[Path] = []
    for fn in geradores:
        out = fn(wb, output_dir)
        size = out.stat().st_size
        with out.open("r", encoding="utf-8") as f:
            n_lines = sum(1 for _ in f)
        print(f"  -> {out.relative_to(repo_root)} ({n_lines} linhas, {size} bytes)")
        out_files.append(out)

    print(f"\nOK: {len(out_files)} arquivos gerados em {output_dir.relative_to(repo_root)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
