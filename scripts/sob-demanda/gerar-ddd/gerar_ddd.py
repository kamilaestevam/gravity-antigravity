"""
Gera a planilha DDD preenchida a partir dos schemas Prisma dos 3 bancos
em escopo (Configurador, Serviços, Pedido) + rotas Express.

Regras aprovadas pelo dono em 2026-04-21.
Entrada: planilha_geral_gravity (2).xlsx (template)
Saída:   planilha_geral_gravity_PREENCHIDA.xlsx
"""
from __future__ import annotations
import os, re, json, glob
from dataclasses import dataclass, field
from collections import Counter, defaultdict
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = r'C:\Users\danie\gravity-antigravity'
TEMPLATE = r'C:\Users\danie\Downloads\planilha_geral_gravity (2).xlsx'
OUTPUT   = r'C:\Users\danie\Downloads\planilha_geral_gravity_PREENCHIDA.xlsx'

SCHEMAS = [
    # (label, bank_label, produto_label, schema_path)
    ('Configurador', 'Banco de Dados - Configurador', 'Configurador',
        os.path.join(ROOT, 'configurador', 'prisma', 'schema.prisma')),
    ('Serviços',     'Banco de Dados - Serviços',    'Serviços (Tenant)',
        os.path.join(ROOT, 'servicos-global', 'tenant', 'prisma', 'schema.prisma')),
    ('Pedido',       'Banco de Dados - Pedido',      'Pedido',
        os.path.join(ROOT, 'produto', 'pedido', 'server', 'prisma', 'schema.prisma')),
]

SCALAR_TYPES = {'String','Int','BigInt','Float','Decimal','Boolean','DateTime','Json','Bytes'}
AUDIT_FIELDS = {'id','tenant_id','id_organizacao','company_id','id_workspace','created_at','updated_at','deleted_at'}

# Mapeamento DDD conhecido (Mandamento 03 — tabela oficial)
DDD_LEGACY_MAP = {
    'tenant_id'   : 'id_organizacao',
    'company_id'  : 'id_workspace',
    'user_id'     : 'id_usuario',
    'role'        : 'tipo_usuario',
    'is_super_admin': 'is_gravity_admin',
}

# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------
@dataclass
class Field:
    name: str
    type_token: str           # ex: String?, Usuario[], TipoUsuario
    base_type: str            # String, Usuario, TipoUsuario
    is_list: bool
    is_optional: bool
    attrs: str                # resto da linha após o tipo (inclui @map, @default, etc)
    map_to: str | None = None
    default: str | None = None
    is_unique: bool = False
    is_id: bool = False
    relation: str | None = None  # cardinalidade se relation

@dataclass
class Model:
    name: str
    fields: list[Field] = field(default_factory=list)

@dataclass
class Enum:
    name: str
    values: list[str] = field(default_factory=list)

@dataclass
class Schema:
    label: str
    models: list[Model] = field(default_factory=list)
    enums:  list[Enum]  = field(default_factory=list)

def parse_schema(path: str, label: str) -> Schema:
    with open(path, encoding='utf-8') as f:
        lines = f.readlines()

    sch = Schema(label=label)
    model: Model | None = None
    enum: Enum | None = None

    for raw in lines:
        s = raw.strip()
        if s.startswith('//') or not s:
            continue
        if s.startswith('generator ') or s.startswith('datasource '):
            continue
        if s.startswith('model '):
            model = Model(name=s.split()[1])
            sch.models.append(model)
            enum = None
            continue
        if s.startswith('enum '):
            enum = Enum(name=s.split()[1])
            sch.enums.append(enum)
            model = None
            continue
        if s == '}':
            model = None; enum = None
            continue
        if model is None and enum is None:
            continue
        if s.startswith('@@'):  # index/unique/map a nível de modelo
            continue

        if enum is not None:
            m = re.match(r'^([A-Z_][A-Z0-9_]*)', s)
            if m:
                enum.values.append(m.group(1))
            continue

        m = re.match(r'^\s*([a-zA-Z_][a-zA-Z0-9_]*)\s+([A-Za-z][A-Za-z0-9_]*(?:\[\])?(?:\?)?)\s*(.*)$', raw.rstrip())
        if not m:
            continue
        fname, ftype, attrs = m.groups()
        is_list = ftype.endswith('[]')
        is_opt  = ftype.endswith('?')
        base    = re.sub(r'[\[\]\?]', '', ftype)
        f = Field(name=fname, type_token=ftype, base_type=base,
                  is_list=is_list, is_optional=is_opt, attrs=attrs)
        # @map
        mm = re.search(r'@map\("([^"]+)"\)', attrs)
        if mm: f.map_to = mm.group(1)
        md = re.search(r'@default\((.+?)\)\s*(?:@|$)', attrs)
        if md: f.default = md.group(1).strip()
        else:
            md2 = re.search(r'@default\(([^)]*\([^)]*\)[^)]*)\)', attrs)
            if md2: f.default = md2.group(1).strip()
            else:
                md3 = re.search(r'@default\(([^)]*)\)', attrs)
                if md3: f.default = md3.group(1).strip()
        if '@unique' in attrs: f.is_unique = True
        if '@id' in attrs:     f.is_id = True
        if '@relation' in attrs or base[:1].isupper() and base not in SCALAR_TYPES:
            # Classificar depois quando soubermos enums/models
            pass
        model.fields.append(f)
    return sch

# ---------------------------------------------------------------------------
# Classificação
# ---------------------------------------------------------------------------
def classify_field(f: Field, all_models: set[str], all_enums: set[str]) -> str:
    if f.base_type in SCALAR_TYPES: return 'scalar'
    if f.base_type in all_enums:    return 'enum_ref'
    if f.base_type in all_models:   return 'relation'
    return 'unknown'

# ---------------------------------------------------------------------------
# Nomenclatura DDD
# ---------------------------------------------------------------------------
def ddd_name(field: Field, entity: str, name_counts: Counter, kind: str) -> str:
    """
    Aplica as regras:
    - @map presente + diferente do nome Prisma → usa o valor do @map (já é DDD ou alvo desejado)
    - tenant_id → id_organizacao, etc.
    - Nomes genéricos (status, tipo, nome, descricao, etc) + existem em >1 model → sufixar
    """
    if kind == 'relation':
        # Para relations não temos DDD no banco; devolvemos string vazia
        return ''
    base = field.name
    # Se há @map explícito, ele já revela o DDD alvo
    if field.map_to and field.map_to != base:
        return field.map_to
    # Mapeamento oficial da Regra 03
    if base in DDD_LEGACY_MAP:
        return DDD_LEGACY_MAP[base]
    # Heurística: name → nome_<entidade>, email → email_<entidade>
    entity_slug = re.sub(r'(?<!^)(?=[A-Z])', '_', entity).lower()
    # Remover prefixos de módulo (ex: AtividadesDados → atividade)
    entity_slug = re.sub(r'_dados$|_data$', '', entity_slug)
    if base == 'name':  return f'nome_{entity_slug}'
    if base == 'email': return f'email_{entity_slug}'
    GENERIC = {'status','tipo','nome','descricao','titulo','categoria','prioridade','codigo'}
    if base in GENERIC and name_counts[base] > 1:
        return f'{base}_{entity_slug}'
    return base  # já em DDD ou não-ambíguo

def ddd_back_front(prisma_name: str, ddd: str) -> str:
    """Regra 5: Back DDD = Front DDD = nome DDD do banco."""
    return ddd or prisma_name

def tela_atual(field_name: str) -> str:
    """Heurística simples: snake → Title Case human."""
    if not field_name: return ''
    parts = field_name.replace('id_','').split('_')
    return ' '.join(p.capitalize() for p in parts if p)

def tela_ddd(atual: str) -> str:
    """Expande abreviações comuns. Mantém Nº e %."""
    out = atual
    out = re.sub(r'\bQtd\.?\b', 'Quantidade', out, flags=re.I)
    out = re.sub(r'\bQtde\b', 'Quantidade', out, flags=re.I)
    out = re.sub(r'\bNum\.?\b', 'Número', out, flags=re.I)
    out = re.sub(r'\bDesc\.?\b', 'Descrição', out, flags=re.I)
    return out

# ---------------------------------------------------------------------------
# Descoberta de contagens para heurística DDD
# ---------------------------------------------------------------------------
def build_name_counts(schemas: list[Schema]) -> Counter:
    c = Counter()
    for s in schemas:
        for m in s.models:
            for f in m.fields:
                c[f.name] += 1
    return c

# ---------------------------------------------------------------------------
# Geração de linhas
# ---------------------------------------------------------------------------
COLS = [
    'Local','Tabela','Entidade',
    'Nome no banco - PostgreSQL','Nome no banco - Prisma','Nome no banco - DDD',
    'Nome no back - Atual','Nome no back - DDD',
    'Nome no front - Atual','Nome no front - DDD',
    'Nome em tela - Atual','Nome em tela - DDD',
    'Local Tela','Tipo Dado','Produto Gravity',
    'Natureza','Tipo de Dado','Formato','Validação',
    'Obrigatório','Editável','Valor Padrão','Exemplo',
    'Descrição','Componente','Origem','Padrão Universal',
]

def origem(f: Field) -> str:
    if f.is_id: return 'Sistema'
    if f.default and 'now()' in f.default: return 'Sistema'
    if f.default and 'cuid()' in f.default: return 'Sistema'
    if f.default and 'uuid()' in f.default: return 'Sistema'
    return 'Manual'

def obrigatorio(f: Field) -> str:
    return 'Não' if f.is_optional else 'Sim'

def editavel(f: Field) -> str:
    if f.is_id: return 'Não'
    if f.name in AUDIT_FIELDS: return 'Não'
    return 'Sim'

def tipo_dado_and_componente(kind: str, f: Field, enums_set: set[str]) -> tuple[str,str,str,str]:
    """Retorna (Tipo Dado, Tipo de Dado detalhado, Formato, Componente)"""
    if kind == 'relation':
        card = 'N:1' if not f.is_list else '1:N'
        fmt = f'{card} → {f.base_type}'
        return ('Relação', 'Relation', fmt, '—')
    if kind == 'enum_ref':
        return ('Enum', f.base_type, f'→ {f.base_type}', 'select')
    # scalar
    base = f.base_type
    comp_map = {
        'String':'text','Int':'number','BigInt':'number','Float':'number',
        'Decimal':'number','Boolean':'toggle','DateTime':'date',
        'Json':'display','Bytes':'hidden',
    }
    # se é id ou chave estrangeira, esconde
    if f.is_id or f.name.endswith('_id'):
        comp = 'hidden'
    else:
        comp = comp_map.get(base, 'text')
    # Tipo Dado exibido
    td_display = 'SUID' if f.is_id else base
    return (td_display, base, '', comp)

def make_field_row(sch: Schema, bank_label: str, produto: str,
                   model: Model, f: Field, name_counts: Counter,
                   all_models: set[str], all_enums: set[str]) -> list:
    kind = classify_field(f, all_models, all_enums)
    pg_name     = f.map_to if f.map_to else (f.name if kind != 'relation' else '')
    prisma_name = f.name
    ddd         = ddd_name(f, model.name, name_counts, kind)
    # Se a coluna física já é DDD (ex: @map("id_organizacao_usuario")), DDD == pg_name
    if kind == 'relation':
        ddd_banco = ''
        back_atual  = prisma_name
        back_ddd    = ddd_back_front_for_relation(prisma_name)
        front_atual = prisma_name
        front_ddd   = back_ddd
    else:
        ddd_banco = ddd or pg_name
        back_atual  = prisma_name
        back_ddd    = ddd_back_front(prisma_name, ddd_banco)
        front_atual = prisma_name
        front_ddd   = back_ddd

    tela_a = tela_atual(prisma_name)
    tela_d = tela_ddd(tela_a)

    tipo_dado, tipo_detalhe, fmt, comp = tipo_dado_and_componente(kind, f, all_enums)

    # Natureza
    if kind == 'relation':
        natureza = 'virtual'
    elif f.is_id:
        natureza = 'sistema'
    else:
        natureza = 'fisico'

    # Validação (heurística leve)
    validacao = ''
    if kind == 'scalar' and f.base_type == 'String' and not f.is_id and not f.name.endswith('_id'):
        validacao = 'min(1) max(255)'

    # Padrão universal
    padrao = 'sim' if f.name in AUDIT_FIELDS else ''

    # Valor Padrão (só se existe @default simples)
    vdef = ''
    if f.default:
        vd = f.default
        if re.search(r'\b(now|cuid|uuid|autoincrement)\(\)', vd):
            vdef = ''  # gerado pelo sistema
        elif vd.startswith('"') and vd.endswith('"'):
            vdef = vd[1:-1]
        else:
            vdef = vd

    # Descrição
    if kind == 'relation':
        desc = f'Relação Prisma ({"1:N" if f.is_list else "N:1"}) para {f.base_type} — navegação, não coluna física.'
    elif kind == 'enum_ref':
        desc = f'Valor do enum {f.base_type}. Ver linha dedicada do enum para valores permitidos.'
    elif f.base_type == 'Json':
        desc = 'Campo JSON — estrutura interna pode mudar sem migration; não é contrato DDD.'
    elif f.name in AUDIT_FIELDS:
        desc = 'Campo padrão de auditoria / isolamento.'
    else:
        desc = ''

    exemplo = ''  # deixar vazio por default (dono preenche)

    return [
        bank_label,                     # Local
        model.name,                     # Tabela
        model.name,                     # Entidade
        pg_name,                        # Nome no banco - PostgreSQL
        prisma_name,                    # Nome no banco - Prisma
        ddd_banco,                      # Nome no banco - DDD
        back_atual,                     # Nome no back - Atual
        back_ddd,                       # Nome no back - DDD
        front_atual,                    # Nome no front - Atual
        front_ddd,                      # Nome no front - DDD
        tela_a,                         # Nome em tela - Atual
        tela_d,                         # Nome em tela - DDD
        local_tela_para_tabela(model.name, produto),   # Local Tela
        tipo_dado,                      # Tipo Dado
        produto,                        # Produto Gravity
        natureza,                       # Natureza
        tipo_detalhe,                   # Tipo de Dado
        fmt,                            # Formato
        validacao,                      # Validação
        obrigatorio(f),                 # Obrigatório
        editavel(f),                    # Editável
        vdef,                           # Valor Padrão
        exemplo,                        # Exemplo
        desc,                           # Descrição
        comp,                           # Componente
        origem(f),                      # Origem
        padrao,                         # Padrão Universal
    ]

def ddd_back_front_for_relation(prisma_name: str) -> str:
    """users → usuarios, workspaces → workspaces, organization → organizacao."""
    table = {
        'users':'usuarios','user':'usuario',
        'organization':'organizacao','organizations':'organizacoes',
        'company':'workspace','companies':'workspaces',
    }
    return table.get(prisma_name, prisma_name)

def make_enum_row(sch: Schema, bank_label: str, produto: str, e: Enum) -> list:
    return [
        bank_label, e.name, e.name,
        e.name, e.name, e.name,                      # PG/Prisma/DDD = nome do enum (tipo real do PostgreSQL)
        e.name, e.name,                              # back atual/ddd = nome do enum
        e.name, e.name,                              # front atual/ddd
        e.name, e.name,                              # tela
        local_tela_para_tabela(e.name, produto),     # Local Tela
        'Enum', produto, 'sistema', 'Enum',
        '|'.join(e.values),                          # Formato
        'valor deve estar na lista',                 # Validação
        'Sim','Não','','','Definição de enum — usada por campos que referenciam este tipo.',
        '—','Sistema','',
    ]

# Mapeamento tabela -> Local na Tela (breadcrumb hierárquico).
# Match por prefixo, case-insensitive.
TABLE_TO_SCREEN = [
    # --- Configurador ---
    ('Organizacao',              'Admin / Organizações'),
    ('Empresa',                  'Admin / Workspaces'),
    ('UsuarioEmpresa',           'Admin / Workspaces / Membros'),
    ('UsuarioWorkspace',         'Admin / Workspaces / Membros'),
    ('UsuarioPermissao',         'Admin / Permissões'),
    ('Usuario',                  'Admin / Usuários; Hub / Perfil'),
    ('AssinaturaProdutoGravity', 'Admin / Produtos Gravity / Assinaturas'),
    ('ConfiguracaoProduto',      'Admin / Produtos Gravity / Configuração'),
    ('ProdutoGravity',           'Admin / Produtos Gravity'),
    ('TokenServico',             'Admin / API Cockpit / Tokens'),
    ('TokenAcesso',              'Admin / API Cockpit / Tokens'),
    ('Fatura',                   'Admin / Financeiro / Faturas'),
    ('Cobranca',                 'Admin / Financeiro / Cobranças'),
    ('Assinatura',               'Admin / Financeiro / Assinaturas'),
    ('Deploy',                   'Admin / Deploy'),
    ('AmbienteDeploy',           'Admin / Deploy'),
    ('MetricasGemini',           'Admin / Segurança / Telemetria'),
    ('LogGerador',               'Admin / Segurança / Logs'),
    ('EscopoToken',              'Admin / API Cockpit / Tokens'),
    ('OrganizacaoStatus',        'Admin / Organizações (seletor)'),
    ('StatusAssinaturaProdutoGravity', 'Admin / Financeiro (seletor)'),
    ('StatusEmpresa',            'Admin / Workspaces (seletor)'),
    ('StatusProduto',            'Admin / Produtos Gravity (seletor)'),
    ('StatusDeploy',             'Admin / Deploy (seletor)'),
    ('TipoUsuario',              'Login / Sign-up; Admin / Usuários (seletor)'),
    ('TipoMembroEmpresa',        'Admin / Workspaces / Membros (seletor)'),
    ('TipoCobranca',             'Admin / Financeiro (seletor)'),
    ('TipoLimiteUsuario',        'Admin / Usuários (seletor)'),
    ('FaturaStatus',             'Admin / Financeiro / Faturas (seletor)'),

    # --- Serviços (Tenant) ---
    ('Atividades',               'Atividades'),
    ('Agenda',                   'Agenda'),
    ('ApiCockpit',               'API Cockpit'),
    ('ConectorErp',              'Conector ERP'),
    ('Cronometro',               'Cronômetro'),
    ('Dashboard',                'Dashboard'),
    ('Email',                    'Email'),
    ('Gabi',                     'Gabi'),
    ('HistoricoGlobal',          'Histórico Global'),
    ('Historico',                'Histórico'),
    ('NcmSync',                  'NCM Sync'),
    ('Ncm',                      'NCM'),
    ('Notificacao',              'Notificações'),
    ('PreferenciasUsuario',      'Preferências / Usuário'),
    ('Preferencia',              'Preferências'),
    ('Relatorio',                'Relatórios'),
    ('Whatsapp',                 'WhatsApp'),
    ('Disponibilidade',          'Agenda / Disponibilidade'),
    ('Reserva',                  'Agenda / Reservas'),
    ('Slot',                     'Agenda / Slots'),

    # --- Pedido ---
    ('PedidoItemLote',           'Pedido / Detalhe / Itens / Lote'),
    ('PedidoItem',               'Pedido / Detalhe / Itens'),
    ('PedidoAnexo',              'Pedido / Detalhe / Anexos'),
    ('PedidoHistorico',          'Pedido / Detalhe / Histórico'),
    ('Pedido',                   'Pedidos / Lista; Pedidos / Detalhe; Pedidos / Kanban; Pedidos / Dashboard'),
    ('TrackingItem',             'Pedidos / Histórico / Tracking'),
    ('StatusPedido',             'Pedidos (seletor)'),
    ('CampoCustom',              'Pedidos / Configurações / Campos Customizados'),
    ('Configuracao',             'Pedidos / Configurações'),
    ('Incoterm',                 'Pedidos (seletor)'),
    ('Moeda',                    'Pedidos (seletor)'),
]

def local_tela_para_tabela(table_name: str, produto: str = '') -> str:
    """Mapeia nome da tabela (modelo/enum) para breadcrumb de tela."""
    for prefix, screen in TABLE_TO_SCREEN:
        if table_name.startswith(prefix) or table_name == prefix:
            return screen
    # fallbacks por produto
    if 'Pedido' in produto: return 'Pedidos'
    if 'Tenant' in produto or 'Serviços' in produto: return '—'
    if 'Configurador' in produto: return 'Admin'
    return ''

# ---------------------------------------------------------------------------
# APIs
# ---------------------------------------------------------------------------
API_COLS = ['Método','Rota','Tipo','Tipo Auth','Request Schema','Response Schema','Entidade Retornada','Consumidor','Descrição']

def parse_route_files() -> list[list]:
    dirs = [
        ('Configurador', os.path.join(ROOT, 'servicos-global', 'configurador', 'server')),
        ('Serviços',     os.path.join(ROOT, 'servicos-global', 'tenant')),
        ('Pedido',       os.path.join(ROOT, 'produto', 'pedido', 'server')),
    ]
    rows = []
    for label, root in dirs:
        for path in glob.glob(os.path.join(root, '**', '*.ts'), recursive=True):
            norm = path.replace('\\','/')
            if 'node_modules' in norm: continue
            if '.test.' in norm or '.spec.' in norm: continue
            if '/__tests__/' in norm or '/tests/' in norm: continue
            try:
                with open(path, encoding='utf-8') as f: src = f.read()
            except: continue
            for m in re.finditer(r"\b(\w*[Rr]outer|app|api)\.(get|post|put|patch|delete)\(\s*['\"`]([^'\"`]+)['\"`]", src):
                method = m.group(2).upper()
                route  = m.group(3)
                is_internal = '/internal' in route or '/api/internal' in route
                tipo = 'interno S2S' if is_internal else 'público'
                auth = 'x-internal-key' if is_internal else 'Clerk JWT'
                ctx = src[m.start():m.start()+500]
                req_schema = ''
                res_schema = ''
                sm = re.search(r'(\w+Schema)\.parse', ctx)
                if sm: req_schema = sm.group(1)
                rm = re.search(r'(\w+ResponseSchema)', ctx)
                if rm: res_schema = rm.group(1)
                # Reconstruir prefixo do router a partir do arquivo — procurar app.use('/prefix', xxxRouter)
                router_name = m.group(1)
                prefix = ''
                if router_name and 'router' in router_name.lower():
                    # Buscar em qualquer arquivo referência a este router
                    # (simplificado: deixa só a rota relativa, o prefixo pode ser adicionado manualmente)
                    pass
                rows.append([method, route, tipo, auth, req_schema, res_schema, '', label, ''])
    return rows

# ---------------------------------------------------------------------------
# Escrita
# ---------------------------------------------------------------------------
def main():
    schemas = [(label, bank, prod, parse_schema(p, label)) for (label,bank,prod,p) in SCHEMAS]
    all_schemas = [t[3] for t in schemas]

    # Coleta universos de nomes (models e enums) cross-bancos
    all_models = set()
    all_enums  = set()
    for s in all_schemas:
        all_models.update(m.name for m in s.models)
        all_enums.update(e.name  for e in s.enums)

    name_counts = build_name_counts(all_schemas)

    # Carrega template e preserva primeira aba estrutura; adiciona coluna "Padrão Universal"
    wb = load_workbook(TEMPLATE)
    ws = wb.active
    # Limpar linhas de exemplo existentes (linhas 2..max_row)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    # Inserir coluna "Padrão Universal" na posição 27 (após Origem)
    # Detectar posição de "Origem"
    header = [c.value for c in ws[1]]
    try:
        origem_idx = header.index('Origem') + 1   # 1-based
    except ValueError:
        origem_idx = 26
    ws.insert_cols(origem_idx + 1)
    ws.cell(row=1, column=origem_idx + 1, value='Padrão Universal')

    # Estilo do header novo (roubado da célula anterior)
    try:
        src_cell = ws.cell(row=1, column=origem_idx)
        tgt_cell = ws.cell(row=1, column=origem_idx + 1)
        tgt_cell.font = src_cell.font.copy()
        tgt_cell.fill = PatternFill('solid', start_color='4F46E5')
        tgt_cell.alignment = src_cell.alignment.copy() if src_cell.alignment else Alignment(horizontal='center')
    except: pass

    # Escrever linhas
    row_idx = 2
    for label, bank, produto, sch in schemas:
        # Enums primeiro — 1 linha cada
        for e in sch.enums:
            vals = make_enum_row(sch, bank, produto, e)
            # Precisamos expandir para o total de colunas do template
            write_row(ws, row_idx, vals)
            row_idx += 1
        # Campos de cada model
        for model in sch.models:
            for fld in model.fields:
                vals = make_field_row(sch, bank, produto, model, fld, name_counts, all_models, all_enums)
                write_row(ws, row_idx, vals)
                row_idx += 1

    # Aba APIs
    if 'APIs' in wb.sheetnames:
        del wb['APIs']
    ws_api = wb.create_sheet('APIs')
    ws_api.append(API_COLS)
    # Bolder header
    for col_i, _ in enumerate(API_COLS, start=1):
        c = ws_api.cell(row=1, column=col_i)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', start_color='1F2937')
        c.alignment = Alignment(horizontal='center')

    api_rows = parse_route_files()
    # Dedup por (método, rota)
    seen = set()
    for r in api_rows:
        key = (r[0], r[1])
        if key in seen: continue
        seen.add(key)
        ws_api.append(r)

    wb.save(OUTPUT)

    # Resumo
    print(f'Arquivo gerado: {OUTPUT}')
    print(f'Linhas de campos: {row_idx - 2}')
    print(f'Linhas de APIs:   {len(seen)}')
    print()
    for label, bank, produto, sch in schemas:
        nf = sum(len(m.fields) for m in sch.models)
        ne = len(sch.enums)
        print(f'  {label:<18} models={len(sch.models):>3}  fields={nf:>4}  enums={ne:>3}')

def write_row(ws, row_idx, vals):
    """Escreve vals nas primeiras len(COLS) colunas da linha."""
    for col_i, val in enumerate(vals, start=1):
        ws.cell(row=row_idx, column=col_i, value=val)

if __name__ == '__main__':
    main()
