# -*- coding: utf-8 -*-
"""
Gera tabela_pedido_populada_10.xlsx — 10 pedidos com itens variáveis.
Schema base: servicos-global/produto/pedido/prisma/schema.prisma (models Pedido + PedidoItem)

Saída: C:/Users/danie/gravity-antigravity/tabela_pedido_populada_10.xlsx
"""
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────

ID_ORG = "org_test_cde_pro"
ID_WS  = "ws_test_main"

# Quantidade de itens por pedido (padrão pedido pelo usuário)
ITENS_POR_PEDIDO = [1, 2, 3, 5, 1, 4, 2, 7, 3, 6]  # total = 34 itens

# ─────────────────────────────────────────────────────────────────────────────
# DADOS REALISTAS
# ─────────────────────────────────────────────────────────────────────────────

# Cada tupla é um cenário de pedido completo
PEDIDOS_DATA = [
    # (idx, tipo_op, status, incoterm, moeda, taxa_camb, exportador, importador, fabricante, pais_origem, cobertura, num_volumes)
    (1,  "importacao",  "rascunho",      "FOB",  "USD", 5.4250, "FXC_TECH_001", "GVT_BR_IMP_001", "FXC_TECH_001",  "CN", "com_cobertura",  240),
    (2,  "importacao",  "aberto",        "CIF",  "EUR", 5.8900, "BSCH_DE_002",   "GVT_BR_IMP_001", "BSCH_DE_002",   "DE", "com_cobertura",  120),
    (3,  "importacao",  "em_andamento",  "FOB",  "USD", 5.4180, "SMSG_KR_003",   "GVT_BR_IMP_001", "SMSG_KR_003",   "KR", "com_cobertura",  500),
    (4,  "exportacao",  "aprovado",      "FCA",  "USD", 5.4320, "GVT_BR_EXP_001", "ACME_US_004",   "GVT_BR_EXP_001", "BR", "sem_cobertura",   80),
    (5,  "importacao",  "transferencia", "DDP",  "JPY", 0.0364, "MTSH_JP_005",   "GVT_BR_IMP_001", "MTSH_JP_005",   "JP", "com_cobertura",  300),
    (6,  "importacao",  "consolidado",   "EXW",  "USD", 5.4100, "BYDA_CN_006",   "GVT_BR_IMP_001", "BYDA_CN_006",   "CN", "com_cobertura",  150),
    (7,  "exportacao",  "aberto",        "FOB",  "USD", 5.4250, "GVT_BR_EXP_001", "EUROIMP_007",   "GVT_BR_EXP_001", "BR", "com_cobertura",   60),
    (8,  "importacao",  "em_andamento",  "CIP",  "USD", 5.4180, "XIAO_CN_008",   "GVT_BR_IMP_001", "XIAO_CN_008",   "CN", "com_cobertura", 1200),
    (9,  "importacao",  "cancelado",     "FOB",  "EUR", 5.8950, "MAGN_DE_009",   "GVT_BR_IMP_001", "MAGN_DE_009",   "DE", "sem_cobertura",  100),
    (10, "exportacao",  "rascunho",      "FCA",  "USD", 5.4400, "GVT_BR_EXP_001", "TARG_US_010",   "GVT_BR_EXP_001", "BR", "com_cobertura",  450),
]

# Itens — cada lista representa os itens do pedido N (sequencia 1..n)
# (part_number, ncm, descricao, qtd, valor_unitario, unidade, peso_unit, peso_bruto_unit, cubagem_unit)
ITENS_DATA = {
    # Pedido 1 — China, roteadores Foxconn (1 item)
    1: [
        ("FXC-RT-AC1900", "8517.62.59", "Roteador Wi-Fi 6 dual-band AC1900",            240, 28.50, "PCS", 0.420, 0.580, 0.0080),
    ],
    # Pedido 2 — Alemanha, sensores Bosch (2 itens)
    2: [
        ("BSCH-PMS-100",  "9027.50.20", "Sensor de pressão industrial 0-100bar",        80,  145.20, "PCS", 0.180, 0.290, 0.0012),
        ("BSCH-TEMP-DS",  "9025.19.10", "Sensor de temperatura digital -40~+125 C",     40,  62.40,  "PCS", 0.060, 0.120, 0.0005),
    ],
    # Pedido 3 — Coréia, telas LCD Samsung (3 itens)
    3: [
        ("SMSG-LCD-15.6", "8528.59.20", "Painel LCD 15.6 polegadas Full HD IPS",        200, 78.00, "PCS", 0.520, 0.680, 0.0085),
        ("SMSG-LCD-13.3", "8528.59.20", "Painel LCD 13.3 polegadas Full HD IPS",        180, 64.50, "PCS", 0.380, 0.510, 0.0062),
        ("SMSG-LED-22",   "8528.72.00", "Monitor LED 22 polegadas full HD",             120, 145.00, "PCS", 3.200, 4.100, 0.0420),
    ],
    # Pedido 4 — EUA, exportação calçados brasileiros (5 itens)
    4: [
        ("BR-SP-AIR-42",   "6403.99.90", "Tenis esportivo Air Pro tamanho 42",          150, 32.50, "PAR", 0.480, 0.620, 0.0048),
        ("BR-SP-AIR-43",   "6403.99.90", "Tenis esportivo Air Pro tamanho 43",          150, 32.50, "PAR", 0.490, 0.640, 0.0050),
        ("BR-SP-AIR-44",   "6403.99.90", "Tenis esportivo Air Pro tamanho 44",          120, 32.50, "PAR", 0.510, 0.660, 0.0052),
        ("BR-SP-RUN-41",   "6403.99.90", "Tenis Running ProMax tamanho 41",             100, 45.80, "PAR", 0.460, 0.610, 0.0046),
        ("BR-SP-RUN-42",   "6403.99.90", "Tenis Running ProMax tamanho 42",             100, 45.80, "PAR", 0.470, 0.620, 0.0048),
    ],
    # Pedido 5 — Japão, instrumento óptico Mitsubishi (1 item)
    5: [
        ("MTSH-MIC-X1000", "9011.20.20", "Microscopio binocular X1000 com camera",      30,  1850.00, "UN",  8.500, 12.300, 0.1850),
    ],
    # Pedido 6 — China, baterias BYD (4 itens)
    6: [
        ("BYDA-BAT-100AH", "8507.60.00", "Bateria de litio 100Ah 12V LiFePO4",          50, 285.00, "UN", 14.200, 16.500, 0.0250),
        ("BYDA-BAT-200AH", "8507.60.00", "Bateria de litio 200Ah 12V LiFePO4",          40, 540.00, "UN", 27.500, 31.000, 0.0480),
        ("BYDA-BMS-12V",   "8504.40.20", "BMS para bateria 12V com bluetooth",          50, 78.50,  "UN",  0.420, 0.580, 0.0018),
        ("BYDA-CABO-50A",  "8544.42.00", "Cabo de potencia 50A 4AWG par 2m",            100, 14.20, "PAR", 1.200, 1.400, 0.0035),
    ],
    # Pedido 7 — Exportacao, soja para EU (2 itens)
    7: [
        ("BR-SOJA-A1",  "1201.90.00", "Soja em graos tipo exportacao A1",               45000, 0.42, "KG", 1.000, 1.000, 0.0011),
        ("BR-SOJA-A2",  "1201.90.00", "Soja em graos tipo exportacao A2",               15000, 0.39, "KG", 1.000, 1.000, 0.0011),
    ],
    # Pedido 8 — China, eletronicos Xiaomi (7 itens)
    8: [
        ("XIAO-PHN-13",   "8517.13.00", "Smartphone Mi 13 128GB",                       300, 285.00, "PCS", 0.220, 0.450, 0.0030),
        ("XIAO-PHN-13PR", "8517.13.00", "Smartphone Mi 13 Pro 256GB",                   200, 425.00, "PCS", 0.230, 0.480, 0.0032),
        ("XIAO-WTC-S2",   "9102.12.10", "Smartwatch Mi Watch S2 GPS",                   150, 92.50,  "PCS", 0.120, 0.250, 0.0015),
        ("XIAO-EAR-PRO3", "8518.30.00", "Fone Bluetooth Mi Buds 3 Pro",                 250, 38.40,  "PAR", 0.080, 0.180, 0.0008),
        ("XIAO-PWR-20K",  "8504.40.90", "Powerbank Mi 20000mAh USB-C 33W",              200, 28.90,  "PCS", 0.380, 0.520, 0.0045),
        ("XIAO-CAB-1M",   "8544.42.00", "Cabo USB-C 1m com tecido trancado",            500, 3.20,   "PCS", 0.050, 0.080, 0.0003),
        ("XIAO-CHRG-67W", "8504.40.90", "Carregador GaN 67W USB-C PD",                  150, 18.80,  "PCS", 0.180, 0.280, 0.0012),
    ],
    # Pedido 9 — Alemanha, ferramentas eletricas (3 itens) — CANCELADO
    9: [
        ("MAGN-FUR-18V",  "8467.21.00", "Furadeira de impacto a bateria 18V brushless", 40,  185.00, "UN", 1.800, 2.400, 0.0140),
        ("MAGN-PAR-18V",  "8467.22.00", "Parafusadeira a bateria 18V 60Nm",             40,  165.00, "UN", 1.500, 2.100, 0.0125),
        ("MAGN-BAT-18V",  "8507.60.00", "Bateria 18V 5.0Ah ion litio",                  80,  68.50,  "UN",  0.620, 0.780, 0.0028),
    ],
    # Pedido 10 — Exportacao cafe especial (6 itens)
    10: [
        ("BR-CAF-ARAB-450", "0901.21.00", "Cafe arabica torrado moido 450g especialidade",    8000, 4.85, "PCT", 0.480, 0.520, 0.0008),
        ("BR-CAF-ARAB-250", "0901.21.00", "Cafe arabica torrado moido 250g especialidade",    6000, 2.95, "PCT", 0.270, 0.295, 0.0005),
        ("BR-CAF-ROBU-500", "0901.22.00", "Cafe robusta torrado em graos 500g",               5000, 3.10, "PCT", 0.530, 0.570, 0.0009),
        ("BR-CAF-DESC-200", "0901.30.00", "Cafe descafeinado torrado moido 200g",             3000, 5.40, "PCT", 0.220, 0.245, 0.0004),
        ("BR-CAF-BLEND-1KG","0901.21.00", "Cafe blend especial 1kg gourmet",                  2000, 11.20, "PCT", 1.030, 1.080, 0.0017),
        ("BR-CAF-CAPS-60",  "2101.11.10", "Capsulas espresso 60 unidades",                    4000, 14.80, "CX", 0.420, 0.480, 0.0035),
    ],
}

EMPRESA_NOMES = {
    "FXC_TECH_001":   ("Foxconn Industrial Internet Co Ltd", "92.318.520/0001-90", "Shenzhen", "Guangdong", "CN"),
    "BSCH_DE_002":    ("Robert Bosch GmbH", None, "Stuttgart", "BW", "DE"),
    "SMSG_KR_003":    ("Samsung Electronics Co Ltd", None, "Suwon", "Gyeonggi", "KR"),
    "ACME_US_004":    ("Acme Trading LLC", None, "Miami", "FL", "US"),
    "MTSH_JP_005":    ("Mitsubishi Electric Corporation", None, "Tokyo", "Tokyo", "JP"),
    "BYDA_CN_006":    ("BYD Auto Industry Co Ltd", None, "Shenzhen", "Guangdong", "CN"),
    "EUROIMP_007":    ("EuroImp Trading GmbH", None, "Hamburg", "HH", "DE"),
    "XIAO_CN_008":    ("Xiaomi Communications Co Ltd", None, "Beijing", "Beijing", "CN"),
    "MAGN_DE_009":    ("Magnusson Tools GmbH", None, "Munich", "BY", "DE"),
    "TARG_US_010":    ("Target Corp Importing Division", None, "Minneapolis", "MN", "US"),
    "GVT_BR_IMP_001": ("Gravity Comercio Importacao Ltda", "47.215.836/0001-15", "Sao Paulo", "SP", "BR"),
    "GVT_BR_EXP_001": ("Gravity Comercio Exportacao Ltda", "47.215.836/0002-04", "Sao Paulo", "SP", "BR"),
}

CONDICAO_PAGAMENTO = ["30 days T/T", "60 days L/C", "Cash in advance 100%", "30/60/90 days T/T", "L/C at sight", "Cash on delivery", "TT 30%/70% before shipment", "Open account 60 days"]

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def cuid_pedido(idx: int) -> str:
    """Gera CUID-like determinístico para pedido (formato pedi_id_xxxxxxx-26)"""
    return f"pedi_id_{str(1000000 + idx).rjust(7, '0')}-26"

def cuid_item(idx_pedido: int, idx_item: int) -> str:
    return f"pite_id_{str(idx_pedido * 100 + idx_item).rjust(7, '0')}-26"

def cuid_status(idx: int) -> str:
    return f"stps_id_{str(2000000 + idx).rjust(7, '0')}-26"

def cuid_emp(suid: str, idx: int) -> str:
    return f"empr_id_{str(3000000 + idx).rjust(7, '0')}-26"

def cuid_contrato_cambio(idx: int) -> str:
    return f"cccm_id_{str(4000000 + idx).rjust(7, '0')}-26"

# ─────────────────────────────────────────────────────────────────────────────
# CRIA A WORKBOOK
# ─────────────────────────────────────────────────────────────────────────────

wb = Workbook()

# Estilo de header
header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
header_fill = PatternFill('solid', start_color='2D3748')  # cinza escuro
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

cell_font = Font(name='Arial', size=10)
cell_align_left = Alignment(horizontal='left', vertical='center')
cell_align_right = Alignment(horizontal='right', vertical='center')

# ─────────────────────────────────────────────────────────────────────────────
# ABA 1 — PEDIDO
# ─────────────────────────────────────────────────────────────────────────────

ws_pedido = wb.active
ws_pedido.title = "Pedido"

PEDIDO_COLS = [
    "id_pedido", "id_organizacao", "id_workspace", "tipo_operacao_pedido", "numero_pedido",
    "status_pedido", "id_status_pedido", "id_importacao_exportador_pedido",
    "id_exportacao_importador_pedido", "id_fabricante_pedido", "incoterm_pedido", "moeda_pedido",
    "valor_total_pedido", "casas_decimais_valor_pedido", "quantidade_total_pedido",
    "casas_decimais_quantidade_pedido", "unidade_comercializada_pedido", "condicao_pagamento_pedido",
    "numero_proforma_pedido", "numero_invoice_pedido", "referencia_importador_pedido",
    "referencia_exportador_pedido", "referencia_fabricante_pedido", "valor_total_cambio_pedido",
    "moeda_cambio_pedido", "taxa_cambio_estimada_pedido", "contrato_cambio_id_pedido",
    "data_emissao_pedido", "detalhes_operacionais_pedido", "dados_extras_importacao_pedido",
    "ids_origem_consolidacao_pedido", "cnpj_importador_pedido", "data_consolidacao_pedido",
    "data_exclusao_pedido", "peso_liquido_total_pedido", "peso_bruto_total_pedido",
    "cubagem_total_pedido", "casas_decimais_peso_pedido", "casas_decimais_cubagem_pedido",
    "data_criacao_pedido", "data_atualizacao_pedido", "cobertura_cambial_pedido",
    "quantidade_volumes_pedido", "data_documento_pedido", "data_documento_proforma_pedido",
    "data_documento_invoice_pedido", "data_prevista_pedido_pronto", "data_confirmada_pedido_pronto",
    "data_meta_pedido_pronto", "data_prevista_inspecao_pedido", "data_confirmada_inspecao_pedido",
    "data_meta_inspecao_pedido", "data_prevista_coleta_pedido", "data_confirmada_coleta_pedido",
    "data_meta_coleta_pedido", "data_previsao_recebimento_rascunho_pedido",
    "data_confirmacao_recebimento_rascunho_pedido", "data_meta_recebimento_rascunho_pedido",
    "data_previsao_aprovacao_rascunho_pedido", "data_confirmacao_aprovacao_rascunho_pedido",
    "data_meta_aprovacao_rascunho_pedido", "data_previsao_recebimento_rascunho_proforma_pedido",
    "data_confirmacao_recebimento_rascunho_proforma_pedido", "data_meta_recebimento_rascunho_proforma_pedido",
    "data_previsao_aprovacao_rascunho_proforma_pedido", "data_confirmacao_aprovacao_rascunho_proforma_pedido",
    "data_meta_aprovacao_rascunho_proforma_pedido", "data_previsao_envio_original_proforma_pedido",
    "data_confirmacao_envio_original_proforma_pedido", "data_meta_envio_original_proforma_pedido",
    "data_previsao_recebimento_original_proforma_pedido",
    "data_confirmacao_recebimento_original_proforma_pedido",
    "data_meta_recebimento_original_proforma_pedido", "data_previsao_recebimento_rascunho_invoice_pedido",
    "data_confirmacao_recebimento_rascunho_invoice_pedido", "data_meta_recebimento_rascunho_invoice_pedido",
    "data_previsao_aprovacao_rascunho_invoice_pedido", "data_confirmacao_aprovacao_rascunho_invoice_pedido",
    "data_meta_aprovacao_rascunho_invoice_pedido", "data_previsao_envio_original_invoice_pedido",
    "data_confirmacao_envio_original_invoice_pedido", "data_meta_envio_original_invoice_pedido",
    "data_previsao_recebimento_original_invoice_pedido",
    "data_confirmacao_recebimento_original_invoice_pedido",
    "data_meta_recebimento_original_invoice_pedido",
]

# Header
for col_idx, col_name in enumerate(PEDIDO_COLS, start=1):
    cell = ws_pedido.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align

# Linhas de pedido
for row_idx, ped in enumerate(PEDIDOS_DATA, start=2):
    idx, tipo_op, status, incoterm, moeda, taxa_camb, exp_suid, imp_suid, fab_suid, pais, cobertura, num_volumes = ped
    itens = ITENS_DATA[idx]

    qtd_total = sum(it[3] for it in itens)        # soma quantidade_inicial_item
    valor_total = round(sum(it[3] * it[4] for it in itens), 2)
    peso_liq_total = round(sum(it[3] * it[6] for it in itens), 3)
    peso_bru_total = round(sum(it[3] * it[7] for it in itens), 3)
    cubagem_total = round(sum(it[3] * it[8] for it in itens), 4)

    valor_cambio = round(valor_total * taxa_camb, 2)
    cnpj_imp = EMPRESA_NOMES[imp_suid][1] if EMPRESA_NOMES[imp_suid][1] else "47.215.836/0001-15"
    unidade_comum = itens[0][5]  # primeira unidade

    # Datas em sequência lógica baseadas em data_emissao
    base_dt = datetime(2026, 1, 5) + timedelta(days=(idx - 1) * 12)

    # Função helper para datas opcionais (None se status=cancelado em alguns campos finais)
    def fmt_dt(d):
        return d.strftime("%Y-%m-%d") if d else None

    # Datas previstas / confirmadas / meta — sequência de etapas
    d_emissao = base_dt
    d_doc_pedido = base_dt + timedelta(days=2)
    d_doc_proforma = base_dt + timedelta(days=4)
    d_doc_invoice = base_dt + timedelta(days=18)

    d_prev_pronto = base_dt + timedelta(days=20)
    d_conf_pronto = base_dt + timedelta(days=22) if status not in ("rascunho", "aberto", "cancelado") else None
    d_meta_pronto = base_dt + timedelta(days=18)

    d_prev_insp = base_dt + timedelta(days=22)
    d_conf_insp = base_dt + timedelta(days=24) if status in ("em_andamento", "aprovado", "transferencia", "consolidado") else None
    d_meta_insp = base_dt + timedelta(days=20)

    d_prev_col = base_dt + timedelta(days=26)
    d_conf_col = base_dt + timedelta(days=28) if status in ("aprovado", "transferencia", "consolidado") else None
    d_meta_col = base_dt + timedelta(days=24)

    # Draft pedido
    d_prev_rec_dr_ped = base_dt + timedelta(days=3)
    d_conf_rec_dr_ped = base_dt + timedelta(days=4) if status not in ("rascunho",) else None
    d_meta_rec_dr_ped = base_dt + timedelta(days=2)

    d_prev_apv_dr_ped = base_dt + timedelta(days=5)
    d_conf_apv_dr_ped = base_dt + timedelta(days=6) if status not in ("rascunho",) else None
    d_meta_apv_dr_ped = base_dt + timedelta(days=4)

    # Draft proforma
    d_prev_rec_dr_prof = base_dt + timedelta(days=7)
    d_conf_rec_dr_prof = base_dt + timedelta(days=8) if status not in ("rascunho", "aberto") else None
    d_meta_rec_dr_prof = base_dt + timedelta(days=6)

    d_prev_apv_dr_prof = base_dt + timedelta(days=9)
    d_conf_apv_dr_prof = base_dt + timedelta(days=10) if status not in ("rascunho", "aberto") else None
    d_meta_apv_dr_prof = base_dt + timedelta(days=8)

    d_prev_env_or_prof = base_dt + timedelta(days=11)
    d_conf_env_or_prof = base_dt + timedelta(days=12) if status not in ("rascunho", "aberto") else None
    d_meta_env_or_prof = base_dt + timedelta(days=10)

    d_prev_rec_or_prof = base_dt + timedelta(days=14)
    d_conf_rec_or_prof = base_dt + timedelta(days=15) if status in ("em_andamento", "aprovado", "transferencia", "consolidado") else None
    d_meta_rec_or_prof = base_dt + timedelta(days=12)

    # Draft invoice
    d_prev_rec_dr_inv = base_dt + timedelta(days=18)
    d_conf_rec_dr_inv = base_dt + timedelta(days=19) if status in ("em_andamento", "aprovado", "transferencia", "consolidado") else None
    d_meta_rec_dr_inv = base_dt + timedelta(days=16)

    d_prev_apv_dr_inv = base_dt + timedelta(days=20)
    d_conf_apv_dr_inv = base_dt + timedelta(days=21) if status in ("aprovado", "transferencia", "consolidado") else None
    d_meta_apv_dr_inv = base_dt + timedelta(days=18)

    d_prev_env_or_inv = base_dt + timedelta(days=22)
    d_conf_env_or_inv = base_dt + timedelta(days=23) if status in ("aprovado", "transferencia", "consolidado") else None
    d_meta_env_or_inv = base_dt + timedelta(days=20)

    d_prev_rec_or_inv = base_dt + timedelta(days=25)
    d_conf_rec_or_inv = base_dt + timedelta(days=26) if status in ("transferencia", "consolidado") else None
    d_meta_rec_or_inv = base_dt + timedelta(days=22)

    d_consolidacao = base_dt + timedelta(days=35) if status == "consolidado" else None
    d_exclusao = None  # nenhum excluído neste mock

    # Compor número
    ano = base_dt.year
    prefixo = "IMP" if tipo_op == "importacao" else "EXP"
    numero_pedido = f"{prefixo}-{ano}-{str(idx).zfill(4)}"
    numero_proforma = f"PI-{prefixo}-{ano}-{str(idx).zfill(4)}"
    numero_invoice = f"CI-{prefixo}-{ano}-{str(idx).zfill(4)}"

    detalhes_op = (
        f'{{"nome_exportador":"{EMPRESA_NOMES[exp_suid][0]}",'
        f'"nome_importador":"{EMPRESA_NOMES[imp_suid][0]}",'
        f'"nome_fabricante":"{EMPRESA_NOMES[fab_suid][0]}",'
        f'"pais_origem":"{pais}","via":"maritima"}}'
    )
    dados_extras = '{"observacao":"Operacao mock para testes DDD","gerado_por":"script"}'

    row_data = {
        "id_pedido": cuid_pedido(idx),
        "id_organizacao": ID_ORG,
        "id_workspace": ID_WS,
        "tipo_operacao_pedido": tipo_op,
        "numero_pedido": numero_pedido,
        "status_pedido": status,
        "id_status_pedido": cuid_status(idx),
        "id_importacao_exportador_pedido": cuid_emp(exp_suid, idx) if tipo_op == "importacao" else None,
        "id_exportacao_importador_pedido": cuid_emp(imp_suid, idx) if tipo_op == "exportacao" else None,
        "id_fabricante_pedido": cuid_emp(fab_suid, idx),
        "incoterm_pedido": incoterm,
        "moeda_pedido": moeda,
        "valor_total_pedido": valor_total,
        "casas_decimais_valor_pedido": 2,
        "quantidade_total_pedido": qtd_total,
        "casas_decimais_quantidade_pedido": 0 if unidade_comum in ("PCS", "PAR", "UN", "CX", "PCT") else 3,
        "unidade_comercializada_pedido": unidade_comum,
        "condicao_pagamento_pedido": CONDICAO_PAGAMENTO[(idx - 1) % len(CONDICAO_PAGAMENTO)],
        "numero_proforma_pedido": numero_proforma,
        "numero_invoice_pedido": numero_invoice,
        "referencia_importador_pedido": f"REF-IMP-{idx:04d}",
        "referencia_exportador_pedido": f"REF-EXP-{idx:04d}",
        "referencia_fabricante_pedido": f"REF-FAB-{idx:04d}",
        "valor_total_cambio_pedido": valor_cambio,
        "moeda_cambio_pedido": "BRL",
        "taxa_cambio_estimada_pedido": taxa_camb,
        "contrato_cambio_id_pedido": cuid_contrato_cambio(idx) if cobertura == "com_cobertura" else None,
        "data_emissao_pedido": fmt_dt(d_emissao),
        "detalhes_operacionais_pedido": detalhes_op,
        "dados_extras_importacao_pedido": dados_extras,
        "ids_origem_consolidacao_pedido": "[]",
        "cnpj_importador_pedido": cnpj_imp,
        "data_consolidacao_pedido": fmt_dt(d_consolidacao),
        "data_exclusao_pedido": fmt_dt(d_exclusao),
        "peso_liquido_total_pedido": peso_liq_total,
        "peso_bruto_total_pedido": peso_bru_total,
        "cubagem_total_pedido": cubagem_total,
        "casas_decimais_peso_pedido": 3,
        "casas_decimais_cubagem_pedido": 4,
        "data_criacao_pedido": fmt_dt(d_emissao),
        "data_atualizacao_pedido": fmt_dt(base_dt + timedelta(days=10)),
        "cobertura_cambial_pedido": cobertura,
        "quantidade_volumes_pedido": num_volumes,
        "data_documento_pedido": fmt_dt(d_doc_pedido),
        "data_documento_proforma_pedido": fmt_dt(d_doc_proforma),
        "data_documento_invoice_pedido": fmt_dt(d_doc_invoice),
        "data_prevista_pedido_pronto": fmt_dt(d_prev_pronto),
        "data_confirmada_pedido_pronto": fmt_dt(d_conf_pronto),
        "data_meta_pedido_pronto": fmt_dt(d_meta_pronto),
        "data_prevista_inspecao_pedido": fmt_dt(d_prev_insp),
        "data_confirmada_inspecao_pedido": fmt_dt(d_conf_insp),
        "data_meta_inspecao_pedido": fmt_dt(d_meta_insp),
        "data_prevista_coleta_pedido": fmt_dt(d_prev_col),
        "data_confirmada_coleta_pedido": fmt_dt(d_conf_col),
        "data_meta_coleta_pedido": fmt_dt(d_meta_col),
        "data_previsao_recebimento_rascunho_pedido": fmt_dt(d_prev_rec_dr_ped),
        "data_confirmacao_recebimento_rascunho_pedido": fmt_dt(d_conf_rec_dr_ped),
        "data_meta_recebimento_rascunho_pedido": fmt_dt(d_meta_rec_dr_ped),
        "data_previsao_aprovacao_rascunho_pedido": fmt_dt(d_prev_apv_dr_ped),
        "data_confirmacao_aprovacao_rascunho_pedido": fmt_dt(d_conf_apv_dr_ped),
        "data_meta_aprovacao_rascunho_pedido": fmt_dt(d_meta_apv_dr_ped),
        "data_previsao_recebimento_rascunho_proforma_pedido": fmt_dt(d_prev_rec_dr_prof),
        "data_confirmacao_recebimento_rascunho_proforma_pedido": fmt_dt(d_conf_rec_dr_prof),
        "data_meta_recebimento_rascunho_proforma_pedido": fmt_dt(d_meta_rec_dr_prof),
        "data_previsao_aprovacao_rascunho_proforma_pedido": fmt_dt(d_prev_apv_dr_prof),
        "data_confirmacao_aprovacao_rascunho_proforma_pedido": fmt_dt(d_conf_apv_dr_prof),
        "data_meta_aprovacao_rascunho_proforma_pedido": fmt_dt(d_meta_apv_dr_prof),
        "data_previsao_envio_original_proforma_pedido": fmt_dt(d_prev_env_or_prof),
        "data_confirmacao_envio_original_proforma_pedido": fmt_dt(d_conf_env_or_prof),
        "data_meta_envio_original_proforma_pedido": fmt_dt(d_meta_env_or_prof),
        "data_previsao_recebimento_original_proforma_pedido": fmt_dt(d_prev_rec_or_prof),
        "data_confirmacao_recebimento_original_proforma_pedido": fmt_dt(d_conf_rec_or_prof),
        "data_meta_recebimento_original_proforma_pedido": fmt_dt(d_meta_rec_or_prof),
        "data_previsao_recebimento_rascunho_invoice_pedido": fmt_dt(d_prev_rec_dr_inv),
        "data_confirmacao_recebimento_rascunho_invoice_pedido": fmt_dt(d_conf_rec_dr_inv),
        "data_meta_recebimento_rascunho_invoice_pedido": fmt_dt(d_meta_rec_dr_inv),
        "data_previsao_aprovacao_rascunho_invoice_pedido": fmt_dt(d_prev_apv_dr_inv),
        "data_confirmacao_aprovacao_rascunho_invoice_pedido": fmt_dt(d_conf_apv_dr_inv),
        "data_meta_aprovacao_rascunho_invoice_pedido": fmt_dt(d_meta_apv_dr_inv),
        "data_previsao_envio_original_invoice_pedido": fmt_dt(d_prev_env_or_inv),
        "data_confirmacao_envio_original_invoice_pedido": fmt_dt(d_conf_env_or_inv),
        "data_meta_envio_original_invoice_pedido": fmt_dt(d_meta_env_or_inv),
        "data_previsao_recebimento_original_invoice_pedido": fmt_dt(d_prev_rec_or_inv),
        "data_confirmacao_recebimento_original_invoice_pedido": fmt_dt(d_conf_rec_or_inv),
        "data_meta_recebimento_original_invoice_pedido": fmt_dt(d_meta_rec_or_inv),
    }

    for col_idx, col_name in enumerate(PEDIDO_COLS, start=1):
        v = row_data.get(col_name)
        cell = ws_pedido.cell(row=row_idx, column=col_idx, value=v)
        cell.font = cell_font

# Auto-width básico
for col_idx in range(1, len(PEDIDO_COLS) + 1):
    ws_pedido.column_dimensions[get_column_letter(col_idx)].width = 22

ws_pedido.freeze_panes = "B2"

# ─────────────────────────────────────────────────────────────────────────────
# ABA 2 — ItensPedido
# ─────────────────────────────────────────────────────────────────────────────

ws_item = wb.create_sheet("ItensPedido")

ITEM_COLS = [
    "id_item", "id_organizacao", "id_workspace", "id_pedido", "sequencia_item_pedido",
    "part_number_item", "ncm_item", "descricao_item", "unidade_comercializada_item",
    "quantidade_inicial_item", "quantidade_atual_item", "quantidade_pronta_item",
    "quantidade_transferida_item", "quantidade_cancelada_item", "casas_decimais_quantidade_item",
    "moeda_item", "valor_total_item", "valor_por_unidade_item", "casas_decimais_valor_item",
    "cobertura_cambial_item", "nome_exportador_item", "nome_importador_item",
    "nome_fabricante_item", "referencia_importador_item", "referencia_exportador_item",
    "referencia_fabricante_item", "incoterm_item", "condicao_pagamento_item", "data_emissao_item",
    "peso_liquido_unitario_item", "peso_bruto_unitario_item", "cubagem_unitaria_item",
    "casas_decimais_peso_item", "casas_decimais_cubagem_item", "dados_extras_importacao_item",
    "data_criacao_item", "data_atualizacao_item", "data_consolidacao_item", "data_exclusao_item",
    "data_prevista_item_pronto", "data_confirmada_item_pronto", "data_meta_item_pronto",
    "data_prevista_inspecao_item", "data_confirmada_inspecao_item", "data_meta_inspecao_item",
    "data_prevista_coleta_item", "data_confirmada_coleta_item", "data_meta_coleta_item",
]

for col_idx, col_name in enumerate(ITEM_COLS, start=1):
    cell = ws_item.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align

row_idx = 2
for ped in PEDIDOS_DATA:
    idx, tipo_op, status, incoterm, moeda, taxa_camb, exp_suid, imp_suid, fab_suid, pais, cobertura, num_volumes = ped
    base_dt = datetime(2026, 1, 5) + timedelta(days=(idx - 1) * 12)
    itens = ITENS_DATA[idx]

    nome_exp = EMPRESA_NOMES[exp_suid][0]
    nome_imp = EMPRESA_NOMES[imp_suid][0]
    nome_fab = EMPRESA_NOMES[fab_suid][0]

    for seq, (pn, ncm, desc, qtd, valor_unit, unid, peso_liq, peso_bru, cub) in enumerate(itens, start=1):
        valor_total_item = round(qtd * valor_unit, 2)

        # Datas item — alinhadas com pedido
        d_emissao = base_dt
        d_prev_pronto = base_dt + timedelta(days=20)
        d_conf_pronto = base_dt + timedelta(days=22) if status not in ("rascunho", "aberto", "cancelado") else None
        d_meta_pronto = base_dt + timedelta(days=18)
        d_prev_insp = base_dt + timedelta(days=22)
        d_conf_insp = base_dt + timedelta(days=24) if status in ("em_andamento", "aprovado", "transferencia", "consolidado") else None
        d_meta_insp = base_dt + timedelta(days=20)
        d_prev_col = base_dt + timedelta(days=26)
        d_conf_col = base_dt + timedelta(days=28) if status in ("aprovado", "transferencia", "consolidado") else None
        d_meta_col = base_dt + timedelta(days=24)
        d_consolidacao = base_dt + timedelta(days=35) if status == "consolidado" else None

        def fmt_dt(d):
            return d.strftime("%Y-%m-%d") if d else None

        item_row = {
            "id_item": cuid_item(idx, seq),
            "id_organizacao": ID_ORG,
            "id_workspace": ID_WS,
            "id_pedido": cuid_pedido(idx),
            "sequencia_item_pedido": seq,
            "part_number_item": pn,
            "ncm_item": ncm,
            "descricao_item": desc,
            "unidade_comercializada_item": unid,
            "quantidade_inicial_item": qtd,
            "quantidade_atual_item": qtd,
            "quantidade_pronta_item": 0,
            "quantidade_transferida_item": 0,
            "quantidade_cancelada_item": 0,
            "casas_decimais_quantidade_item": 0 if unid in ("PCS", "PAR", "UN", "CX", "PCT") else 3,
            "moeda_item": moeda,
            "valor_total_item": valor_total_item,
            "valor_por_unidade_item": valor_unit,
            "casas_decimais_valor_item": 2,
            "cobertura_cambial_item": cobertura,
            "nome_exportador_item": nome_exp,
            "nome_importador_item": nome_imp,
            "nome_fabricante_item": nome_fab,
            "referencia_importador_item": f"REF-IMP-{idx:04d}-{seq:02d}",
            "referencia_exportador_item": f"REF-EXP-{idx:04d}-{seq:02d}",
            "referencia_fabricante_item": f"REF-FAB-{idx:04d}-{seq:02d}",
            "incoterm_item": incoterm,
            "condicao_pagamento_item": CONDICAO_PAGAMENTO[(idx - 1) % len(CONDICAO_PAGAMENTO)],
            "data_emissao_item": fmt_dt(d_emissao),
            "peso_liquido_unitario_item": peso_liq,
            "peso_bruto_unitario_item": peso_bru,
            "cubagem_unitaria_item": cub,
            "casas_decimais_peso_item": 3,
            "casas_decimais_cubagem_item": 4,
            "dados_extras_importacao_item": '{"observacao":"Item gerado para testes"}',
            "data_criacao_item": fmt_dt(d_emissao),
            "data_atualizacao_item": fmt_dt(base_dt + timedelta(days=10)),
            "data_consolidacao_item": fmt_dt(d_consolidacao),
            "data_exclusao_item": None,
            "data_prevista_item_pronto": fmt_dt(d_prev_pronto),
            "data_confirmada_item_pronto": fmt_dt(d_conf_pronto),
            "data_meta_item_pronto": fmt_dt(d_meta_pronto),
            "data_prevista_inspecao_item": fmt_dt(d_prev_insp),
            "data_confirmada_inspecao_item": fmt_dt(d_conf_insp),
            "data_meta_inspecao_item": fmt_dt(d_meta_insp),
            "data_prevista_coleta_item": fmt_dt(d_prev_col),
            "data_confirmada_coleta_item": fmt_dt(d_conf_col),
            "data_meta_coleta_item": fmt_dt(d_meta_col),
        }

        for col_idx, col_name in enumerate(ITEM_COLS, start=1):
            cell = ws_item.cell(row=row_idx, column=col_idx, value=item_row.get(col_name))
            cell.font = cell_font

        row_idx += 1

for col_idx in range(1, len(ITEM_COLS) + 1):
    ws_item.column_dimensions[get_column_letter(col_idx)].width = 22
ws_item.freeze_panes = "B2"

# ─────────────────────────────────────────────────────────────────────────────
# SAVE
# ─────────────────────────────────────────────────────────────────────────────

OUTPUT = "C:/Users/danie/gravity-antigravity/tabela_pedido_populada_10.xlsx"
wb.save(OUTPUT)

# Stats
total_itens = sum(ITENS_POR_PEDIDO)
print(f"OK — gerado {OUTPUT}")
print(f"   - Aba 'Pedido': 10 linhas, {len(PEDIDO_COLS)} colunas")
print(f"   - Aba 'ItensPedido': {total_itens} linhas, {len(ITEM_COLS)} colunas")
